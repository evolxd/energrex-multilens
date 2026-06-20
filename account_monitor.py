"""
ENERGREX — 账户持仓监控模块
============================
数据来源：Firstrade 网站手动导出 CSV → watchdog 自动检测 → 解析入库
数据存储：SQLite  data/energrex.db
监控目录：~/Downloads/  检测 export*.csv（当日文件）
"""

import os, pathlib, datetime, sqlite3, threading, time, logging, shutil, re, json
import urllib.request
import numpy as np
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import pytz

try:
    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler
    _WATCHDOG_OK = True
except ImportError:
    _WATCHDOG_OK = False

try:
    from apscheduler.schedulers.background import BackgroundScheduler
    from apscheduler.triggers.cron import CronTrigger
    _SCHED_OK = True
except ImportError:
    _SCHED_OK = False

# ════════════════════════════════════════════════════════
# 路径 & 配置
# ════════════════════════════════════════════════════════
_ROOT         = pathlib.Path(__file__).parent
_DOWNLOADS    = pathlib.Path.home() / "Downloads"
_FT_DIR       = _ROOT / "data" / "firstrade"
_FT_DIR.mkdir(parents=True, exist_ok=True)
_LATEST_CSV   = _FT_DIR / "latest.csv"

_env = _ROOT / ".env"
if _env.exists():
    for _line in _env.read_text(encoding="utf-8").splitlines():
        _line = _line.strip()
        if _line and not _line.startswith("#") and "=" in _line:
            _k, _v = _line.split("=", 1)
            os.environ.setdefault(_k.strip(), _v.strip())

ACCT_CFG = [
    {"id": "account_1", "label": "账户一"},
    {"id": "account_2", "label": "账户二"},
]

_ET  = pytz.timezone("America/New_York")
_log = logging.getLogger("energrex.account")

_MD_BASE = "https://api.marketdata.app/v1"
_MD_KEY  = os.environ.get("MARKETDATA_API_KEY", "")

# Database boundary lives in account.db; these aliases keep the existing
# account_monitor call sites stable while the monolith is split incrementally.
from account.db import DB_PATH as _DB_PATH
from account.db import SCREENSHOT_DIR as _SS_DIR
from account.db import db as _db
from account.db import init_db as _init_db


_init_db()




def _fetch_qqq_close(date: str) -> float | None:
    """返回 QQQ 在 date（YYYY-MM-DD）当天或之前最近交易日的收盘价，结果缓存入 DB。"""
    conn = _db()
    row = conn.execute(
        "SELECT close_price FROM qqq_daily_price WHERE date=?", (date,)).fetchone()
    conn.close()
    if row:
        return float(row[0])
    try:
        import yfinance as yf
        dt = datetime.datetime.strptime(date, "%Y-%m-%d")
        start = (dt - datetime.timedelta(days=6)).strftime("%Y-%m-%d")
        end   = (dt + datetime.timedelta(days=1)).strftime("%Y-%m-%d")
        hist  = yf.Ticker("QQQ").history(start=start, end=end)
        if hist.empty:
            return None
        hist.index = pd.to_datetime(hist.index).strftime("%Y-%m-%d")
        available = [d for d in hist.index if d <= date]
        if not available:
            return None
        close_date  = available[-1]
        close_price = float(hist.loc[close_date, "Close"])
        conn = _db()
        conn.execute(
            "INSERT OR REPLACE INTO qqq_daily_price (date, close_price) VALUES (?,?)",
            (date, close_price))
        conn.commit(); conn.close()
        return close_price
    except Exception as e:
        _log.warning(f"QQQ fetch failed for {date}: {e}")
        return None


@st.cache_data(ttl=3600)
def _fetch_qqq_range(start_date: str, end_date: str) -> dict:
    """返回 {YYYY-MM-DD: close} for QQQ over [start_date, end_date)。结果缓存1小时。"""
    try:
        import yfinance as yf
        hist = yf.Ticker("QQQ").history(start=start_date, end=end_date, auto_adjust=True)
        if hist.empty:
            return {}
        hist.index = pd.to_datetime(hist.index).tz_localize(None)
        return {d.strftime("%Y-%m-%d"): float(p) for d, p in hist["Close"].items()}
    except Exception:
        return {}


def _get_benchmark_chart_data(acct_id: str):
    """返回 (nav_df, qqq_df, base_date, base_nav)，用于 AI赋能基准图表。"""
    conn = _db()
    nav_df = pd.read_sql_query(
        "SELECT date, nav FROM daily_nav WHERE account_id=? AND date>=? ORDER BY date",
        conn, params=(acct_id, _AI_START))
    conn.close()
    if nav_df.empty:
        return pd.DataFrame(), pd.DataFrame(), None, 0.0

    base_date = nav_df["date"].iloc[0]
    base_nav  = float(nav_df["nav"].iloc[0])

    qqq_rows = []
    for d in nav_df["date"]:
        p = _fetch_qqq_close(d)
        if p:
            qqq_rows.append({"date": d, "close": p})
    qqq_df = pd.DataFrame(qqq_rows)

    nav_df["rel_pct"] = (nav_df["nav"] / base_nav - 1) * 100
    if not qqq_df.empty:
        base_qqq = float(qqq_df["close"].iloc[0])
        qqq_df["rel_pct"] = (qqq_df["close"] / base_qqq - 1) * 100

    return nav_df, qqq_df, base_date, base_nav


_HIST_START = "2025-06-13"   # 历史视图起点（过去约1年）
_AI_START   = "2026-06-16"   # AI赋能对比起点（周一）

def _effective_base_capital(acct_id: str) -> float:
    """资金基数 = 当前账户净值 + 历史出金合计（去重后）。"""
    conn = _db()
    nav_row = conn.execute(
        "SELECT nav FROM daily_nav WHERE account_id=? ORDER BY date DESC LIMIT 1",
        (acct_id,)).fetchone()
    out_row = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM cash_flow "
        "WHERE account_id=? AND type='出金'",
        (acct_id,)).fetchone()
    conn.close()
    nav = float(nav_row[0]) if nav_row else 0.0
    out = float(out_row[0]) if out_row else 0.0
    return nav + out if (nav + out) > 0 else 34330.88


def _get_history_vs_qqq(acct_id: str, base_capital: float | None = None):
    """
    返回 (acct_df, qqq_df, base_capital) 用于历史表现 vs QQQ 图表。
    acct_df: date / cumulative_pnl / rel_pct（按每日平仓P&L累加）
    qqq_df:  date / close / rel_pct（从 _HIST_START 归一化）
    base_capital：当前净值 + 历史出金，自动计算（可手动覆盖）。
    """
    if base_capital is None:
        base_capital = _effective_base_capital(acct_id)
    today = datetime.datetime.now(_ET).strftime("%Y-%m-%d")

    # ── 账户累计P&L（按平仓日聚合）──
    conn = _db()
    trades = pd.read_sql_query(
        "SELECT close_date AS date, SUM(realized_pnl) AS daily_pnl "
        "FROM option_realized_trades "
        "WHERE account_id=? AND close_date>=? "
        "GROUP BY close_date ORDER BY close_date",
        conn, params=(acct_id, _HIST_START))
    conn.close()

    if trades.empty:
        return pd.DataFrame(), pd.DataFrame(), base_capital

    trades["cumulative_pnl"] = trades["daily_pnl"].cumsum()
    trades["rel_pct"]        = trades["cumulative_pnl"] / base_capital * 100

    # ── QQQ 批量拉取 ──
    qqq_df = pd.DataFrame()
    try:
        import yfinance as yf
        end = (datetime.datetime.strptime(today, "%Y-%m-%d")
               + datetime.timedelta(days=1)).strftime("%Y-%m-%d")
        raw = yf.Ticker("QQQ").history(start=_HIST_START, end=end)
        if not raw.empty:
            raw.index = pd.to_datetime(raw.index).strftime("%Y-%m-%d")
            qqq_df = raw.reset_index()[["Date", "Close"]].rename(
                columns={"Date": "date", "Close": "close"})
            base_qqq       = float(qqq_df["close"].iloc[0])
            qqq_df["rel_pct"] = (qqq_df["close"] / base_qqq - 1) * 100
            # 批量缓存到 DB
            conn = _db()
            for _, row in qqq_df.iterrows():
                conn.execute(
                    "INSERT OR REPLACE INTO qqq_daily_price (date, close_price) VALUES (?,?)",
                    (row["date"], row["close"]))
            conn.commit(); conn.close()
    except Exception as e:
        _log.warning(f"QQQ history fetch failed: {e}")

    return trades, qqq_df, base_capital


def _compute_twr_series(acct_id: str, start_date: str) -> tuple:
    """
    TWR（时间加权收益率）序列，排除出入金影响。
    每日子期间收益率 r_i = (NAV_end - NAV_start - CF_i) / NAV_start
    Returns (df, cashflows_dict)
      df: date / total_equity / twr_pct
      cashflows_dict: {YYYY-MM-DD: amount}  CF<0=出金, CF>0=入金
    """
    conn = _db()
    df = pd.read_sql_query(
        "SELECT DATE(sync_time) AS date, total_equity "
        "FROM account_balance WHERE account_id=? AND DATE(sync_time)>=? "
        "ORDER BY sync_time",
        conn, params=(acct_id, start_date))
    cf_rows = conn.execute(
        "SELECT trade_date, SUM(amount) AS cf FROM transactions "
        "WHERE account_id=? AND type IN ('提款','存款','DEPOSIT','WITHDRAWAL') "
        "AND trade_date>=? GROUP BY trade_date",
        (acct_id, start_date)).fetchall()
    conn.close()

    if df.empty:
        return pd.DataFrame(), {}

    df = (df.groupby("date", as_index=False)["total_equity"].last()
            .sort_values("date").reset_index(drop=True))
    cashflows = {r[0]: float(r[1]) for r in cf_rows}

    nav_arr  = df["total_equity"].tolist()
    date_arr = df["date"].tolist()
    twr_factor  = 1.0
    twr_pct_list = [0.0]
    for i in range(1, len(date_arr)):
        nav_end   = float(nav_arr[i])
        nav_start = float(nav_arr[i - 1])
        cf = cashflows.get(date_arr[i], 0.0)
        r  = (nav_end - nav_start - cf) / nav_start if nav_start > 0 else 0.0
        twr_factor *= (1.0 + r)
        twr_pct_list.append((twr_factor - 1) * 100)
    df["twr_pct"] = twr_pct_list
    return df, cashflows




# ────────────────────────────────────────────────────────
# 期权持仓辅助函数
# ────────────────────────────────────────────────────────
from account.repository import load_balance_history as _load_balance_history
from account.repository import load_latest_balance as _load_latest_balance
from account.repository import load_positions as _load_positions
from account.repository import load_transactions as _load_transactions
from account.repository import record_daily_nav as _record_daily_nav
from account.repository import save_balance as _save_balance
from account.repository import save_positions as _save_positions
from account.repository import save_transactions as _save_transactions


_OCC_RE = re.compile(r'^([A-Z]{1,6})(\d{2})(\d{2})(\d{2})([CP])(\d{8})$')


from account.options_repository import delete_options_position as _delete_options_position
from account.options_repository import derive_open_options as _derive_open_options
from account.options_repository import load_latest_portfolio_greeks_snapshot as _load_latest_portfolio_greeks_snapshot
from account.options_repository import load_options_positions as _load_options_positions
from account.options_repository import load_realized_trades as _load_realized_trades
from account.options_repository import replace_realized_trades_and_fifo_costs as _replace_realized_trades_and_fifo_costs
from account.options_repository import save_portfolio_greeks_snapshot as _save_portfolio_greeks_snapshot
from account.options_repository import save_options_positions as _save_options_positions
from account.options_repository import update_option_market_snapshot as _update_option_market_snapshot
from account.options_repository import update_option_unit_cost as _update_option_unit_cost
from account.options import option_market_value as _option_market_value
from account.options import OCC_RE as _OCC_RE
from account.options import parse_occ as _parse_occ
from account.options import parse_occ_sym as _parse_occ_sym
from account.fifo import calculate_fifo_matches as _calculate_fifo_matches
from account.risk import bs_greeks as _bs_greeks
from account.risk import calculate_option_position_greeks as _calculate_option_position_greeks
from account.risk import delta_drift_trigger as _delta_drift_trigger
from account.risk import summarize_portfolio_greeks as _summarize_portfolio_greeks
from account.risk import vix_spike_trigger as _vix_spike_trigger
from account.marketdata import fetch_option_quote as _fetch_option_quote_md
from account.marketdata import fetch_underlying_prices as _fetch_underlying_prices_md
from account.marketdata import get_atm_iv_batch as _get_atm_iv_batch_md
from account.marketdata import get_spot_prices_batch as _get_spot_prices_batch_md
from account.marketdata import get_vix_snapshot as _get_vix_snapshot_md
from account.importers import detect_csv_type as _detect_csv_type
from account.importers import import_positions_csv as _account_import_positions_csv
from account.importers import import_transactions_csv as _account_import_transactions_csv
from account.importers import parse_date as _parse_date
from account.importers import parse_money as _parse_money
from account.importers import process_csv_file as _account_process_csv_file


def _refresh_options_prices(acct_id: str) -> tuple[int, int]:
    """拉取所有期权现价，更新 DB。返回 (更新数, 删除数)。"""
    df = _load_options_positions(acct_id)
    if df.empty:
        return 0, 0

    today = datetime.date.today()
    updated = 0
    deleted = 0

    for _, row in df.iterrows():
        sym    = row["symbol"]
        expiry = row.get("expiry", "")

        # 删除已到期（到期日 < 今天）
        try:
            if expiry and datetime.date.fromisoformat(expiry) < today:
                _delete_options_position(acct_id, sym)
                deleted += 1
                continue
        except Exception:
            pass

        quote = _md_options_quote(sym)
        if not quote:
            continue

        price = quote.get("mid") or quote.get("last")
        if price is None:
            continue

        qty       = row.get("quantity") or 0
        unit_cost = row.get("unit_cost")

        # Signed market value: short option positions are liabilities.
        market_value = _option_market_value(qty, price)

        # day_pnl：涨跌 × 数量 × 100（多头赚跌亏，空头反之）
        change = quote.get("change") or 0
        day_pnl = change * qty * 100   # qty < 0 for short → change up = loss

        # total_pnl = (current - cost) × qty × 100（需要 unit_cost）
        total_pnl = None
        if unit_cost is not None:
            total_pnl = (price - unit_cost) * qty * 100

        iv    = quote.get("iv")
        delta = quote.get("delta")
        gamma = quote.get("gamma")
        theta = quote.get("theta")
        vega  = quote.get("vega")

        _update_option_market_snapshot(
            acct_id,
            sym,
            price=price,
            market_value=market_value,
            day_pnl=day_pnl,
            total_pnl=total_pnl,
            iv=iv,
            delta=delta,
            gamma=gamma,
            theta=theta,
            vega=vega,
        )
        updated += 1

    return updated, deleted


def _fifo_match_options(acct_id: str) -> dict:
    conn = _db()
    rows = conn.execute(
        "SELECT trade_date, type, symbol, quantity, price, amount "
        "FROM transactions WHERE account_id=? ORDER BY trade_date, id",
        (acct_id,),
    ).fetchall()
    conn.close()

    result = _calculate_fifo_matches(rows)
    _replace_realized_trades_and_fifo_costs(
        acct_id,
        realized=result["realized"],
        fifo_costs=result["fifo_costs"],
    )
    return result["summary"]



def _compute_iv_regime(acct_id: str, min_samples: int = 20) -> dict:
    """
    读取 iv_history + 当前 options_positions.iv，计算每个 symbol 的
    IV Rank / PIV，汇总组合 IV Regime 状态。
    逻辑与 update_iv_regime.py 完全一致（HIGH_PIV=0.85 / EXTREME_PIV=0.95）。
    """
    HIGH_PIV    = 0.85
    EXTREME_PIV = 0.95
    LOW_PIV     = 0.30

    conn = _db()
    hist_rows = conn.execute(
        "SELECT symbol, iv FROM iv_history WHERE account_id=? ORDER BY timestamp",
        (acct_id,)).fetchall()
    cur_rows = conn.execute(
        "SELECT symbol, iv FROM options_positions "
        "WHERE account_id=? AND iv IS NOT NULL",
        (acct_id,)).fetchall()
    conn.close()

    history: dict = {}
    for r in hist_rows:
        history.setdefault(r["symbol"], []).append(float(r["iv"]))

    results = []
    for cur in cur_rows:
        sym = cur["symbol"]
        current_iv = float(cur["iv"])
        sample = list(history.get(sym, []))
        if current_iv not in sample:
            sample.append(current_iv)
        n = len(sample)
        iv_min, iv_max = min(sample), max(sample)
        iv_rank = (current_iv - iv_min) / (iv_max - iv_min) if iv_max > iv_min else None
        piv = sum(1 for v in sample if v <= current_iv) / n if n else None

        if n < min_samples:
            status = "INSUFFICIENT_HISTORY"
        elif piv is not None and piv >= EXTREME_PIV:
            status = "EXTREME_IV"
        elif piv is not None and piv >= HIGH_PIV:
            status = "HIGH_IV"
        elif piv is not None and piv < LOW_PIV:
            status = "LOW_IV"
        else:
            status = "NORMAL"

        results.append({
            "symbol": sym, "iv": current_iv, "n": n,
            "iv_rank": iv_rank, "piv": piv, "status": status,
        })

    if not results:
        return {"status": "NO_DATA", "positions": [], "max_piv": None}

    sufficient = [r for r in results if r["status"] != "INSUFFICIENT_HISTORY"]
    if not sufficient:
        port_status = "INSUFFICIENT_HISTORY"
    elif any(r["status"] == "EXTREME_IV" for r in sufficient):
        port_status = "EXTREME_IV"
    elif any(r["status"] == "HIGH_IV" for r in sufficient):
        port_status = "HIGH_IV"
    elif all(r["piv"] is not None and r["piv"] < LOW_PIV for r in sufficient):
        port_status = "LOW_IV"
    else:
        port_status = "NORMAL"

    max_piv = max((r for r in results if r["piv"] is not None),
                  key=lambda x: x["piv"], default=None)

    return {"status": port_status, "positions": results, "max_piv": max_piv}


# ─────────────────────────────────────────────────────────────────
# Step 3: Risk Snapshot
# ─────────────────────────────────────────────────────────────────
# ── Beta 硬编码兜底值（yfinance 失效时使用）────────────────────────────────
_BETA_BASE = {
    "ARM": 2.95, "AVGO": 2.13, "META": 1.48, "PLTR": 1.81, "VST": 1.52,
    "FCX": 2.14, "MRVL": 2.61, "NVDA": 1.83, "PANW": 1.09, "CRWV": 2.84,
    "QQQ": 1.31, "NOK": 1.42, "CRWD": 1.50, "DDOG": 1.70, "NET": 1.65,
    "AMD": 1.90, "TSLA": 2.30, "MSFT": 1.20, "GOOGL": 1.15, "AMZN": 1.25,
    "SMCI": 2.20, "MU": 1.60, "SNOW": 2.10, "ORCL": 1.10, "ZS": 1.55,
    "SPY": 1.00, "IWM": 1.10, "GLD": 0.05,
}

# ── Beta 缓存（yfinance 每周刷新，存 data/beta_cache.json）────────────────
_BETA_CACHE_PATH  = _ROOT / "data" / "beta_cache.json"
_BETA_CACHE_DAYS  = 8   # 超过8天视为过期，重新拉取


def _load_beta_cache() -> dict:
    """读 beta_cache.json；过期或损坏时返回 {}。"""
    try:
        raw  = json.loads(_BETA_CACHE_PATH.read_text(encoding="utf-8"))
        age  = (datetime.datetime.now() -
                datetime.datetime.fromisoformat(raw["updated_at"])).days
        if age <= _BETA_CACHE_DAYS:
            return {k: float(v) for k, v in raw.get("betas", {}).items()}
    except Exception:
        pass
    return {}


def _save_beta_cache(betas: dict) -> None:
    payload = {
        "updated_at": datetime.datetime.now().isoformat(),
        "betas":      {k: round(v, 4) for k, v in betas.items()},
    }
    _BETA_CACHE_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def _refresh_beta_spy() -> None:
    """从 yfinance 拉取所有已知标的的 5年月度 beta，更新 _BETA_SPY 并写缓存。"""
    import yfinance as yf
    tickers = list(_BETA_BASE.keys())
    new_betas: dict = {}
    failed: list    = []

    for sym in tickers:
        try:
            b = yf.Ticker(sym).info.get("beta")
            if b is not None and 0.01 < float(b) < 15:
                new_betas[sym] = round(float(b), 3)
            else:
                failed.append(sym)
        except Exception:
            failed.append(sym)

    # 用 yfinance 结果覆盖 base，失败的保留 base 值
    merged = {**_BETA_BASE, **new_betas}
    _BETA_SPY.update(merged)
    _save_beta_cache(new_betas)

    changes = {k: (round(_BETA_BASE[k], 3), round(new_betas[k], 3))
               for k in new_betas if abs(new_betas[k] - _BETA_BASE[k]) > 0.05}
    _log.info(
        f"[beta_refresh] 成功={len(new_betas)} 失败={failed or '无'} "
        f"变化较大={changes or '无'}"
    )


# 运行时 _BETA_SPY：启动时合并 base + 缓存（缓存优先）
_BETA_SPY = {**_BETA_BASE, **_load_beta_cache()}

_RISK_LIMITS = {
    "max_leverage":          4.0,
    "max_beta_delta_ratio":  3.5,
    "stress_warning":        0.08,
    "stress_de_risk":        0.12,
    "stress_hard_stop":      0.15,
    "drawdown_freeze":       0.20,
    "drawdown_de_risk":      0.30,
}


def _compute_risk_snapshot(acct_id: str) -> dict:
    """
    计算 Beta 加权 Delta、杠杆、压力测试损失。
    逻辑与 snapshot_risk.py 一致：
      stress = q * mult * (d*ds + 0.5*g*ds²) + q * mult * v * iv_shock
    -10% 下跌: ds=-10%×S, iv_shock=+8 pts
    -20% 下跌: ds=-20%×S, iv_shock=+16 pts
    """
    bal = _load_latest_balance(acct_id)
    equity = float(bal.get("total_equity") or 0)
    if equity <= 0:
        return {"error": "no_equity"}

    conn = _db()

    # ── TWR-based drawdown（排除出入金）──────────────────────────────
    _nav_rows = conn.execute(
        "SELECT DATE(sync_time) AS d, total_equity FROM account_balance "
        "WHERE account_id=? ORDER BY sync_time", (acct_id,)).fetchall()
    _cf_rows_dd = conn.execute(
        "SELECT trade_date, SUM(amount) AS cf FROM transactions "
        "WHERE account_id=? AND type IN ('提款','存款','DEPOSIT','WITHDRAWAL') "
        "GROUP BY trade_date", (acct_id,)).fetchall()
    _nav_by_d = {}
    for r in _nav_rows:
        _nav_by_d[str(r[0])] = float(r[1])
    _cf_map_dd = {r[0]: float(r[1]) for r in _cf_rows_dd}
    _dates_dd = sorted(_nav_by_d)
    drawdown = 0.0
    if len(_dates_dd) >= 2:
        _twr_f = _peak_f = 1.0
        _prev = float(_nav_by_d[_dates_dd[0]])
        for _d in _dates_dd[1:]:
            _nav = float(_nav_by_d[_d])
            _cf  = _cf_map_dd.get(_d, 0.0)
            _r   = (_nav - _prev - _cf) / _prev if _prev > 0 else 0.0
            _twr_f *= (1.0 + _r)
            if _twr_f > _peak_f:
                _peak_f = _twr_f
            _dd = _twr_f / _peak_f - 1
            if _dd < -drawdown:
                drawdown = abs(_dd)
            _prev = _nav

    opts = conn.execute(
        "SELECT symbol, quantity, current_price, market_value, strike, expiry "
        "FROM options_positions WHERE account_id=? AND current_price IS NOT NULL",
        (acct_id,)).fetchall()
    stks = conn.execute(
        "SELECT symbol, quantity, market_value "
        "FROM positions WHERE account_id=? AND position_type='stock'",
        (acct_id,)).fetchall()
    conn.close()

    # Collect underlyings for option stress test
    underlyings = set()
    for o in opts:
        mo = _OCC_RE.match((o["symbol"] or "").upper())
        if mo:
            underlyings.add(mo.group(1))
    und_prices = _fetch_underlying_prices(tuple(sorted(underlyings))) if underlyings else {}
    _iv_map    = _get_atm_iv_batch(tuple(sorted(underlyings))) if underlyings else {}
    _today     = datetime.date.today()

    gross       = 0.0
    delta_notl  = 0.0   # Delta-adjusted notional: Σ|Δ × qty × 100 × S|，spread 两腿自然对冲
    beta_delta  = 0.0
    theta_tot   = 0.0
    vega_tot    = 0.0
    gamma_tot   = 0.0
    stress_10   = 0.0
    stress_20   = 0.0
    nearest_expiry_date = None
    nearest_expiry_sym  = ""

    for s in stks:
        sym = str(s["symbol"] or "").upper()
        q   = float(s["quantity"] or 0)
        mv  = float(s["market_value"] or 0)
        s_price = mv / q if q else 0.0
        b = _BETA_BASE.get(sym, 1.0)   # 用硬编码保守值，不受 yfinance 刷新影响
        gross      += abs(mv)
        delta_notl += abs(mv)           # 股票 delta=1
        beta_delta += q * s_price * 1.0 * b
        ds10 = -0.10 * s_price
        ds20 = -0.20 * s_price
        stress_10 += q * ds10
        stress_20 += q * ds20

    for o in opts:
        sym   = (o["symbol"] or "").upper()
        mo    = _OCC_RE.match(sym)
        und   = mo.group(1) if mo else sym
        b     = _BETA_BASE.get(und, 1.0)   # 用硬编码保守值，不受 yfinance 刷新影响
        q     = float(o["quantity"] or 0)
        mult  = 100.0
        price = float(o["current_price"] or 0)
        mv    = float(o["market_value"] or 0)
        S     = und_prices.get(und, 0.0)

        # ── 始终用 BS 公式计算 Greeks，不依赖 DB 中可能为 NULL 的列 ──
        d = g = th = vg = 0.0
        if mo and S > 0:
            _K = float(o["strike"] or 0) or float(mo.group(6)) / 1000.0
            _opt_type = "call" if mo.group(5) == "C" else "put"
            _iv_entry = _iv_map.get(und)
            _iv = _iv_entry["iv"] if _iv_entry else 0.30
            try:
                _exp_date = datetime.date.fromisoformat(str(o["expiry"]))
            except Exception:
                _exp_date = datetime.date(2000 + int(mo.group(2)), int(mo.group(3)), int(mo.group(4)))
            _dte = max(0, (_exp_date - _today).days)
            if _dte > 0 and _K > 0:
                _gr = _bs_greeks(S, _K, _dte / 365.0, _iv, _opt_type)
                d, g, th, vg = _gr["delta"], _gr["gamma"], _gr["theta"], _gr["vega"]

        gross      += abs(q * mult * S) if S > 0 else abs(q * mult * price)
        # Delta-adjusted notional: 价差两腿 delta 符号相反，自然抵消，不双计
        if S > 0 and abs(d) > 0.001:
            delta_notl += abs(q * mult * d * S)
        else:
            # Delta 未算出（深度 OTM / 到期）：用期权市值代替，避免高估
            delta_notl += abs(mv) if abs(mv) > 0 else abs(q * mult * price)
        theta_tot  += q * mult * th
        vega_tot   += q * mult * vg
        gamma_tot  += abs(q) * g * mult

        # nearest expiry — OCC groups: 1=und 2=YY 3=MM 4=DD
        if mo:
            try:
                _exp = datetime.date(2000 + int(mo.group(2)), int(mo.group(3)), int(mo.group(4)))
                if nearest_expiry_date is None or _exp < nearest_expiry_date:
                    nearest_expiry_date = _exp
                    nearest_expiry_sym  = und
            except ValueError:
                pass

        if S > 0:
            beta_delta += q * mult * d * S * b
            ds10 = -0.10 * S
            ds20 = -0.20 * S
            pnl10 = q * mult * (d * ds10 + 0.5 * g * ds10 * ds10) + q * mult * vg * 8
            pnl20 = q * mult * (d * ds20 + 0.5 * g * ds20 * ds20) + q * mult * vg * 16
            stress_10 += pnl10
            stress_20 += pnl20

    leverage          = gross / equity if equity else None
    leverage_delta    = delta_notl / equity if equity else None   # Delta 口径，价差不双计
    beta_delta_ratio  = beta_delta / equity if equity else None
    stress_10_ratio   = stress_10 / equity if equity else None
    stress_20_ratio   = stress_20 / equity if equity else None

    lim = _RISK_LIMITS
    s10 = abs(stress_10_ratio) if stress_10_ratio else 0
    if s10 >= lim["stress_hard_stop"]:
        risk_status = "RED_HARD_STOP"
    elif s10 >= lim["stress_de_risk"]:
        risk_status = "ORANGE_DE_RISK"
    elif s10 >= lim["stress_warning"]:
        risk_status = "YELLOW_WARNING"
    else:
        risk_status = "GREEN"

    if drawdown >= lim["drawdown_de_risk"]:
        dd_status = "RED_MANDATORY_DE_RISK"
    elif drawdown >= lim["drawdown_freeze"]:
        dd_status = "ORANGE_FREEZE_NEW_RISK"
    else:
        dd_status = "GREEN"

    return {
        "equity":           equity,
        "drawdown":         drawdown,
        "gross_notional":   round(gross, 0),
        "delta_notional":   round(delta_notl, 0),
        "leverage":         round(leverage, 2)      if leverage       is not None else None,
        "leverage_delta":   round(leverage_delta, 2) if leverage_delta is not None else None,
        "beta_delta":       round(beta_delta, 0),
        "beta_delta_ratio": round(beta_delta_ratio, 4) if beta_delta_ratio is not None else None,
        "theta_per_day":       round(theta_tot, 2),
        "vega_per_pt":         round(vega_tot, 2),
        "gamma_total":         round(gamma_tot, 4),
        "stress_10":           round(stress_10, 0),
        "stress_10_ratio":     round(stress_10_ratio, 4)  if stress_10_ratio  is not None else None,
        "stress_20":           round(stress_20, 0),
        "stress_20_ratio":     round(stress_20_ratio, 4)  if stress_20_ratio  is not None else None,
        "nearest_expiry_date": nearest_expiry_date,
        "nearest_expiry_sym":  nearest_expiry_sym,
        "risk_status":         risk_status,
        "drawdown_status":     dd_status,
    }


def _compute_options_cost_ratio(acct_id: str) -> dict:
    """期权总成本 / 最新账户净值，从 Portfolio_Config.json 读取上限。"""
    import json as _json
    conn = _db()
    cost_row = conn.execute(
        "SELECT SUM(ABS(quantity) * unit_cost * 100) FROM options_positions "
        "WHERE account_id=? AND unit_cost IS NOT NULL AND quantity IS NOT NULL",
        (acct_id,)).fetchone()
    nav_row = conn.execute(
        "SELECT total_equity FROM account_balance WHERE account_id=? "
        "ORDER BY sync_time DESC LIMIT 1",
        (acct_id,)).fetchone()
    conn.close()

    total_cost = float(cost_row[0] or 0) if cost_row else 0.0
    nav        = float(nav_row[0] or 0) if nav_row and nav_row[0] else 0.0
    ratio      = total_cost / nav if nav > 0 else None

    cfg_path = pathlib.Path(r"C:\Users\evolx\Documents\ENERGREX期权量化系统\Portfolio_Config.json")
    limit = 0.50
    try:
        with open(cfg_path, "r", encoding="utf-8") as _f:
            _cfg = _json.load(_f)
        limit = float(_cfg.get("options_cost_ratio_limit", 0.50))
    except Exception:
        pass

    if ratio is None:
        status = "gray"
    elif ratio >= limit * 1.5:
        status = "red"
    elif ratio >= limit:
        status = "amber"
    else:
        status = "green"

    return {"total_cost": total_cost, "nav": nav, "ratio": ratio, "limit": limit, "status": status}


# ─────────────────────────────────────────────────────────────────
# Step 4: AI Score
# ─────────────────────────────────────────────────────────────────
@st.cache_data(ttl=3600)
def _load_ai_scores() -> dict:
    """读取 results_validated.csv，返回 {TICKER: score}。"""
    csv_path = _ROOT / "results_validated.csv"
    if not csv_path.exists():
        return {}
    try:
        df = pd.read_csv(csv_path)
        col = "final_score_recalculated" if "final_score_recalculated" in df.columns else "final_score"
        if "ticker" not in df.columns or col not in df.columns:
            return {}
        return {str(t).upper(): float(s) for t, s in
                zip(df["ticker"], df[col]) if pd.notna(s)}
    except Exception:
        return {}


def _score_label(score: float | None) -> str:
    if score is None:
        return ""
    if score >= 65:
        return f"{score:.0f} ⭐"
    if score >= 55:
        return f"{score:.0f} ✅"
    if score >= 45:
        return f"{score:.0f} 🟡"
    if score >= 35:
        return f"{score:.0f} ⚠️"
    return f"{score:.0f} 🔴"


# ─────────────────────────────────────────────────────────────────
# Module A: 事件风险日历
# ─────────────────────────────────────────────────────────────────
_POLY_KEY = os.getenv("POLYGON_API_KEY", "")

# 财报日历：指定标的（ETF QQQ 跳过）
_EARNINGS_TICKERS = ("FCX", "CRWV", "META", "NOK", "NVDA", "PANW", "PLTR", "MRVL")
_ETF_TICKERS      = {"QQQ", "SPY", "IWM", "GLD", "TLT"}


@st.cache_data(ttl=3600)
def _fetch_earnings_dates(tickers_tuple: tuple) -> dict:
    """
    获取各标的下次财报日（1小时缓存）。
    优先 Polygon.io /vX/reference/financials（季度最新申报日 +91 天估算下次），
    回退到 yfinance。ETF 自动跳过。
    """
    result = {t: None for t in tickers_tuple}
    today = datetime.date.today()
    today_str = today.isoformat()

    for t in tickers_tuple:
        if t.upper() in _ETF_TICKERS:
            continue

        # ── Polygon ──────────────────────────────────────────────────
        if _POLY_KEY:
            try:
                url = (f"https://api.polygon.io/vX/reference/financials"
                       f"?ticker={t}&timeframe=quarterly&order=desc&limit=4"
                       f"&apiKey={_POLY_KEY}")
                with urllib.request.urlopen(url, timeout=6) as r:
                    data = json.loads(r.read())
                filing_dates = sorted(
                    [f["filing_date"] for f in data.get("results", [])
                     if f.get("filing_date")],
                    reverse=True)
                if filing_dates:
                    last = datetime.date.fromisoformat(filing_dates[0])
                    nxt  = last + datetime.timedelta(days=91)
                    if nxt.isoformat() <= today_str:
                        nxt = nxt + datetime.timedelta(days=91)
                    result[t] = nxt.isoformat()
                    continue
            except Exception as _pe:
                _log.debug(f"Polygon earnings {t}: {_pe}")

        # ── yfinance 回退 ────────────────────────────────────────────
        try:
            import yfinance as yf
            ticker = yf.Ticker(t)
            try:
                eds = ticker.get_earnings_dates(limit=8)
                if eds is not None and not eds.empty:
                    tz     = eds.index.tz
                    now_ts = pd.Timestamp.now(tz=tz) if tz else pd.Timestamp.now()
                    future = eds[eds.index > now_ts]
                    if not future.empty:
                        result[t] = future.index.min().strftime('%Y-%m-%d')
                        continue
            except Exception:
                pass
            try:
                cal = ticker.calendar
                if isinstance(cal, dict):
                    ed = cal.get('Earnings Date') or cal.get('earningsDate') or []
                    if isinstance(ed, (list, tuple)) and ed:
                        result[t] = str(ed[0])[:10]
                    elif isinstance(ed, str) and ed:
                        result[t] = ed[:10]
            except Exception:
                pass
        except Exception:
            pass

    return result


def _get_event_calendar(acct_id: str, window_days: int = 30) -> list[dict]:
    """
    合并期权到期事件（从 DB）+ 财报日（yfinance），
    返回未来 window_days 天内的事件列表，按 days_left 升序。
    风险级别: CRITICAL(DTE≤14) / HIGH(DTE≤21 或 财报≤7天) / MEDIUM
    """
    today = datetime.date.today()
    DTE_ACTION = 14
    DTE_REVIEW = 21
    events: list[dict] = []

    conn = _db()
    opt_rows = conn.execute(
        "SELECT symbol, expiry FROM options_positions WHERE account_id=?",
        (acct_id,)).fetchall()
    conn.close()

    underlyings: set = set()
    by_exp: dict = {}
    for r in opt_rows:
        sym = (r["symbol"] or "").upper()
        mo = _OCC_RE.match(sym)
        if mo:
            und = mo.group(1)
            underlyings.add(und)
            if r["expiry"]:
                by_exp[(und, r["expiry"])] = by_exp.get((und, r["expiry"]), 0) + 1

    # Option expiry events
    for (und, exp), count in sorted(by_exp.items()):
        try:
            exp_date = datetime.date.fromisoformat(exp)
        except ValueError:
            continue
        dl = (exp_date - today).days
        if dl < 0 or dl > window_days:
            continue
        level = ("CRITICAL" if dl <= DTE_ACTION else
                 "HIGH"     if dl <= DTE_REVIEW else "MEDIUM")
        events.append({
            "date": exp, "days_left": dl, "event_type": "option_expiry",
            "symbol": und, "title": f"{und} {count}腿到期",
            "risk_level": level, "source": "portfolio",
        })

    # Earnings events from yfinance
    if underlyings:
        earnings = _fetch_earnings_dates(tuple(sorted(underlyings)))
        for und, ed_str in earnings.items():
            if not ed_str:
                continue
            try:
                ed = datetime.date.fromisoformat(ed_str)
            except ValueError:
                continue
            dl = (ed - today).days
            if dl < 0 or dl > window_days:
                continue
            level = "HIGH" if dl <= 7 else "MEDIUM"
            events.append({
                "date": ed_str, "days_left": dl, "event_type": "earnings",
                "symbol": und, "title": f"{und} 财报",
                "risk_level": level, "source": "yfinance",
            })

    # Dedupe + sort
    seen: set = set()
    out = []
    for e in sorted(events, key=lambda x: (x["days_left"], x["symbol"])):
        key = (e["date"], e["event_type"], e["symbol"])
        if key not in seen:
            seen.add(key)
            out.append(e)
    return out


# ─────────────────────────────────────────────────────────────────
# Module B: 交易绩效分析
# ─────────────────────────────────────────────────────────────────
def _compute_performance_stats(acct_id: str) -> dict | None:
    """从 option_realized_trades 计算全套绩效指标。数据为空返回 None。"""
    df = _load_realized_trades(acct_id)
    if df.empty:
        return None

    df = df.sort_values("close_date").reset_index(drop=True)
    df["pnl"]  = pd.to_numeric(df["realized_pnl"], errors="coerce").fillna(0)
    df["win"]  = (df["pnl"] > 0).astype(int)
    df["trade_num"]      = range(1, len(df) + 1)
    df["cumulative_pnl"] = df["pnl"].cumsum()
    df["rolling20_wr"]   = df["win"].rolling(20, min_periods=1).mean()
    df["rolling20_avg"]  = df["pnl"].rolling(20, min_periods=1).mean()
    df["rolling50_wr"]   = df["win"].rolling(50, min_periods=1).mean()
    df["rolling50_avg"]  = df["pnl"].rolling(50, min_periods=1).mean()
    df["pnl_peak"]       = df["cumulative_pnl"].cummax()
    df["drawdown"]       = df["cumulative_pnl"] - df["pnl_peak"]

    wins   = df.loc[df["pnl"] > 0, "pnl"]
    losses = df.loc[df["pnl"] < 0, "pnl"]
    avg_win  = float(wins.mean())   if len(wins)   else 0.0
    avg_loss = float(losses.mean()) if len(losses) else 0.0
    pf = abs(avg_win / avg_loss) if avg_loss else None

    by_strat = {}
    for strat, grp in df.groupby("strategy_type"):
        w = (grp["pnl"] > 0).sum()
        by_strat[str(strat)] = {
            "count": len(grp), "win_rate": w / len(grp),
            "avg_pnl": float(grp["pnl"].mean()), "total": float(grp["pnl"].sum()),
        }

    by_und = {}
    for und, grp in df.groupby("underlying"):
        w = (grp["pnl"] > 0).sum()
        by_und[str(und)] = {
            "count": len(grp), "win_rate": w / len(grp),
            "avg_pnl": float(grp["pnl"].mean()), "total": float(grp["pnl"].sum()),
        }

    by_combo = {}
    if "combo_strategy" in df.columns:
        for cs, grp in df.groupby("combo_strategy"):
            w = (grp["pnl"] > 0).sum()
            by_combo[str(cs)] = {
                "count": len(grp), "win_rate": w / len(grp),
                "avg_pnl": float(grp["pnl"].mean()), "total": float(grp["pnl"].sum()),
            }

    # EWMA 加权胜率（λ=0.94，近期权重高）
    _lam = 0.94
    _n   = len(df)
    _w   = np.array([_lam ** (_n - 1 - i) for i in range(_n)], dtype=float)
    _w  /= _w.sum()
    ewma_wr = float((_w * df["win"].values).sum())

    return {
        "df":            df,
        "total_pnl":     float(df["pnl"].sum()),
        "count":         len(df),
        "win_rate":      float(df["win"].mean()),
        "avg_pnl":       float(df["pnl"].mean()),
        "avg_win":       avg_win,
        "avg_loss":      avg_loss,
        "profit_factor": pf,
        "max_loss":      float(df["pnl"].min()),
        "max_gain":      float(df["pnl"].max()),
        "max_drawdown":  float(df["drawdown"].min()),
        "latest_wr20":   float(df["rolling20_wr"].iloc[-1]),
        "ewma_wr":       ewma_wr,
        "by_strat":      by_strat,
        "by_und":        by_und,
        "by_combo":      by_combo,
    }


_RF_RATE = 0.045   # 2026 risk-free rate (3-month T-bill proxy)


def _md_options_quote(symbol: str) -> dict | None:
    return _fetch_option_quote_md(symbol, api_key=_MD_KEY, logger=_log)


@st.cache_data(ttl=300)
def _fetch_underlying_prices(tickers_tuple: tuple) -> dict:
    return _fetch_underlying_prices_md(tickers_tuple, logger=_log)


@st.cache_data(ttl=300)
def _get_spot_prices_batch(tickers: tuple) -> dict[str, float]:
    return _get_spot_prices_batch_md(tickers)


@st.cache_data(ttl=300)
def _get_vix_snapshot() -> dict:
    return _get_vix_snapshot_md()


@st.cache_data(ttl=600)
def _get_atm_iv_batch(tickers: tuple) -> dict[str, dict]:
    return _get_atm_iv_batch_md(tickers, api_key=_MD_KEY)


def _pos_status_light(dte: int, qty: int, bs_delta: float) -> str:
    """每持仓状态灯：🔴需立即行动 / 🟡需关注 / 🟢正常"""
    if dte < 21:
        return "🔴"
    if dte < 45 or (qty < 0 and abs(bs_delta) > 0.70):
        return "🟡"
    return "🟢"


def _compute_portfolio_greeks(acct_id: str) -> dict:
    """
    BS 公式计算全组合 Greeks，聚合组合层面，检测三重触发。
    IV 优先级：MarketData.app ATM IV → yfinance ATM IV → 数据库存储值。
    返回 {rows, totals, triggers, n_contracts, vix, iv_sources}。
    """
    conn = _db()
    pos_rows = conn.execute(
        "SELECT symbol, quantity, unit_cost, current_price, iv, expiry, strike, direction "
        "FROM options_positions WHERE account_id=?", (acct_id,)).fetchall()
    conn.close()
    last_snap = _load_latest_portfolio_greeks_snapshot(acct_id)

    today = datetime.date.today()

    # ── 解析持仓，收集标的列表 ──────────────────────────────────
    # OCC 解析只用于提取 root（标的代码）；expiry/strike/direction 直接读 DB 列
    underlyings: set[str] = set()
    parsed: list[tuple] = []
    for r in pos_rows:
        sym = str(r[0]).strip().upper()
        p = _parse_occ(sym)      # {root, direction, strike, expiry} — used only for root
        if int(r[1] or 0) != 0:
            root = p["root"] if p else re.match(r'^([A-Z]{1,6})', sym).group(1)
            underlyings.add(root)
            parsed.append((r, root))

    # ── 批量拉取现价 + ATM IV ───────────────────────────────────
    spot_prices  = _get_spot_prices_batch(tuple(sorted(underlyings)))
    atm_iv_map   = _get_atm_iv_batch(tuple(sorted(underlyings)))

    rows_out: list[dict] = []

    for r, und in parsed:
        sym      = str(r[0]).strip().upper()
        qty      = int(r[1])
        db_iv    = float(r[4]) if r[4] is not None else None
        # Option type from OCC symbol (C/P), not direction column (which stores long/short)
        _occ_m   = _OCC_RE.match(sym)
        opt_type = ("call" if _occ_m and _occ_m.group(5) == "C" else "put")
        expiry_str = str(r[5])                                # "YYYY-MM-DD"
        K          = float(r[6])                              # strike

        # IV 来源优先级
        iv_entry = atm_iv_map.get(und)
        if iv_entry:
            iv_raw  = iv_entry["iv"]
            iv_src  = iv_entry["src"]       # "md" or "yf"
        elif db_iv is not None:
            iv_raw  = db_iv
            iv_src  = "db"
        else:
            iv_raw  = 0.50
            iv_src  = "default"

        row_out = _calculate_option_position_greeks(
            symbol=sym,
            underlying=und,
            option_type=opt_type,
            quantity=qty,
            strike=K,
            expiry=expiry_str,
            spot_price=spot_prices.get(und),
            current_price=float(r[3] or 0),
            iv=iv_raw,
            iv_source=iv_src,
            today=today,
        )
        if not row_out:
            continue

        row_out["status"] = _pos_status_light(row_out["dte"], qty, row_out["_raw"]["bs_delta"])
        rows_out.append(row_out)

    summary = _summarize_portfolio_greeks(rows_out)
    rows_out = summary["rows"]
    totals = summary["totals"]
    avg_delta = summary["avg_delta"]
    n_contracts = summary["n_contracts"]

    _save_portfolio_greeks_snapshot(
        acct_id,
        total_delta=totals["delta"],
        total_gamma=totals["gamma"],
        total_theta=totals["theta"],
        total_vega=totals["vega"],
        n_contracts=n_contracts,
    )

    # ── 三重触发检测 ───────────────────────────────────────────
    triggers: list[dict] = []

    # 触发一：Delta 漂移 ±0.10（均值/合约）
    if last_snap:
        _drift_trigger = _delta_drift_trigger(last_snap[0], last_snap[1], avg_delta)
        if _drift_trigger:
            triggers.append({
                "level": _drift_trigger["level"],
                "msg":   (f"⚠️ Delta 漂移 {_drift_trigger['drift']:+.3f}/合约（阈值 ±{_drift_trigger['threshold']}）"
                          f"— 均值 Delta 从 {_drift_trigger['previous_avg_delta']:+.3f} 变为 {_drift_trigger['current_avg_delta']:+.3f}"),
            })

    # 触发三A：VIX 单日涨幅 >15%
    vix_snap = _get_vix_snapshot()
    _vix_trigger = _vix_spike_trigger(vix_snap)
    if _vix_trigger:
        triggers.append({
            "level": _vix_trigger["level"],
            "msg":   f"🔴 VIX 单日 {_vix_trigger['change_pct']:+.1f}% — 全组合紧急风险检查！",
        })

    # 触发三B：财报前 7 天
    earns = _fetch_earnings_dates(tuple(sorted(underlyings - _ETF_TICKERS)))
    for und, ed in earns.items():
        if not ed:
            continue
        try:
            dl = (datetime.date.fromisoformat(ed) - today).days
            if 0 <= dl <= 7:
                triggers.append({
                    "level": "HIGH",
                    "msg":   f"🟠 {und} 财报 {dl} 天后（{ed}）— 升级高优先级",
                })
        except Exception:
            pass

    return {
        "rows":     rows_out,
        "totals": totals,
        "n_contracts": n_contracts,
        "triggers":    triggers,
        "vix":         vix_snap,
        "by_und":      summary["by_und"],
        "top_long":    summary["top_long"],
        "top_short":   summary["top_short"],
        "iv_src_counts": summary["iv_src_counts"],
    }


# ─────────────────────────────────────────────────────────────────
# IV 监控 — 从期权链提取ATM IV，本地积累历史计算IVRank
# IVRank endpoint (/v1/options/ivrank/) 当前订阅不可用（返回no_data）
# 改用 /v1/options/chain/ 提取ATM IV + SQLite本地历史
# ─────────────────────────────────────────────────────────────────

def _ensure_iv_atm_history_table() -> None:
    conn = _db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS iv_atm_history (
            id      INTEGER PRIMARY KEY AUTOINCREMENT,
            symbol  TEXT NOT NULL,
            date    TEXT NOT NULL,
            iv_pct  REAL NOT NULL,
            UNIQUE(symbol, date)
        )
    """)
    conn.commit()
    conn.close()


def _save_iv_reading(symbol: str, iv_pct: float) -> None:
    """写入今天的IV读数（已有则忽略）。"""
    _ensure_iv_atm_history_table()
    today = datetime.date.today().isoformat()
    conn = _db()
    conn.execute(
        "INSERT OR IGNORE INTO iv_atm_history (symbol, date, iv_pct) VALUES (?,?,?)",
        (symbol, today, round(iv_pct, 2)))
    conn.commit()
    conn.close()


def _compute_ivrank_from_history(symbol: str, current_iv: float) -> dict:
    """
    从本地历史计算IVRank。
    至少需要30天数据才输出rank；不足时返回ivr=None。
    """
    _ensure_iv_atm_history_table()
    conn = _db()
    rows = conn.execute(
        "SELECT iv_pct FROM iv_atm_history WHERE symbol=? "
        "AND date >= date('now','-365 days') ORDER BY date",
        (symbol,)).fetchall()
    conn.close()
    ivs = [r[0] for r in rows] + [current_iv]
    if len(ivs) < 30:
        return {"ivr": None, "iv_low": None, "iv_high": None,
                "iv_days": len(ivs), "note": f"积累中({len(ivs)}/30天)"}
    lo, hi = min(ivs), max(ivs)
    ivr = round((current_iv - lo) / (hi - lo) * 100, 1) if hi > lo else 50.0
    return {"ivr": ivr, "iv_low": round(lo, 1), "iv_high": round(hi, 1),
            "iv_days": len(ivs), "note": None}


@st.cache_data(ttl=900)
def _fetch_ivrank_md(symbol: str) -> dict:
    """
    从 MarketData.app 期权链提取ATM IV（delta最接近0.50的call，20-90 DTE）。
    本地积累历史后自动计算IVRank。
    返回 {iv, ivr, iv_low, iv_high, iv_days, error, source}
    """
    out = {"iv": None, "ivr": None, "iv_low": None, "iv_high": None,
           "iv_days": 0, "error": None, "source": "chain"}
    if not _MD_KEY:
        out["error"] = "未配置 MARKETDATA_API_KEY"; return out
    try:
        url = (f"{_MD_BASE}/options/chain/{symbol}/"
               f"?side=call&minDte=10&maxDte=90&token={_MD_KEY}")
        with urllib.request.urlopen(url, timeout=12) as resp:
            d = json.loads(resp.read())

        if d.get("s") != "ok":
            # 把真实错误原因透传出来
            err_msg = d.get("errmsg") or d.get("s") or "期权链无数据"
            out["error"] = f"chain返回 {err_msg}"; return out

        deltas = d.get("delta", [])
        ivs    = d.get("iv",    [])
        if not deltas:
            out["error"] = "期权链数据为空（该标的可能无期权）"; return out

        # 找 delta 最接近 0.50 的 call（ATM）
        best_iv, best_diff = None, 99.0
        for delta, iv in zip(deltas, ivs):
            if delta is None or iv is None or float(iv) <= 0:
                continue
            diff = abs(float(delta) - 0.50)
            if diff < best_diff:
                best_diff = diff
                best_iv = float(iv)

        if best_iv is None:
            out["error"] = "期权链中无有效delta/iv数据"; return out

        iv_pct = round(best_iv * 100, 1)   # chain返回的是小数(0.48 = 48%)
        out["iv"] = iv_pct

        # 存入历史 + 计算IVRank
        _save_iv_reading(symbol, iv_pct)
        rank_info = _compute_ivrank_from_history(symbol, iv_pct)
        out.update(rank_info)

    except urllib.error.HTTPError as e:
        out["error"] = f"HTTP {e.code} — {e.reason}"
    except urllib.error.URLError as e:
        out["error"] = f"网络错误 — {e.reason}"
    except Exception as e:
        out["error"] = str(e)[:140]
    return out


def _fetch_pltr_iv_snapshot() -> dict:
    """向后兼容的 PLTR 快照 — 直接转发到链式IV接口。"""
    return _fetch_ivrank_md("PLTR")


_IV_WATCH_SYMS = ("PLTR", "NOK", "NVDA", "FCX", "META", "PANW", "CRWV", "QQQ")


@st.cache_data(ttl=900)
def _fetch_iv_monitor_batch(symbols: tuple) -> dict:
    """批量拉取 IVRank，结果缓存15分钟。返回 {sym: {...}} 字典。"""
    return {s: _fetch_ivrank_md(s) for s in symbols}


def _compute_qqq_hedge_plan(acct_id: str, target_bd_ratio: float = 1.50) -> dict:
    """
    计算 QQQ Put Debit Spread 对冲数量。
    target_bd_ratio: 目标 Beta-Delta/净值 (小数，如 1.50 = 150%)。
    返回三套方案 (A=标准$35宽, B=宽幅$60, C=保留现有+补充)。
    """
    import json as _json, math as _math
    from scipy.stats import norm as _sn

    # ── 当前快照 ──────────────────────────────────────────────────
    snap = _compute_risk_snapshot(acct_id)
    if "error" in snap:
        return {"error": "no_data"}

    equity      = snap["equity"]
    current_bd  = snap["beta_delta"]                   # $ amount
    current_bdr = (snap["beta_delta_ratio"] or 0)      # fraction (e.g. 3.23)

    target_bd    = target_bd_ratio * equity
    bd_to_hedge  = current_bd - target_bd               # positive ⟹ need to reduce

    # ── QQQ 价格 + IV ─────────────────────────────────────────────
    qqq_price = _fetch_underlying_prices(("QQQ",)).get("QQQ", 0.0)
    if qqq_price <= 0:
        return {"error": "no_qqq_price"}

    _iv_d      = _fetch_ivrank_md("QQQ")
    qqq_iv_pct = _iv_d.get("iv") or 20.0    # in %
    qqq_iv     = qqq_iv_pct / 100            # decimal for BS
    b_qqq      = _BETA_SPY.get("QQQ", 1.31)

    # ── 内联 BS put 定价 ──────────────────────────────────────────
    def _bs_put_price(S, K, T, sigma):
        if T <= 1e-6 or sigma <= 1e-6 or S <= 0:
            return max(K - S, 0.0)
        try:
            d1 = (_math.log(S / K) + (_RF_RATE + 0.5 * sigma**2) * T) / (sigma * _math.sqrt(T))
            d2 = d1 - sigma * _math.sqrt(T)
            return K * _math.exp(-_RF_RATE * T) * _sn.cdf(-d2) - S * _sn.cdf(-d1)
        except Exception:
            return max(K - S, 0.0)

    # ── 现有 QQQ 持仓 ─────────────────────────────────────────────
    conn = _db()
    _q_rows = conn.execute(
        "SELECT symbol, quantity, current_price, delta, market_value "
        "FROM options_positions WHERE account_id=? AND symbol LIKE 'QQQ%'",
        (acct_id,)).fetchall()
    conn.close()

    existing_legs = []
    existing_bd   = 0.0
    for r in _q_rows:
        sym = (r["symbol"] or "").upper()
        mo  = _OCC_RE.match(sym)
        if not mo:
            continue
        q      = float(r["quantity"]  or 0)
        d      = float(r["delta"]     or 0)
        price  = float(r["current_price"] or 0)
        strike = float(mo.group(6)) / 1000
        opt_t  = mo.group(5)                        # "C" or "P"
        exp_s  = f"20{mo.group(2)}-{mo.group(3)}-{mo.group(4)}"
        existing_legs.append({
            "sym": sym, "qty": q, "type": opt_t,
            "strike": strike, "expiry": exp_s,
            "delta": d, "price": price,
        })
        existing_bd += q * 100 * d * qqq_price * b_qqq

    n_existing = sum(int(l["qty"]) for l in existing_legs
                     if l["type"] == "P" and l["qty"] > 0)

    # ── 参考买入行权价（最近 $5 整数）────────────────────────────
    _long_ks = [l["strike"] for l in existing_legs
                if l["type"] == "P" and l["qty"] > 0]
    if _long_ks:
        ref_buy_k = float(max(_long_ks))
    else:
        ref_buy_k = float(round(qqq_price * 0.97 / 5) * 5)

    # ── 计划到期：约 90 DTE ───────────────────────────────────────
    plan_exp     = datetime.date.today() + datetime.timedelta(days=90)
    plan_dte     = 90
    plan_occ_exp = plan_exp.strftime("%y%m%d")

    # ── 当前期权成本率（用于执行后估算）─────────────────────────
    _ocr_now = _compute_options_cost_ratio(acct_id)
    _cur_cost = (_ocr_now["ratio"] or 0) * equity

    # ── 内联方案计算 ──────────────────────────────────────────────
    def _plan(buy_k, sell_k):
        T = plan_dte / 365.0
        gb = _bs_greeks(qqq_price, buy_k,  T, qqq_iv, "put")
        gs = _bs_greeks(qqq_price, sell_k, T, qqq_iv, "put")
        d_buy, th_buy   = gb["delta"], gb["theta"]
        d_sell, th_sell = gs["delta"], gs["theta"]

        # BD contribution per spread: buy 1 put at buy_k + sell 1 put at sell_k
        #   buy leg:  +1 × 100 × d_buy  × qqq_price × b_qqq  (negative, d_buy<0)
        #   sell leg: -1 × 100 × d_sell × qqq_price × b_qqq  (positive, since d_sell<0)
        #   net = 100 × (d_buy - d_sell) × qqq_price × b_qqq  (negative, reduces BD)
        bd_ps = 100 * (d_buy - d_sell) * qqq_price * b_qqq

        if bd_to_hedge > 0 and bd_ps < 0:
            n_total = _math.ceil(bd_to_hedge / (-bd_ps))
        else:
            n_total = 0

        p_buy  = _bs_put_price(qqq_price, buy_k,  T, qqq_iv) * 100
        p_sell = _bs_put_price(qqq_price, sell_k, T, qqq_iv) * 100
        cost_ps = p_buy - p_sell     # net debit per spread

        post_bdr    = ((current_bd + n_total * bd_ps) / equity * 100) if equity else 0
        theta_chg   = n_total * (th_buy - th_sell) * 100    # $/day change
        new_cost_tot = _cur_cost + n_total * cost_ps
        new_ocr     = (new_cost_tot / equity * 100) if equity else 0

        return {
            "buy_strike":    buy_k,  "sell_strike": sell_k,
            "n_total":       n_total,
            "bd_per_spread": round(bd_ps, 0),
            "d_long":        round(d_buy,  3),
            "d_short":       round(d_sell, 3),
            "cost_per_spread": round(cost_ps, 2),
            "total_cost":    round(n_total * cost_ps, 0),
            "post_bd_ratio": round(post_bdr, 1),
            "theta_change":  round(theta_chg, 2),
            "new_ocr":       round(new_ocr, 1),
        }

    plan_a = _plan(ref_buy_k, ref_buy_k - 35)
    plan_b = _plan(ref_buy_k, ref_buy_k - 60)

    # Plan C: 保留现有 + 补充（用方案A的价差结构）
    remaining_bd = bd_to_hedge + existing_bd     # still-unhedged BD after existing positions
    if remaining_bd <= 0:
        n_add = 0
    elif plan_a["bd_per_spread"] >= 0:
        n_add = 0
    else:
        n_add = _math.ceil(remaining_bd / (-plan_a["bd_per_spread"]))
    post_bdc = ((current_bd + existing_bd + n_add * plan_a["bd_per_spread"])
                / equity * 100) if equity else 0
    add_cost_c = n_add * plan_a["cost_per_spread"]

    plan_c = {
        "n_existing":     n_existing,
        "n_additional":   n_add,
        "buy_strike":     ref_buy_k,
        "sell_strike":    ref_buy_k - 35,
        "cost_per_spread": plan_a["cost_per_spread"],
        "additional_cost": round(add_cost_c, 0),
        "post_bd_ratio":  round(post_bdc, 1),
        "existing_bd":    round(existing_bd, 0),
    }

    return {
        "equity":          equity,
        "current_bd_ratio": round(current_bdr * 100, 1),
        "target_bd_ratio":  round(target_bd_ratio * 100, 1),
        "bd_to_hedge":      round(bd_to_hedge, 0),
        "qqq_price":        round(qqq_price, 2),
        "qqq_iv":           round(qqq_iv_pct, 1),
        "b_qqq":            b_qqq,
        "existing_legs":    existing_legs,
        "existing_bd":      round(existing_bd, 0),
        "n_existing":       n_existing,
        "plan_dte":         plan_dte,
        "plan_occ_exp":     plan_occ_exp,
        "plan_exp_str":     plan_exp.strftime("%Y-%m-%d"),
        "plan_a":           plan_a,
        "plan_b":           plan_b,
        "plan_c":           plan_c,
    }


def _pltr_ivr_signal(iv: float, ivr: float | None) -> tuple[str, str, str]:
    """返回 (icon, 建议文字, hex颜色)。"""
    if ivr is None:
        return "⚪", "历史数据不足，等待积累", _MUTED
    if ivr < 25:
        return "⚪", "等待 — IV历史偏低，权利金太薄", _MUTED
    if ivr < 40:
        return "⚪", f"等待 — IVR {ivr:.0f}%（需 >40）", _MUTED
    # IVR >= 40
    if 40 <= ivr < 60:
        if iv >= 50:
            return "🟡", "可以卖1张 Call", _AMB
        return "⚪", f"等待 — IV {iv:.0f}% 不足（需 >50%）", _MUTED
    if 60 <= ivr < 80:
        if iv >= 60:
            return "🟠", "建议卖2张 Call", _AMB
        if iv >= 50:
            return "🟡", "可以卖1张 Call（IVR满足，IV未达60%）", _AMB
        return "⚪", f"等待 — IV {iv:.0f}% 不足", _MUTED
    # IVR >= 80
    if iv >= 70:
        return "🔴", "最佳时机，卖2-3张 Call", _RED
    if iv >= 60:
        return "🟠", "建议卖2张 Call（IVR极高，IV未达70%）", _AMB
    if iv >= 50:
        return "🟡", "可以卖1张 Call（IVR极高，IV偏低）", _AMB
    return "⚪", f"等待 — IV {iv:.0f}% 不足", _MUTED


# ─────────────────────────────────────────────────────────────────
# Module C-SIM: 策略模拟器 — 辅助函数
# ─────────────────────────────────────────────────────────────────

def _ensure_sim_table() -> None:
    """Create simulation_log table if it doesn't exist."""
    conn = _db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS simulation_log (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            acct_id   TEXT    NOT NULL,
            sim_ts    TEXT    NOT NULL,
            actions   TEXT,
            before_m  TEXT,
            after_m   TEXT,
            scenarios TEXT,
            notes     TEXT
        )
    """)
    conn.commit()
    conn.close()


def _compute_sim_impact(acct_id: str, sim_actions: list, base_snap: dict,
                        hplan: dict | None = None) -> dict:
    """
    Apply sim_actions to base_snap and return post-simulation metrics.
    action types: 'close_underlying', 'qqq_hedge', 'no_sim', 'no_change'
    """
    equity  = base_snap.get("equity", 1)
    bd_delta_sim    = theta_delta_sim = vega_delta_sim = s10_delta_sim = s20_delta_sim = cash_delta_sim = 0.0
    descs: list[str] = []

    # Collect underlyings needed for close actions
    _unds = {a["underlying"] for a in sim_actions if a.get("type") == "close_underlying"}
    _und_p = _fetch_underlying_prices(tuple(sorted(_unds))) if _unds else {}

    for action in sim_actions:
        atype = action.get("type", "no_change")

        if atype == "close_underlying":
            und  = action["underlying"]
            S    = _und_p.get(und, 0.0)
            b    = _BETA_SPY.get(und, 1.0)
            mult = 100.0

            conn  = _db()
            _rows = conn.execute(
                "SELECT quantity, current_price, delta, gamma, theta, vega, market_value "
                "FROM options_positions WHERE account_id=? AND symbol LIKE ?",
                (acct_id, f"{und}%")).fetchall()
            conn.close()

            _mv_sum = 0.0
            for r in _rows:
                q   = float(r[0] or 0)
                d   = float(r[2] or 0)
                g   = float(r[3] or 0)
                th  = float(r[4] or 0)
                vg  = float(r[5] or 0)
                mv  = float(r[6] or 0)
                bd_delta_sim    -= q * mult * d * S * b if S > 0 else 0.0
                theta_delta_sim -= q * mult * th
                vega_delta_sim  -= q * mult * vg
                _mv_sum += mv
                if S > 0:
                    ds10 = -0.10 * S;  ds20 = -0.20 * S
                    s10_delta_sim -= (q * mult * (d * ds10 + 0.5 * g * ds10**2)
                               + q * mult * vg * 8)
                    s20_delta_sim -= (q * mult * (d * ds20 + 0.5 * g * ds20**2)
                               + q * mult * vg * 16)
            cash_delta_sim += _mv_sum
            descs.append(f"关闭 {und} 期权（收回约${_mv_sum:+,.0f}）")

        elif atype == "qqq_hedge":
            if not hplan or "error" in hplan:
                descs.append("QQQ对冲（数据不足，跳过）");  continue
            _pa = hplan.get("plan_a", {})
            n   = _pa.get("n_total", 0)
            if n <= 0:
                descs.append("QQQ对冲（现有对冲已足够）");  continue
            S   = hplan["qqq_price"];  iv  = hplan["qqq_iv"] / 100
            b   = hplan["b_qqq"];      T   = hplan["plan_dte"] / 365.0
            bk  = _pa["buy_strike"];   sk  = _pa["sell_strike"]
            mult = 100.0
            gb = _bs_greeks(S, bk, T, iv, "put")
            gs = _bs_greeks(S, sk, T, iv, "put")
            dn  = gb["delta"] - gs["delta"];   gn  = gb["gamma"] - gs["gamma"]
            thn = gb["theta"] - gs["theta"];   vgn = gb["vega"]  - gs["vega"]
            bd_delta_sim    += n * mult * dn  * S * b
            theta_delta_sim += n * mult * thn
            vega_delta_sim  += n * mult * vgn
            ds10 = -0.10 * S;  ds20 = -0.20 * S
            s10_delta_sim += n * (mult * (dn * ds10 + 0.5 * gn * ds10**2) + mult * vgn * 8)
            s20_delta_sim += n * (mult * (dn * ds20 + 0.5 * gn * ds20**2) + mult * vgn * 16)
            cash_delta_sim -= _pa.get("total_cost", 0)
            descs.append(f"QQQ Put Spread×{n}张（方案A，成本${_pa.get('total_cost',0):,.0f}）")

        else:
            descs.append(action.get("label", "持有（不变）"))

    new_bd   = base_snap.get("beta_delta", 0)    + bd_delta_sim
    new_bdr  = new_bd / equity if equity else 0
    new_th   = base_snap.get("theta_per_day", 0) + theta_delta_sim
    new_s10  = base_snap.get("stress_10", 0)     + s10_delta_sim
    new_s20  = base_snap.get("stress_20", 0)     + s20_delta_sim
    new_s10r = new_s10 / equity if equity else 0
    new_s20r = new_s20 / equity if equity else 0

    return {
        "bd_delta":    round(bd_delta_sim, 0),   "theta_delta": round(theta_delta_sim, 2),
        "s10_delta":   round(s10_delta_sim, 0),  "s20_delta":   round(s20_delta_sim, 0),
        "cash_delta":  round(cash_delta_sim, 0),
        "beta_delta":       round(new_bd,  0),
        "beta_delta_ratio": round(new_bdr * 100, 1),
        "theta_per_day":    round(new_th,  2),
        "stress_10":        round(new_s10, 0),
        "stress_10_ratio":  round(new_s10r * 100, 1),
        "stress_20":        round(new_s20, 0),
        "stress_20_ratio":  round(new_s20r * 100, 1),
        "actions":     descs,
    }


def _run_scenarios(snap: dict) -> dict:
    """Approximate P/L for 3 market scenarios."""
    equity = snap.get("equity", 1) or 1
    theta  = snap.get("theta_per_day", 0)
    s10    = snap.get("stress_10", 0)
    s20    = snap.get("stress_20", 0)
    vega   = snap.get("vega_per_pt", 0)
    bd     = snap.get("beta_delta", 0)

    calm   = theta * 30 + bd * 0.01           # 30天 Theta + 小幅市场漂移
    drop10 = s10                               # 已经计算好
    # +10%: 反向 delta/gamma，但 IV 通常随上涨下降 ~8pts → vega 损失
    rally10 = -s10 * 0.70 + vega * (-8)

    return {
        "calm":    round(calm, 0),
        "drop10":  round(drop10, 0),
        "rally10": round(rally10, 0),
        "equity":  equity,
    }


def _save_sim_record(acct_id: str, actions: list, before: dict,
                     after: dict, scenarios: dict) -> int:
    import json as _json
    _ensure_sim_table()
    conn = _db()
    cur  = conn.execute(
        "INSERT INTO simulation_log (acct_id,sim_ts,actions,before_m,after_m,scenarios) "
        "VALUES (?,?,?,?,?,?)",
        (acct_id, datetime.datetime.now().isoformat(),
         _json.dumps(actions, ensure_ascii=False),
         _json.dumps(before,  ensure_ascii=False),
         _json.dumps(after,   ensure_ascii=False),
         _json.dumps(scenarios, ensure_ascii=False)))
    conn.commit()
    rid = cur.lastrowid
    conn.close()
    return rid


def _load_sim_records(acct_id: str, limit: int = 8) -> list[dict]:
    import json as _json
    try:
        _ensure_sim_table()
        conn  = _db()
        rows  = conn.execute(
            "SELECT id,sim_ts,actions,before_m,after_m,scenarios "
            "FROM simulation_log WHERE acct_id=? ORDER BY id DESC LIMIT ?",
            (acct_id, limit)).fetchall()
        conn.close()
        out = []
        for r in rows:
            out.append({
                "id":        r[0],
                "sim_ts":    (r[1] or "")[:16].replace("T", " "),
                "actions":   _json.loads(r[2] or "[]"),
                "before":    _json.loads(r[3] or "{}"),
                "after":     _json.loads(r[4] or "{}"),
                "scenarios": _json.loads(r[5] or "{}"),
            })
        return out
    except Exception:
        return []


# ─────────────────────────────────────────────────────────────────
# Module C: 候选交易建议
# ─────────────────────────────────────────────────────────────────
def _generate_recommendations(acct_id: str) -> list[dict]:
    """
    基于组合识别结果生成建议（以价差为单位，不拆单腿）。
    同时叠加：IV Regime / 压力测试 / AI评分 / 新机会候选。
    """
    snap      = _compute_risk_snapshot(acct_id)
    iv_regime = _compute_iv_regime(acct_id)
    ai_scores = _load_ai_scores()
    portfolios = _build_spread_portfolios(acct_id)

    iv_status = iv_regime.get("status", "NO_DATA")
    stress10  = snap.get("stress_10_ratio") or 0.0
    leverage  = snap.get("leverage_delta")  or snap.get("leverage") or 0.0  # 优先用 Delta 口径
    held_und  = {p["underlying"] for p in portfolios}

    _RISK_PRIORITY = {
        "CRITICAL": "🔴 紧急",
        "HIGH":     "🟠 高",
        "MEDIUM":   "🟡 中",
        "LOW":      "🟢 低",
    }
    _RISK_EMOJI = {
        "CRITICAL": "🔴", "HIGH": "🟠", "MEDIUM": "🟡", "LOW": "🟢",
    }

    recs = []
    idx  = 1

    # 1. Portfolio-level recommendations
    for p in portfolios:
        und        = p["underlying"]
        rl         = p["risk_level"]
        priority   = _RISK_PRIORITY.get(rl, "🟡 中")
        score      = ai_scores.get(und)
        ptype      = p["type"]
        pnl        = p.get("current_pnl") or 0
        dte_v      = p.get("dte")
        pnl_pct    = p.get("pnl_pct")

        # Build action string
        action_parts: list[str] = [p["recommendation"]]

        # AI score supplement
        if score is not None:
            action_parts.append(f"AI评分 {_score_label(score)}")

        # IV-driven advice for short vega spreads (credit spreads)
        is_short_vega = "Credit" in ptype or "Bear Call" in ptype or "Bull Put" in ptype
        if is_short_vega and iv_status in ("HIGH_IV", "EXTREME_IV"):
            action_parts.append(f"IV Regime={iv_status}，卖权环境有利，可持有至 80% 利润后平仓")
        elif "Debit" in ptype or "Bull Call" in ptype or "Bear Put" in ptype:
            if iv_status == "LOW_IV":
                action_parts.append("低 IV 环境，买权价差成本低，有利于持有")

        # Portfolio stress supplement
        if abs(stress10) >= _RISK_LIMITS["stress_hard_stop"]:
            action_parts.append("⚠️ 组合压力超限(≥15%)，优先减仓")

        # Spread-specific hold/roll/close logic
        if rl == "LOW" and pnl_pct is not None:
            if pnl_pct >= 50:
                action_parts.append("已盈利50%+，可考虑提前平仓锁利")
            else:
                action_parts.append("建议继续持有至75%利润或到期2周前评估")
        elif rl == "MEDIUM" and "Diagonal" in ptype:
            action_parts.append("关注近月腿到期节点，提前15天制定展期方案")

        trigger_parts = [f"组合类型={ptype}", f"风险={rl}"]
        if dte_v is not None:
            trigger_parts.append(f"DTE={dte_v}天")
        if pnl != 0:
            trigger_parts.append(f"盈亏=${pnl:+,.0f}")

        recs.append({
            "序号":   idx,
            "优先级": priority,
            "标的":   und,
            "组合":   ptype,
            "手数":   p.get("spread_qty", 0),
            "到期":   p.get("expiry", "—"),
            "DTE":    dte_v if dte_v is not None else "—",
            "AI评分": _score_label(score) if score else "—",
            "行动建议": " ；".join(action_parts),
            "触发原因": " | ".join(trigger_parts),
            "最大盈利": f"${p['max_profit']:,.0f}" if p.get("max_profit") is not None else "—",
            "最大亏损": f"${p['max_loss']:,.0f}"  if p.get("max_loss")   is not None else "无限",
            "当前盈亏": f"${pnl:+,.0f}",
            "_sim_action": {
                "type": "close_underlying", "underlying": und,
                "label": f"关闭 {und} 全部期权持仓",
            },
        })
        idx += 1

    # 2. Portfolio stress hedge suggestion
    if abs(stress10) >= _RISK_LIMITS["stress_de_risk"]:
        recs.append({
            "序号":   idx, "优先级": "🟠 高", "标的": "QQQ",
            "组合":   "宏观对冲（建议）", "手数": 0, "到期": "—", "DTE": "—",
            "AI评分": _score_label(ai_scores.get("QQQ")) if ai_scores.get("QQQ") else "—",
            "行动建议": "建议买入 QQQ Put Debit Spread 作宏观对冲，最大亏损 = 净权利金",
            "触发原因": f"-10% 压力损失 {stress10*100:+.1f}% ≥ {_RISK_LIMITS['stress_de_risk']*100:.0f}% 阈值",
            "最大盈利": "—", "最大亏损": "= 权利金", "当前盈亏": "—",
            "_sim_action": {"type": "qqq_hedge", "label": "执行 QQQ Put Spread 对冲（方案A）"},
        })
        idx += 1

    # 3. New opportunity candidates: high AI score + IV regime fit + risk headroom
    if not snap.get("error") and leverage < _RISK_LIMITS["max_leverage"] * 0.75:
        candidates = [(t, s) for t, s in ai_scores.items()
                      if s >= 70 and t not in held_und]
        for t, s in sorted(candidates, key=lambda x: -x[1])[:3]:
            if iv_status in ("HIGH_IV", "EXTREME_IV"):
                strat  = "卖出 Put Credit Spread（高 IV 收权利金，限定风险）"
                reason = f"IV Regime={iv_status} 适合卖权；{t} AI评分={s:.0f}"
            elif iv_status == "LOW_IV":
                strat  = "买入 Call Debit Spread（低 IV 低成本买权）"
                reason = f"IV Regime=LOW_IV 适合买权；{t} AI评分={s:.0f}"
            else:
                strat  = "观望或小仓 Bull Call Spread（中性 IV）"
                reason = f"{t} AI评分={s:.0f}，IV 正常区间"
            recs.append({
                "序号":   idx, "优先级": "🟢 机会", "标的": t,
                "组合":   "新开仓候选", "手数": 0, "到期": "—", "DTE": "—",
                "AI评分": _score_label(s),
                "行动建议": strat,
                "触发原因": reason,
                "最大盈利": "—", "最大亏损": "= 权利金", "当前盈亏": "—",
                "_sim_action": {"type": "no_sim",
                                "label": f"新开仓 {t}（需指定具体参数，暂不支持模拟）"},
            })
            idx += 1

    return recs


# ─────────────────────────────────────────────────────────────────
# Feature 1: 每日作战简报 — 持久化存储 + 09:35 自动生成
# ─────────────────────────────────────────────────────────────────

def _ensure_daily_briefing_table() -> None:
    conn = _db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS daily_briefing (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            acct_id   TEXT NOT NULL,
            date      TEXT NOT NULL,
            gen_time  TEXT NOT NULL,
            snap_json TEXT,
            recs_json TEXT,
            alerts_json TEXT,
            UNIQUE(acct_id, date)
        )
    """)
    conn.commit()
    conn.close()


def _save_daily_briefing(acct_id: str, snap: dict,
                          recs: list, alerts: list) -> None:
    import json as _j
    _ensure_daily_briefing_table()
    today   = datetime.date.today().isoformat()
    now_str = datetime.datetime.now(_ET).strftime("%Y-%m-%dT%H:%M:%S")
    conn = _db()
    conn.execute(
        "INSERT OR REPLACE INTO daily_briefing "
        "(acct_id, date, gen_time, snap_json, recs_json, alerts_json) "
        "VALUES (?,?,?,?,?,?)",
        (acct_id, today, now_str,
         _j.dumps(snap,   ensure_ascii=False, default=str),
         _j.dumps(recs,   ensure_ascii=False, default=str),
         _j.dumps(alerts, ensure_ascii=False, default=str)))
    conn.commit()
    conn.close()


def _load_today_briefing(acct_id: str) -> dict | None:
    import json as _j
    try:
        _ensure_daily_briefing_table()
        today = datetime.date.today().isoformat()
        conn  = _db()
        row   = conn.execute(
            "SELECT gen_time, snap_json, recs_json, alerts_json "
            "FROM daily_briefing WHERE acct_id=? AND date=?",
            (acct_id, today)).fetchone()
        conn.close()
        if not row:
            return None
        return {
            "gen_time": row[0],
            "snap":     _j.loads(row[1] or "{}"),
            "recs":     _j.loads(row[2] or "[]"),
            "alerts":   _j.loads(row[3] or "[]"),
        }
    except Exception:
        return None


def _generate_and_save_daily_briefing(acct_id: str) -> None:
    """09:35 ET 自动执行：生成简报并存库。"""
    try:
        snap    = _compute_risk_snapshot(acct_id)
        recs    = _generate_recommendations(acct_id)
        alerts  = _check_sell_call_triggers(acct_id)
        # snap 含 datetime.date 对象，需转为 str
        snap_s  = {k: (v.isoformat() if isinstance(v, datetime.date) else v)
                   for k, v in snap.items()}
        _save_daily_briefing(acct_id, snap_s, recs, alerts)
        _log.info(f"Daily briefing saved for {acct_id}: "
                  f"{len(recs)} recs, {len(alerts)} call alerts")
    except Exception as e:
        _log.error(f"Daily briefing generation failed ({acct_id}): {e}")


# ─────────────────────────────────────────────────────────────────
# Feature 2: 卖Call实时监控 — 检查所有 Long Call 持仓触发条件
# ─────────────────────────────────────────────────────────────────

def _check_sell_call_triggers(acct_id: str) -> list[dict]:
    """
    扫描所有 Long Call 持仓，检查是否满足「卖出 Covered Call」触发条件：
      A. 盈利 ≥ 50%（总盈亏 / 建仓成本）
      B. DTE ≤ 14（时间价值衰减加速，需决策）
      C. Delta ≥ 0.72（深度价内，可能需转换或平仓）
      D. IV 极高（IVR ≥ 75，卖Call收权利金时机好）
    返回 list[dict] 每条包含: sym / underlying / expiry / dte / trigger /
                              suggestion / strike / delta / iv / pnl_pct
    """
    triggers: list[dict] = []
    today = datetime.date.today()

    conn = _db()
    rows = conn.execute(
        "SELECT symbol, quantity, unit_cost, current_price, total_pnl, "
        "       delta, iv, expiry, strike "
        "FROM options_positions "
        "WHERE account_id=? AND quantity > 0",   # Long positions only
        (acct_id,)).fetchall()
    conn.close()

    for r in rows:
        sym  = (r["symbol"] or "").upper()
        mo   = _OCC_RE.match(sym)
        if not mo:
            continue
        opt_type = mo.group(5)   # C or P
        if opt_type != "C":
            continue             # only Long Calls

        und    = mo.group(1)
        qty    = int(r["quantity"] or 0)
        cost   = float(r["unit_cost"]     or 0)
        cprice = float(r["current_price"] or 0)
        pnl    = float(r["total_pnl"]     or 0)
        delta  = float(r["delta"]         or 0)
        iv_pct = float(r["iv"]            or 0)
        strike = float(r["strike"]        or (mo and float(mo.group(6)) / 1000) or 0)

        # Parse expiry
        try:
            exp_date = datetime.date(2000 + int(mo.group(2)),
                                     int(mo.group(3)), int(mo.group(4)))
            dte = (exp_date - today).days
        except Exception:
            exp_date = None
            dte = None

        # Compute P&L %
        basis = abs(cost) * qty * 100
        pnl_pct = (pnl / basis * 100) if basis > 1e-6 else None

        # Evaluate trigger conditions
        hit: list[str] = []
        suggestions: list[str] = []

        if pnl_pct is not None and pnl_pct >= 50:
            hit.append(f"A:盈利{pnl_pct:.0f}%")
            suggestions.append(f"已盈利 {pnl_pct:.0f}%，可平仓锁利或滚动至更高行权价卖Call")

        if dte is not None and dte <= 14:
            hit.append(f"B:DTE={dte}")
            if dte <= 7:
                suggestions.append(f"⚠️ 仅剩 {dte} 天到期，需立即决定：平仓/展期/行权")
            else:
                suggestions.append(f"还有 {dte} 天到期，Theta加速衰减，建议在2周内决策")

        if abs(delta) >= 0.72:
            hit.append(f"C:Delta={delta:.2f}")
            suggestions.append(f"Delta={delta:.2f}（深度价内），可考虑行权或平仓转股票")

        # IV check from cached iv_monitor batch
        try:
            _iv_d = _fetch_ivrank_md(und)
            ivr   = _iv_d.get("ivr")
        except Exception:
            ivr   = None
        if ivr is not None and ivr >= 75:
            hit.append(f"D:IVR={ivr:.0f}%")
            suggestions.append(f"{und} IVR={ivr:.0f}%（极高），是对Long Call卖出Covered Call收权利金的好时机")

        if not hit:
            continue

        occ_exp = exp_date.strftime("%y%m%d") if exp_date else "??????"
        # Suggest covered call: strike 5-10% OTM from current underlying price
        # (we don't have underlying price here, use strike as reference)
        suggestions_text = " ；".join(suggestions)

        triggers.append({
            "sym":        sym,
            "underlying": und,
            "expiry":     exp_date.isoformat() if exp_date else "—",
            "dte":        dte,
            "strike":     strike,
            "delta":      round(delta, 3),
            "iv_pct":     iv_pct,
            "ivr":        ivr,
            "pnl_pct":    round(pnl_pct, 1) if pnl_pct is not None else None,
            "pnl":        round(pnl, 0),
            "qty":        qty,
            "triggers":   hit,
            "suggestion": suggestions_text,
            "occ_exp":    occ_exp,
        })

    # Sort: most triggers first, then by DTE ascending
    triggers.sort(key=lambda x: (-len(x["triggers"]), x["dte"] or 9999))
    return triggers


# ─────────────────────────────────────────────────────────────────
# 新预警 C: 价外程度预警（价差组合接近归零）
# ─────────────────────────────────────────────────────────────────

def _check_otm_spread_alerts(acct_id: str) -> list[dict]:
    """
    扫描 options_positions，识别价差组合，检测买入方价差（Debit Spread）是否接近归零。
    触发条件：价差两腿市值合计 < 原始净权利金的 10%。
    不使用 underlying_price vs strike 的简单比较，避免长期限期权误报。
    """
    from collections import defaultdict
    alerts: list[dict] = []
    today = datetime.date.today()

    conn = _db()
    rows = conn.execute(
        "SELECT symbol, quantity, strike, expiry, unit_cost, market_value, current_price "
        "FROM options_positions WHERE account_id=?",
        (acct_id,)).fetchall()
    conn.close()

    # Group legs by (underlying, expiry, option_type)
    groups: dict = defaultdict(list)
    for r in rows:
        sym = (r["symbol"] or "").upper()
        mo = _OCC_RE.match(sym)
        if not mo:
            continue
        und      = mo.group(1)
        exp_str  = f"20{mo.group(2)}-{mo.group(3)}-{mo.group(4)}"
        opt_type = mo.group(5)   # "C" or "P"
        strike   = float(r["strike"] or 0) or float(mo.group(6)) / 1000
        qty      = float(r["quantity"] or 0)
        uc       = float(r["unit_cost"]    or 0)
        # Use market_value if available; fall back to current_price * |qty| * 100
        mv = r["market_value"]
        if mv is None and r["current_price"] is not None:
            cp = float(r["current_price"])
            mv = cp * abs(qty) * 100 * (1 if qty > 0 else -1)
        mv = float(mv) if mv is not None else None
        groups[(und, exp_str, opt_type)].append({
            "sym": sym, "qty": qty, "strike": strike,
            "unit_cost": uc, "market_value": mv,
        })

    checked: set = set()
    for (und, exp_str, opt_type), legs in groups.items():
        longs  = [l for l in legs if l["qty"] > 0]
        shorts = [l for l in legs if l["qty"] < 0]
        if not longs or not shorts:
            continue

        try:
            exp_date = datetime.date.fromisoformat(exp_str)
            dte = (exp_date - today).days
        except Exception:
            dte = None

        for long_leg in longs:
            for short_leg in shorts:
                key = (und, exp_str, opt_type, long_leg["strike"], short_leg["strike"])
                if key in checked:
                    continue
                checked.add(key)

                ls = long_leg["strike"]
                ss = short_leg["strike"]
                high_k = max(ls, ss)
                low_k  = min(ls, ss)
                qty_used = min(abs(long_leg["qty"]), abs(short_leg["qty"]))

                # Determine spread type and whether it's a debit spread
                spread_label = ""
                is_debit     = False

                if opt_type == "P" and ls > ss:
                    spread_label = "Bear Put Spread"
                    is_debit     = True   # paid net premium (long higher-strike P)
                elif opt_type == "P" and ls < ss:
                    spread_label = "Bull Put Spread"
                    is_debit     = False  # received net credit
                elif opt_type == "C" and ls < ss:
                    spread_label = "Bull Call Spread"
                    is_debit     = True   # paid net premium (long lower-strike C)
                elif opt_type == "C" and ls > ss:
                    spread_label = "Bear Call Spread"
                    is_debit     = False  # received net credit

                if not spread_label or not is_debit:
                    continue   # only alert on debit spreads approaching zero

                # Original net cost (debit paid per spread)
                net_cost_per_share = long_leg["unit_cost"] - abs(short_leg["unit_cost"])
                original_cost      = max(net_cost_per_share * qty_used * 100, 0.01)

                # Current spread value (sum of signed market_values)
                lmv = long_leg["market_value"]
                smv = short_leg["market_value"]
                if lmv is None or smv is None:
                    continue   # no market_value data — skip
                current_value = lmv + smv   # long=positive, short=negative → net

                # Trigger if current value < 10% of original cost
                pct_remaining = current_value / original_cost * 100 if original_cost > 0 else 100
                if pct_remaining >= 10.0:
                    continue

                alerts.append({
                    "underlying":     und,
                    "spread_type":    spread_label,
                    "high_strike":    high_k,
                    "low_strike":     low_k,
                    "opt_type":       opt_type,
                    "expiry":         exp_str,
                    "dte":            dte,
                    "original_cost":  round(original_cost, 2),
                    "current_value":  round(current_value, 2),
                    "pct_remaining":  round(pct_remaining, 1),
                    "long_sym":       long_leg["sym"],
                    "short_sym":      short_leg["sym"],
                    "message": (
                        f"{und} {spread_label} ${low_k:.0f}/{high_k:.0f}"
                        f"{'P' if opt_type=='P' else 'C'} "
                        f"当前价差市值 ${current_value:.0f}，"
                        f"仅剩原始成本 ${original_cost:.0f} 的 {pct_remaining:.1f}%，"
                        f"价差接近归零"
                    ),
                })

    alerts.sort(key=lambda x: x["pct_remaining"])   # lowest remaining % first
    return alerts


# ─────────────────────────────────────────────────────────────────
# 新预警 D: 标的单日异动预警（±5% 单日波动）
# ─────────────────────────────────────────────────────────────────

@st.cache_data(ttl=300)
def _fetch_underlying_day_changes(tickers_tuple: tuple) -> dict:
    """yfinance 拉取标的当日涨跌幅（5分钟缓存）。返回 {ticker: {price, prev_close, chg_pct}}"""
    import yfinance as yf
    import pandas as pd
    result = {}
    try:
        tickers = list(tickers_tuple)
        data = yf.download(tickers, period="2d", interval="1d",
                           progress=False, auto_adjust=True)
        closes = data["Close"] if "Close" in data else data
        if isinstance(closes, pd.Series):
            closes = closes.to_frame(name=tickers[0])
        for t in tickers:
            try:
                ser = closes[t].dropna() if t in closes.columns else None
                if ser is None or len(ser) < 2:
                    continue
                prev_c = float(ser.iloc[-2])
                last_c = float(ser.iloc[-1])
                if prev_c > 0:
                    result[t] = {
                        "price":      last_c,
                        "prev_close": prev_c,
                        "chg_pct":    (last_c - prev_c) / prev_c * 100,
                    }
            except Exception:
                pass
    except Exception as e:
        _log.warning(f"[day_changes] yfinance error: {e}")
    return result


def _check_underlying_move_alerts(acct_id: str) -> list[dict]:
    """
    检查所有持仓期权的标的当日涨跌幅，超过 ±5% 触发异动预警。
    """
    alerts: list[dict] = []

    conn = _db()
    rows = conn.execute(
        "SELECT DISTINCT symbol FROM options_positions WHERE account_id=?",
        (acct_id,)).fetchall()
    conn.close()

    underlyings: set = set()
    for r in rows:
        sym = (r["symbol"] or "").upper()
        mo = _OCC_RE.match(sym)
        if mo:
            underlyings.add(mo.group(1))

    if not underlyings:
        return alerts

    day_changes = _fetch_underlying_day_changes(tuple(sorted(underlyings)))

    for und, info in day_changes.items():
        chg = info.get("chg_pct", 0)
        if abs(chg) >= 5.0:
            direction = "暴涨" if chg > 0 else "暴跌"
            alerts.append({
                "underlying": und,
                "price":      info["price"],
                "prev_close": info["prev_close"],
                "chg_pct":    round(chg, 1),
                "direction":  direction,
                "message": (
                    f"{und} 今日{direction} {abs(chg):.1f}%"
                    f"（${info['prev_close']:.2f} → ${info['price']:.2f}），"
                    f"请立即检查相关期权持仓的风险敞口"
                ),
            })

    alerts.sort(key=lambda x: -abs(x["chg_pct"]))
    return alerts


# ─────────────────────────────────────────────────────────────────
# Portfolio / Spread Recognition Engine
# ─────────────────────────────────────────────────────────────────
_DTE_CRITICAL = 14
_DTE_REVIEW   = 21


def _build_spread_portfolios(acct_id: str) -> list[dict]:
    """
    读取 options_positions，自动识别期权组合关系。
    优先级：同标的+同到期日+同方向 → 垂直价差；
            同标的+不同到期日+同方向 → 日历/对角价差；
            其余 → 裸仓。
    返回 list[dict]，按风险等级从高到低排序。
    """
    df = _load_options_positions(acct_id)
    if df.empty:
        return []

    today = datetime.date.today()

    def _days(exp_str):
        try:
            return (datetime.date.fromisoformat(str(exp_str or "")) - today).days
        except Exception:
            return 9999

    def _leg_pnl(leg, qty_frac=1.0):
        cp = leg.get("cur_price")
        uc = leg.get("unit_cost") or 0
        q  = leg.get("qty", 0)
        if cp is None:
            return (leg.get("total_pnl") or 0) * qty_frac
        return (float(cp) - float(uc)) * q * 100 * qty_frac

    # Parse all legs into uniform dicts
    legs = []
    for _, row in df.iterrows():
        sym = str(row.get("symbol", "")).strip().upper()
        p   = _parse_occ(sym)
        if not p:
            continue
        qty = int(row.get("quantity") or 0)
        if qty == 0:
            continue
        # Backward compat: old DB rows stored abs(qty) with direction="short"/"long"
        _db_dir = str(row.get("direction") or "").lower()
        if qty > 0 and _db_dir == "short":
            qty = -qty
        legs.append({
            "symbol":     sym,
            "underlying": p["root"],
            "direction":  p.get("option_type", p["direction"].lower()),   # "call" / "put"
            "strike":     p["strike"],
            "expiry":     p["expiry"],
            "qty":        qty,
            "unit_cost":  float(row["unit_cost"])     if row.get("unit_cost")     is not None else 0.0,
            "cur_price":  float(row["current_price"]) if row.get("current_price") is not None else None,
            "total_pnl":  float(row["total_pnl"])     if row.get("total_pnl")     is not None else None,
            "delta":      float(row["delta"])          if row.get("delta")         is not None else None,
            "iv":         float(row["iv"])             if row.get("iv")            is not None else None,
            "dte":        _days(p["expiry"]),
        })

    n       = len(legs)
    matched = [False] * n
    portfolios: list[dict] = []

    # ── helpers ──────────────────────────────────────────
    def _make_vertical(li: int, si: int) -> dict:
        ll = legs[li]   # long leg
        sl = legs[si]   # short leg
        qty = min(abs(ll["qty"]), abs(sl["qty"]))
        direction = ll["direction"]
        und  = ll["underlying"]

        if direction == "call":
            stype = "Bull Call Spread" if ll["strike"] < sl["strike"] else "Bear Call Spread"
        else:
            stype = "Bear Put Spread"  if ll["strike"] > sl["strike"] else "Bull Put Spread"

        low_s  = min(ll["strike"], sl["strike"])
        high_s = max(ll["strike"], sl["strike"])
        width  = round(high_s - low_s, 4)

        # positive net = debit paid; negative = credit received
        net_ps   = ll["unit_cost"] - sl["unit_cost"]
        net_tot  = round(net_ps * qty * 100, 2)
        is_debit = net_ps > 0

        if is_debit:
            max_profit = round((width - abs(net_ps)) * qty * 100, 2)
            max_loss   = round(abs(net_tot), 2)
            if direction == "call":
                breakeven = round(ll["strike"] + abs(net_ps), 4)
            else:
                breakeven = round(ll["strike"] - abs(net_ps), 4)
        else:
            max_profit = round(abs(net_tot), 2)
            max_loss   = round((width - abs(net_ps)) * qty * 100, 2)
            if direction == "call":
                breakeven = round(sl["strike"] + abs(net_ps), 4)
            else:
                breakeven = round(sl["strike"] - abs(net_ps), 4)

        long_frac  = qty / abs(ll["qty"])
        short_frac = qty / abs(sl["qty"])
        pnl = _leg_pnl(ll, long_frac) + _leg_pnl(sl, short_frac)

        dte_v = min(ll["dte"], sl["dte"])
        if dte_v <= _DTE_CRITICAL:
            risk_level = "CRITICAL"
        elif dte_v <= _DTE_REVIEW:
            risk_level = "HIGH"
        else:
            risk_level = "LOW"

        if is_debit:
            basis   = abs(max_loss)
            pnl_pct = round(pnl / basis * 100, 1) if basis else 0
            rec = (f"净权利金 ${abs(net_tot):,.0f}（付），"
                   f"当前盈亏 ${pnl:+,.0f}（{pnl_pct:+.0f}% on cost），"
                   f"最大盈利 ${max_profit:,.0f}，盈亏平衡点 ${breakeven:,.2f}")
        else:
            # 收权价差：pnl_pct = 盈亏 / 最大亏损（风险资本），而非 / 权利金收入
            basis   = abs(max_loss)
            pnl_pct = round(pnl / basis * 100, 1) if basis else 0
            rec = (f"收权利金 ${abs(net_tot):,.0f}，"
                   f"当前盈亏 ${pnl:+,.0f}（{pnl_pct:+.0f}% on risk），"
                   f"最大亏损 ${max_loss:,.0f}，盈亏平衡点 ${breakeven:,.2f}")

        if dte_v <= _DTE_CRITICAL:
            rec = f"🚨 DTE={dte_v}天，立即决策展期或平仓！" + rec
        elif dte_v <= _DTE_REVIEW:
            rec = f"⚠️ DTE={dte_v}天，" + rec

        return {
            "id": f"{und}_{stype}_{ll['expiry']}",
            "type": stype, "underlying": und, "direction": direction,
            "legs": [dict(li=li, **ll), dict(li=si, **sl)],
            "spread_qty": qty,
            "expiry": ll["expiry"], "dte": dte_v,
            "low_strike": low_s, "high_strike": high_s, "strike_width": width,
            "is_debit": is_debit,
            "net_per_share": round(net_ps, 4), "net_total": net_tot,
            "max_profit": max_profit, "max_loss": max_loss,
            "breakeven": breakeven,
            "current_pnl": round(pnl, 2),
            "pnl_pct": pnl_pct,
            "risk_level": risk_level,
            "recommendation": rec,
        }

    def _make_diagonal(li: int, si: int) -> dict:
        ll = legs[li]
        sl = legs[si]
        qty = min(abs(ll["qty"]), abs(sl["qty"]))
        direction = ll["direction"]
        und = ll["underlying"]

        # Identify near/far by expiry
        if ll["expiry"] < sl["expiry"]:
            near_leg, far_leg = ll, sl
        else:
            near_leg, far_leg = sl, ll

        near_is_short = near_leg["qty"] < 0
        far_is_long   = far_leg["qty"]  > 0
        is_proper     = near_is_short and far_is_long

        same_strike = abs(near_leg["strike"] - far_leg["strike"]) < 0.001
        if is_proper:
            stype = "Calendar Spread" if same_strike else "Diagonal Spread (LEAPS)"
        else:
            stype = "Reversed Diagonal (⚠️ 买腿先到期)"

        # net cost: positive = debit
        if is_proper:
            net_ps  = far_leg["unit_cost"] - near_leg["unit_cost"]
        else:
            net_ps  = near_leg["unit_cost"] - far_leg["unit_cost"]
        net_tot = round(net_ps * qty * 100, 2)

        long_frac  = qty / abs(ll["qty"])
        short_frac = qty / abs(sl["qty"])
        pnl = _leg_pnl(ll, long_frac) + _leg_pnl(sl, short_frac)

        near_dte = near_leg["dte"]
        if near_dte <= _DTE_CRITICAL:
            risk_level = "CRITICAL"
            rec = (f"🚨 近月腿 DTE={near_dte}天 — 立即决策展期或平仓！"
                   f"当前盈亏 ${pnl:+,.0f}")
        elif not is_proper:
            risk_level = "HIGH"
            rec = (f"⚠️ 买腿（{near_leg['expiry']}）早于卖腿（{far_leg['expiry']}）到期，"
                   f"买腿失效后卖腿变裸空仓！当前盈亏 ${pnl:+,.0f}")
        elif near_dte <= _DTE_REVIEW:
            risk_level = "HIGH"
            rec = (f"⚠️ 近月腿 DTE={near_dte}天，制定展期计划；"
                   f"当前盈亏 ${pnl:+,.0f}")
        else:
            risk_level = "MEDIUM"
            rec = (f"对角价差持有中，Theta 时间优势在近月卖腿；"
                   f"当前盈亏 ${pnl:+,.0f}，净成本 ${abs(net_tot):,.0f}")

        # 正向对角价差（买远月+卖近月）= 净权利金支出封顶最大亏损
        diag_max_loss   = round(abs(net_tot), 2) if (is_proper and net_ps > 0) else None
        diag_pnl_pct    = round(pnl / diag_max_loss * 100, 1) if diag_max_loss else None

        return {
            "id": f"{und}_{stype}_{near_leg['expiry']}_vs_{far_leg['expiry']}",
            "type": stype, "underlying": und, "direction": direction,
            "legs": [dict(li=li, **ll), dict(li=si, **sl)],
            "spread_qty": qty,
            "expiry": near_leg["expiry"], "dte": near_dte,
            "near_expiry": near_leg["expiry"], "far_expiry": far_leg["expiry"],
            "near_strike": near_leg["strike"], "far_strike": far_leg["strike"],
            "is_proper": is_proper,
            "net_per_share": round(net_ps, 4), "net_total": net_tot,
            "max_profit": None, "max_loss": diag_max_loss, "breakeven": None,
            "current_pnl": round(pnl, 2), "pnl_pct": diag_pnl_pct,
            "risk_level": risk_level,
            "recommendation": rec,
        }

    def _make_naked(idx: int) -> dict:
        leg = legs[idx]
        qty = leg["qty"]
        is_long  = qty > 0
        direction = leg["direction"]
        und = leg["underlying"]
        dte_v = leg["dte"]

        stype = f"Naked {'Long' if is_long else 'Short'} {direction.capitalize()}"
        pnl   = _leg_pnl(leg)
        if is_long:
            max_loss_naked = round(leg["unit_cost"] * abs(qty) * 100, 2)
        elif direction == "put":
            # 裸卖 Put：最大亏损 = 行权价×100×手数（标的跌至 0）
            max_loss_naked = round(leg["strike"] * abs(qty) * 100, 2)
        else:
            max_loss_naked = None  # 裸卖 Call：理论无限亏损

        if is_long:
            risk_level = "MEDIUM" if dte_v > _DTE_REVIEW else ("CRITICAL" if dte_v <= _DTE_CRITICAL else "HIGH")
            rec = f"买权持有，最大亏损权利金 ${max_loss_naked:,.0f}，当前盈亏 ${pnl:+,.0f}"
        elif direction == "put":
            risk_level = "CRITICAL" if dte_v <= _DTE_CRITICAL else "HIGH"
            rec = f"⚠️ 裸卖 Put，最大亏损 ${max_loss_naked:,.0f}（标的归零），当前盈亏 ${pnl:+,.0f}"
        else:
            risk_level = "CRITICAL" if dte_v <= _DTE_CRITICAL else "HIGH"
            rec = f"⚠️ 裸卖 Call，风险无限，当前盈亏 ${pnl:+,.0f}"

        if dte_v <= _DTE_CRITICAL:
            rec = f"🚨 DTE={dte_v}天 — 立即处理！" + rec
        elif dte_v <= _DTE_REVIEW:
            rec = f"⚠️ DTE={dte_v}天 — " + rec

        return {
            "id": f"{und}_{stype}_{leg['expiry']}",
            "type": stype, "underlying": und, "direction": direction,
            "legs": [dict(li=idx, **leg)],
            "spread_qty": abs(qty),
            "expiry": leg["expiry"], "dte": dte_v,
            "net_per_share": leg["unit_cost"], "net_total": leg["unit_cost"] * abs(qty) * 100,
            "max_profit": None, "max_loss": max_loss_naked, "breakeven": None,
            "current_pnl": round(pnl, 2), "pnl_pct": None,
            "risk_level": risk_level,
            "recommendation": rec,
        }

    # ── Main matching loop ────────────────────────────────────────
    und_groups: dict[str, list[int]] = {}
    for i, leg in enumerate(legs):
        und_groups.setdefault(leg["underlying"], []).append(i)

    for und, und_indices in sorted(und_groups.items()):

        # Round 1: Same expiry + same direction → vertical spreads
        exp_dir: dict[tuple, list[int]] = {}
        for i in und_indices:
            if matched[i]:
                continue
            exp_dir.setdefault((legs[i]["expiry"], legs[i]["direction"]), []).append(i)

        for (_exp, _dir), group in sorted(exp_dir.items()):
            longs  = [i for i in group if legs[i]["qty"] > 0]
            shorts = [i for i in group if legs[i]["qty"] < 0]
            for li in longs:
                if matched[li]:
                    continue
                for si in shorts:
                    if matched[si]:
                        continue
                    port = _make_vertical(li, si)
                    portfolios.append(port)
                    sq = port["spread_qty"]
                    rem_l = abs(legs[li]["qty"]) - sq
                    rem_s = abs(legs[si]["qty"]) - sq
                    if rem_l <= 0:
                        matched[li] = True
                    else:
                        legs[li]["qty"] = rem_l
                    if rem_s <= 0:
                        matched[si] = True
                    else:
                        legs[si]["qty"] = -rem_s
                    break

        # Round 2: Cross-expiry → diagonal / calendar
        dir_groups: dict[str, list[int]] = {}
        for i in und_indices:
            if matched[i]:
                continue
            dir_groups.setdefault(legs[i]["direction"], []).append(i)

        for _dir, group in dir_groups.items():
            longs  = sorted([i for i in group if legs[i]["qty"] > 0], key=lambda i: legs[i]["expiry"])
            shorts = sorted([i for i in group if legs[i]["qty"] < 0], key=lambda i: legs[i]["expiry"])
            for li in longs:
                if matched[li]:
                    continue
                for si in shorts:
                    if matched[si]:
                        continue
                    if legs[li]["expiry"] == legs[si]["expiry"]:
                        continue
                    port = _make_diagonal(li, si)
                    portfolios.append(port)
                    sq = port["spread_qty"]
                    rem_l = abs(legs[li]["qty"]) - sq
                    rem_s = abs(legs[si]["qty"]) - sq
                    if rem_l <= 0:
                        matched[li] = True
                    else:
                        legs[li]["qty"] = rem_l
                    if rem_s <= 0:
                        matched[si] = True
                    else:
                        legs[si]["qty"] = -rem_s
                    break

        # Round 3: Remaining → naked
        for i in und_indices:
            if not matched[i] and abs(legs[i]["qty"]) > 0:
                portfolios.append(_make_naked(i))
                matched[i] = True

    # Sort: CRITICAL → HIGH → MEDIUM → LOW
    _RISK_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
    portfolios.sort(key=lambda p: (_RISK_ORDER.get(p["risk_level"], 9), p["underlying"]))
    return portfolios


def _compute_exit_analysis(acct_id: str) -> dict:
    """
    Rich layered exit analysis: wraps _build_spread_portfolios and adds
    risk / time / direction layers for each portfolio, sorted by urgency.
    Returns {portfolios: [...], summary: {...}}.
    """
    portfolios = _build_spread_portfolios(acct_id)
    if not portfolios:
        return {"portfolios": [], "summary": {}}

    bal        = _load_latest_balance(acct_id)
    net_equity = float(bal.get("total_equity") or 0)

    underlyings = {p["underlying"] for p in portfolios}
    und_prices  = _fetch_underlying_prices(tuple(sorted(underlyings))) if underlyings else {}

    today = datetime.date.today()
    enriched_list = []

    for port in portfolios:
        und       = port["underlying"]
        und_price = und_prices.get(und)

        # ── Cost basis & equity % ────────────────────────────────
        cost_basis = abs(port.get("max_loss") or port.get("net_total") or 0)
        equity_pct = round(cost_basis / net_equity * 100, 1) if net_equity > 0 else 0.0

        # ── P&L % (fill in if missing) ───────────────────────────
        pnl_pct = port.get("pnl_pct")
        if pnl_pct is None and cost_basis > 0.01:
            pnl_pct = round(port["current_pnl"] / cost_basis * 100, 1)

        # ── DTE info ─────────────────────────────────────────────
        min_dte     = port["dte"]
        short_legs  = [l for l in port["legs"] if (l.get("qty") or 0) < 0]
        short_dtes  = [l["dte"] for l in short_legs]
        min_short_dte = min(short_dtes) if short_dtes else None
        has_short     = bool(short_legs)

        # ── Thesis broken detection (uses yfinance price) ────────
        thesis_broken = False
        thesis_note   = ""
        if und_price:
            ptype = port.get("type", "")
            high_k = port.get("high_strike") or 0
            low_k  = port.get("low_strike")  or 0
            if "Bear Put" in ptype and high_k and und_price > high_k:
                thesis_broken = True
                otm = (und_price - high_k) / high_k * 100
                thesis_note = (f"{und} 现价 ${und_price:.2f} 高于价差上沿 "
                               f"${high_k:.0f}（超出 {otm:.1f}%），看跌假设已被推翻")
            elif "Bull Call" in ptype and low_k and und_price < low_k:
                thesis_broken = True
                otm = (low_k - und_price) / low_k * 100
                thesis_note = (f"{und} 现价 ${und_price:.2f} 低于价差下沿 "
                               f"${low_k:.0f}（偏离 {otm:.1f}%），看涨假设受挫")
            elif "Bear Call" in ptype and high_k and und_price > high_k:
                thesis_broken = True
                otm = (und_price - high_k) / high_k * 100
                thesis_note = (f"{und} 现价 ${und_price:.2f} 高于上沿 "
                               f"${high_k:.0f}（超出 {otm:.1f}%），空头承压")
            elif "Bull Put" in ptype and low_k and und_price < low_k:
                thesis_broken = True
                otm = (low_k - und_price) / low_k * 100
                thesis_note = (f"{und} 现价 ${und_price:.2f} 低于下沿 "
                               f"${low_k:.0f}（偏离 {otm:.1f}%），多头压力加大")
            elif "Naked Long Put" in ptype:
                strike_k = port["legs"][0].get("strike") or 0
                if strike_k and und_price > strike_k * 1.1:
                    otm = (und_price - strike_k) / strike_k * 100
                    thesis_broken = True
                    thesis_note = (f"{und} 现价 ${und_price:.2f} 高于行权价 "
                                   f"${strike_k:.0f}（{otm:.0f}% OTM），看跌假设未兑现")
            elif "Naked Long Call" in ptype:
                strike_k = port["legs"][0].get("strike") or 0
                if strike_k and und_price < strike_k * 0.9:
                    otm = (strike_k - und_price) / strike_k * 100
                    thesis_broken = True
                    thesis_note = (f"{und} 现价 ${und_price:.2f} 低于行权价 "
                                   f"${strike_k:.0f}（{otm:.0f}% OTM），看涨动能不足")

        # ── Urgency score (for sorting) ──────────────────────────
        urgency = 0
        if min_dte is not None:
            if min_dte <= 7:    urgency += 5
            elif min_dte <= 14: urgency += 3
            elif min_dte <= 21: urgency += 1
        if pnl_pct is not None:
            if pnl_pct <= -60:  urgency += 6   # extreme loss → highest priority
            elif pnl_pct <= -45: urgency += 4
            elif pnl_pct <= -30: urgency += 2
            elif pnl_pct >= 45:  urgency += 2
            elif pnl_pct >= 30:  urgency += 1
        if equity_pct > 30: urgency += 3
        elif equity_pct > 20: urgency += 1
        if thesis_broken:   urgency += 4

        # ── Action label & color ─────────────────────────────────
        _p = pnl_pct or 0
        if min_dte is not None and min_dte <= 7:
            action, action_color = "🚨 立即处理", "#FF4B4B"
        elif _p <= -60:
            # Extreme loss: stop regardless of DTE
            action, action_color = "🛑 止损", "#FF4B4B"
        elif thesis_broken and _p <= -20:
            action, action_color = "📉 重新评估", "#FF4B4B"
        elif pnl_pct is not None and pnl_pct >= 50:
            action, action_color = "⚡ 止盈", "#00C853"
        elif pnl_pct is not None and pnl_pct <= -50 and (min_dte or 999) < 30:
            action, action_color = "🛑 止损", "#FF4B4B"
        elif has_short and min_short_dte is not None and min_short_dte <= 21:
            action, action_color = "🔄 滚仓", "#FFB700"
        elif thesis_broken:
            action, action_color = "⚠️ 方向反转", "#FFB700"
        elif pnl_pct is not None and pnl_pct <= -40:
            action, action_color = "👀 关注", "#FFB700"
        else:
            action, action_color = "✅ 持有", "#6B6B6B"

        # ── Why text (layered explanation) ───────────────────────
        why_parts = []

        if thesis_broken and thesis_note:
            why_parts.append(f"【方向】{thesis_note}")

        if pnl_pct is not None:
            dist_stop = pnl_pct - (-50)
            dist_tp   = 50 - pnl_pct
            if pnl_pct >= 50:
                why_parts.append(f"【盈亏】已触发止盈线 +50%，建议锁利或展期")
            elif pnl_pct <= -50:
                why_parts.append(f"【盈亏】已触发止损线，亏损 {abs(pnl_pct):.0f}%")
            elif dist_stop < 15:
                why_parts.append(f"【盈亏】亏损 {pnl_pct:+.0f}%，距止损线 -50% 仅剩 {dist_stop:.0f}%，需密切关注")
            elif dist_tp < 12:
                why_parts.append(f"【盈亏】盈利 {pnl_pct:+.0f}%，距止盈线 +50% 还差 {dist_tp:.0f}%")
            else:
                why_parts.append(f"【盈亏】{pnl_pct:+.0f}%（止盈 +50% / 止损 -50%，当前安全区间）")

        if equity_pct > 25:
            why_parts.append(f"【风险】持仓成本占净值 {equity_pct:.0f}%，集中度偏高（建议单仓 ≤25%净值）")
        elif equity_pct > 15:
            why_parts.append(f"【风险】持仓成本占净值 {equity_pct:.0f}%")

        if min_dte is not None:
            if min_dte <= 7:
                why_parts.append(f"【时间】DTE={min_dte}天，时间价值极速衰减，立即决策")
            elif min_dte <= 14:
                why_parts.append(f"【时间】DTE={min_dte}天，Theta加速衰减，建议本周决策")
            elif min_dte <= 21:
                why_parts.append(f"【时间】DTE={min_dte}天，建议2周内决策")

        if not why_parts:
            why_parts.append("各项指标正常，无需立即行动")

        enriched_list.append({
            **port,
            "cost_basis":     cost_basis,
            "equity_pct":     equity_pct,
            "pnl_pct":        pnl_pct,
            "min_dte":        min_dte,
            "min_short_dte":  min_short_dte,
            "has_short":      has_short,
            "und_price":      und_price,
            "thesis_broken":  thesis_broken,
            "thesis_note":    thesis_note,
            "urgency":        urgency,
            "action":         action,
            "action_color":   action_color,
            "why":            " ；".join(why_parts),
        })

    enriched_list.sort(key=lambda x: -x["urgency"])

    # ── Portfolio-level summary ──────────────────────────────────
    total_cost = sum(p["cost_basis"] for p in enriched_list)
    cost_pct   = round(total_cost / net_equity * 100, 1) if net_equity > 0 else 0.0

    by_und: dict[str, float] = {}
    for p in enriched_list:
        by_und[p["underlying"]] = by_und.get(p["underlying"], 0) + p["cost_basis"]

    top_unds = sorted(by_und.items(), key=lambda x: -x[1])[:3]

    return {
        "portfolios": enriched_list,
        "summary": {
            "total_cost": round(total_cost, 2),
            "net_equity": net_equity,
            "cost_pct":   cost_pct,
            "top_unds":   top_unds,
            "n_broken":   sum(1 for p in enriched_list if p["thesis_broken"]),
        },
    }


def _load_positions_xlsx(acct_id: str) -> pd.DataFrame:
    """读取 data/firstrade/ 下最新的 xlsx 持仓文件，支持中英文列名。"""
    xlsx_files = sorted(_FT_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not xlsx_files:
        return pd.DataFrame()

    latest = xlsx_files[0]
    try:
        df = pd.read_excel(latest, engine="openpyxl")
    except Exception as e:
        _log.warning(f"xlsx read error {latest}: {e}")
        return pd.DataFrame()

    df = df.dropna(how="all")
    if df.empty:
        return pd.DataFrame()

    # 列名映射（中文 → 英文）
    ZH_MAP = {
        "代号": "symbol", "股票代号": "symbol", "ticker": "symbol",
        "数量": "quantity", "股数": "quantity",
        "市值": "market_value", "总市值": "market_value",
        "益损$": "unrealized_pnl", "盈亏$": "unrealized_pnl", "盈亏": "unrealized_pnl",
        "益损%": "unrealized_pnl_pct", "盈亏%": "unrealized_pnl_pct",
        "成本": "cost_basis", "总成本": "cost_basis", "平均成本": "unit_cost",
        "现价": "current_price", "股价": "current_price",
        "描述": "description", "名称": "description",
    }
    df = df.rename(columns=lambda c: ZH_MAP.get(str(c).strip(), str(c).strip().lower()))

    if "symbol" not in df.columns:
        return pd.DataFrame()

    df["symbol"] = df["symbol"].astype(str).str.strip().str.upper()
    df = df[df["symbol"].notna() & (df["symbol"] != "") & (df["symbol"] != "NAN")]

    # 标注持仓类型
    df["position_type"] = df["symbol"].apply(
        lambda s: "option" if _OCC_RE.match(s) else "stock"
    )

    _log.info(f"xlsx {latest.name}: {len(df)} rows from {latest}")
    return df


def _import_from_xlsx_file(path: pathlib.Path, acct_id: str) -> tuple[int, int]:
    """
    读取 Firstrade 持仓 xlsx，写入 positions 表。
    返回 (stocks_count, options_count)。
    """
    ZH_MAP = {
        "代号": "symbol", "股票代号": "symbol", "ticker": "symbol",
        "数量": "quantity", "股数": "quantity",
        "市值": "market_value", "总市值": "market_value",
        "益损$": "unrealized_pnl", "盈亏$": "unrealized_pnl",
        "益损 $": "unrealized_pnl", "益损 %": "unrealized_pnl_pct",
        "益损%": "unrealized_pnl_pct", "盈亏%": "unrealized_pnl_pct",
        "成本": "cost_basis", "总成本": "cost_basis",
        "单位成本": "unit_cost", "价格": "current_price",
        "现价": "current_price", "详细说明": "description",
        "当日益损$": "day_pnl",
    }
    try:
        df = pd.read_excel(path, engine="openpyxl")
    except Exception as e:
        _log.warning(f"xlsx import error {path}: {e}")
        return 0, 0

    df = df.dropna(how="all")
    df = df.rename(columns=lambda c: ZH_MAP.get(str(c).strip(), str(c).strip().lower()))
    if "symbol" not in df.columns:
        return 0, 0

    df["symbol"] = df["symbol"].astype(str).str.strip().str.upper()
    df = df[df["symbol"].notna() & (df["symbol"] != "") & (df["symbol"] != "NAN")]
    df["position_type"] = df["symbol"].apply(
        lambda s: "option" if _OCC_RE.match(s) else "stock"
    )

    def _pf(v):
        return _parse_money(v)

    now = datetime.datetime.now(_ET).isoformat()
    conn = _db()
    conn.execute("DELETE FROM positions WHERE account_id=?", (acct_id,))
    stocks_n = options_n = 0
    for _, row in df.iterrows():
        sym = row.get("symbol")
        if not sym:
            continue
        pos_type = row.get("position_type", "stock")
        conn.execute("""
            INSERT INTO positions
              (account_id, sync_time, symbol, position_type, quantity,
               cost_basis, market_value, unrealized_pnl, unrealized_pnl_pct, description)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (acct_id, now, sym, pos_type,
              _pf(row.get("quantity")), _pf(row.get("cost_basis")),
              _pf(row.get("market_value")), _pf(row.get("unrealized_pnl")),
              _pf(row.get("unrealized_pnl_pct")), str(row.get("description","")).strip()))
        if pos_type == "stock":
            stocks_n += 1
        else:
            options_n += 1
    conn.commit(); conn.close()
    return stocks_n, options_n


def _save_stock_positions(acct_id: str, rows: list[dict]) -> int:
    """
    将手动录入的股票持仓写入 positions 表（权威数据源）。
    计算 cost_basis / market_value / pnl 后 upsert。
    返回保存行数。
    """
    now = datetime.datetime.now(_ET).isoformat()
    conn = _db()
    conn.execute("DELETE FROM positions WHERE account_id=? AND position_type='stock'", (acct_id,))
    count = 0
    for r in rows:
        sym = str(r.get("symbol", "")).strip().upper()
        if not sym:
            continue
        qty  = float(r.get("quantity")     or 0)
        uc   = float(r.get("unit_cost")    or 0)
        cp   = float(r.get("current_price") or 0) or None
        desc = str(r.get("description", "")).strip()
        cost_basis   = round(qty * uc, 2)       if qty and uc else None
        market_value = round(qty * cp, 2)        if qty and cp else None
        pnl          = round(market_value - cost_basis, 2) if (market_value and cost_basis) else None
        pnl_pct      = round(pnl / cost_basis * 100, 2)    if (pnl is not None and cost_basis) else None
        conn.execute("""
            INSERT INTO positions
              (account_id, sync_time, symbol, position_type, quantity, unit_cost,
               cost_basis, current_price, market_value, unrealized_pnl, unrealized_pnl_pct, description)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (acct_id, now, sym, "stock", qty, uc, cost_basis, cp,
              market_value, pnl, pnl_pct, desc))
        count += 1
    conn.commit()
    conn.close()
    return count


def _refresh_stock_prices(acct_id: str) -> dict:
    """
    用 yfinance 刷新 positions 表中股票的现价、市值、浮盈亏。
    返回 {'updated': n, 'failed': [...]}
    """
    try:
        import yfinance as yf
    except ImportError:
        return {"updated": 0, "failed": [], "error": "yfinance not installed"}

    conn = _db()
    rows = conn.execute(
        "SELECT symbol, quantity, cost_basis FROM positions "
        "WHERE account_id=? AND position_type='stock'",
        (acct_id,)).fetchall()
    conn.close()

    if not rows:
        return {"updated": 0, "failed": []}

    syms   = [r[0] for r in rows]
    qty_map  = {r[0]: r[1] for r in rows}
    cost_map = {r[0]: r[2] for r in rows}

    # Batch fetch
    try:
        tickers = yf.Tickers(" ".join(syms))
    except Exception as e:
        return {"updated": 0, "failed": syms, "error": str(e)}

    updated = 0
    failed  = []
    conn    = _db()
    for sym in syms:
        try:
            px = tickers.tickers[sym].fast_info["lastPrice"]
            if not px:
                raise ValueError("no price")
            qty  = qty_map[sym] or 0
            cost = cost_map[sym] or 0
            mv   = round(px * qty, 2)
            pnl  = round(mv - cost, 2)
            pnl_pct = round((pnl / cost * 100) if cost else 0, 2)
            conn.execute("""
                UPDATE positions
                SET current_price=?, market_value=?, unrealized_pnl=?,
                    unrealized_pnl_pct=?, sync_time=?
                WHERE account_id=? AND symbol=? AND position_type='stock'
            """, (round(px, 4), mv, pnl, pnl_pct,
                  datetime.datetime.now(_ET).isoformat(), acct_id, sym))
            updated += 1
        except Exception:
            failed.append(sym)
    conn.commit()
    conn.close()
    return {"updated": updated, "failed": failed}


def _expiry_badge(expiry_str: str) -> str:
    """返回带颜色 emoji 的到期天数标注。"""
    try:
        days = (datetime.date.fromisoformat(expiry_str) - datetime.date.today()).days
        if days < 0:   return f"🔴 已到期"
        if days < 7:   return f"🔴 {days}天"
        if days < 30:  return f"🟠 {days}天"
        return f"🟢 {days}天"
    except Exception:
        return ""


# ════════════════════════════════════════════════════════
# 共享监控状态（跨 rerun 持久化）
# ════════════════════════════════════════════════════════
def _import_positions_csv(df: pd.DataFrame, acct_id: str) -> int:
    return _account_import_positions_csv(
        df,
        acct_id,
        save_positions=_save_positions,
        save_balance=_save_balance,
    )


def _import_transactions_csv(df: pd.DataFrame, acct_id: str) -> int:
    return _account_import_transactions_csv(
        df,
        acct_id,
        save_transactions=_save_transactions,
    )


def _process_csv_file(src: pathlib.Path, acct_id: str = "account_1") -> dict:
    return _account_process_csv_file(
        src,
        acct_id=acct_id,
        latest_csv=_LATEST_CSV,
        save_positions=_save_positions,
        save_transactions=_save_transactions,
        save_balance=_save_balance,
        logger=_log,
    )


@st.cache_resource
def _watch_state() -> dict:
    return {
        "last_file":   None,       # 最近处理的文件名
        "last_time":   None,       # 处理时间 (datetime)
        "last_type":   None,       # positions / transactions
        "last_rows":   0,
        "new_data":    False,      # watchdog 检测到新文件时置 True
        "errors":      [],
    }


# ════════════════════════════════════════════════════════
# Watchdog 文件监听器
# ════════════════════════════════════════════════════════
class _ExportCsvHandler(FileSystemEventHandler if _WATCHDOG_OK else object):
    """监听 Downloads 文件夹，检测 export*.csv 新文件。"""

    def _handle(self, path: str):
        p = pathlib.Path(path)
        if p.suffix.lower() != ".csv":
            return
        if not p.stem.lower().startswith("export"):
            return
        # 只处理今日文件（避免处理旧文件）
        try:
            mdate = datetime.date.fromtimestamp(p.stat().st_mtime)
        except OSError:
            return
        if mdate != datetime.date.today():
            return
        _log.info(f"Detected export CSV: {p.name}")
        try:
            result = _process_csv_file(p)
            ws = _watch_state()
            ws["last_file"] = result.get("file")
            ws["last_time"] = datetime.datetime.now(_ET)
            ws["last_type"] = result.get("type")
            ws["last_rows"] = result.get("rows", 0)
            ws["new_data"]  = True
            if not result.get("ok"):
                ws["errors"].append(result.get("reason", "unknown error"))
        except Exception as e:
            _log.exception(f"CSV import failed: {e}")
            _watch_state()["errors"].append(str(e))

    def on_created(self, event):
        if not event.is_directory:
            time.sleep(0.5)   # 等文件写完
            self._handle(event.src_path)

    def on_moved(self, event):
        if not event.is_directory:
            self._handle(event.dest_path)

    def on_modified(self, event):
        if not event.is_directory:
            self._handle(event.src_path)


@st.cache_resource
def _start_watcher():
    if not _WATCHDOG_OK:
        return None
    handler  = _ExportCsvHandler()
    observer = Observer()
    observer.schedule(handler, str(_DOWNLOADS), recursive=False)
    observer.daemon = True
    observer.start()
    _log.info(f"Watchdog started → {_DOWNLOADS}")
    return observer


# ════════════════════════════════════════════════════════
# CDP 自动化（附着已登录 Chrome，端口 9222）
# ════════════════════════════════════════════════════════
try:
    from selenium import webdriver as _wd
    from selenium.webdriver.chrome.options import Options as _ChromeOpts
    _SEL_OK = True
except ImportError:
    _SEL_OK = False

_CDP_ADDR         = "localhost:9222"
_FT_BALANCE_URL   = "https://invest.firstrade.com/app/balance"
_FT_HISTORY_URL   = "https://invest.firstrade.com/app/history"
_FT_POSITIONS_URL = "https://invest.firstrade.com/app/positions"
_POSITIONS_WRITE  = True    # 每次同步后自动 UPSERT options_positions

# JS：从页面纯文本提取余额数字
# 多标签兜底：简体中文 / 繁体中文 / 英文，适配 Firstrade 不同语言界面
# execute_script 需要顶层 return，IIFE 内 return 不被 selenium 捕获
_BALANCE_JS = r"""
var t = document.body ? (document.body.innerText || '') : '';
// allowZero=true：允许返回 0（如 Margin Balance=$0.00）
// minVal：返回值必须 >= minVal（过滤描述文字中的杂音数字）
function findNth(label, n, allowZero, minVal) {
  var idx = t.indexOf(label);
  if (idx < 0) return null;
  var s = t.slice(idx, idx + 500);
  var re = /[+\-]?\$?\s*[\d,]+\.?\d{0,2}/g;
  var found = [], m;
  while ((m = re.exec(s)) !== null) {
    var v = parseFloat(m[0].replace(/[$,\s]/g, ''));
    var ok = !isNaN(v) && (allowZero ? true : Math.abs(v) > 0.001);
    if (ok && (minVal === undefined || Math.abs(v) >= minVal)) {
      found.push(v);
      if (found.length >= n) break;
    }
  }
  return found.length >= n ? found[n - 1] : null;
}
function findFirst(labels, n, allowZero, minVal) {
  for (var i = 0; i < labels.length; i++) {
    var v = findNth(labels[i], n, allowZero, minVal);
    if (v !== null) return v;
  }
  return null;
}
var day_pnl = findFirst([
  '今日盈亏','今日益損','当日盈亏','日盈亏',
  "Today's Gain/Loss","Today's Change","Day's Gain/Loss",
  "Daily Gain/Loss","Daily Change","Day Change","Change Today"
], 1);
if (day_pnl === null)
  day_pnl = findFirst(['账户总值','帳戶總值','Account Value','Net Account Value'], 2);
// margin_used：Margin Balance 可以是 $0.00，allowZero=true 避免跳过零值
// minVal=undefined（不设下限），n=1（取第一个数字，即 Margin Balance 本身）
var margin_used = findFirst(
  ['融资结余','融資結餘','Margin Balance','Margin Debit Balance','Margin Used'], 1, true);
// margin_available：Cash Buying Power 描述里有 "(Stocks below $3)"，$3 是噪音
// 用 minVal=10 过滤小数字，取第一个 >=10 的值
var margin_available = findFirst(
  ['融资购买力','融資購買力','Cash Buying Power','Margin Buying Power',
   'Available Margin','Buying Power'], 1, false, 10);
return {
  total_equity:     findFirst(['账户总值','帳戶總值','Total Account Value','Account Value','Net Account Value'], 1),
  day_pnl:          day_pnl,
  cash_balance:     findFirst(['现金结余','現金結餘','Cash Balance','Cash & Cash Equivalents','Cash'], 1),
  margin_used:      margin_used,
  margin_available: margin_available
};
"""


@st.cache_resource
def _sync_state() -> dict:
    """跨 rerun 持久化的同步状态字典（供调度线程和 UI 共享）。"""
    return {
        "last_time":   None,   # 最近成功同步的 datetime
        "last_status": "idle", # idle / running / ok / partial / error / no_chrome
        "last_error":  "",
        "chrome_ok":   False,
        "bal_ok":      False,
        "csv_ok":      False,
        "positions_diff": None,  # 最近一次持仓对比结果 dict
    }


# ── 持仓页抓取 JS（四策略兜底）────────────────────────────
_POSITIONS_JS = r"""
return (function() {
    var OCC_RE = /([A-Z]{1,6}\d{6}[CP]\d{8})/i;
    var rows = [], strategy = 'none';

    // 策略 SK: SvelteKit 内联数据（首选）
    // Firstrade 用 SvelteKit，持仓数据内嵌在 <script> 的 resolve() 回调里，
    // 格式: {...,quantity:N,...,symbol:"OCC...",logo:{...长URL...},unitCost:N,secType:2,...}
    // logo URL 可能超过 800 chars，所以从 secType:2 往前搜索 2500 chars 找 OCC 符号和所有字段。
    try {
        var allScript = Array.from(
            document.querySelectorAll('script:not([src])')
        ).map(function(s){ return s.textContent || ''; }).join('\n');

        var stRe = /secType:2/g;
        var stm;
        var seenSK = {};
        while ((stm = stRe.exec(allScript)) !== null) {
            // 以 secType:2 为锚，向前取 2500 chars（覆盖整个 item 对象）
            var ctx = allScript.slice(Math.max(0, stm.index - 2500), stm.index + 200);
            var occm = /([A-Z]{1,6}\d{6}[CP]\d{8})/i.exec(
                ctx.slice(ctx.lastIndexOf('{'))  // 只在最近的 { 之后找
            );
            if (!occm) {
                // fallback：全 ctx 搜索
                occm = /([A-Z]{1,6}\d{6}[CP]\d{8})/i.exec(ctx);
            }
            if (!occm) continue;
            var sym = occm[1].toUpperCase();
            if (seenSK[sym]) continue;
            var qm  = /quantity:(-?\d+)/.exec(ctx);
            var lm  = /(?:^|,)last:([\d.]+)/.exec(ctx);
            var ucm = /unitCost:([\d.]+)/.exec(ctx);
            var glm = /gainloss:(-?[\d.]+)/.exec(ctx);
            if (!qm) continue;
            seenSK[sym] = true;
            rows.push({
                s:        'SK',
                sym:      sym,
                qty:      parseInt(qm[1]),
                last:     lm  ? parseFloat(lm[1])  : null,
                cost:     ucm ? parseFloat(ucm[1]) : null,
                gainloss: glm ? parseFloat(glm[1]) : null,
                cells:    []
            });
            strategy = 'SK';
        }
    } catch(e) {}

    // 策略 A: <table> 标准表格
    if (!rows.length) {
        document.querySelectorAll('table tr').forEach(function(tr) {
            var cells = Array.from(tr.querySelectorAll('td,th')).map(
                function(c){ return (c.innerText||'').replace(/\s+/g,' ').trim(); });
            var sym = '';
            cells.forEach(function(c){
                var mm = OCC_RE.exec(c.replace(/\s/g,'').toUpperCase());
                if (mm && !sym) sym = mm[1];
            });
            if (sym) { rows.push({s:'A', sym:sym, cells:cells}); strategy='A'; }
        });
    }

    // 策略 B: ARIA role=row / role=gridrow
    if (!rows.length) {
        document.querySelectorAll('[role=row],[role=gridrow]').forEach(function(r) {
            var cells = Array.from(r.querySelectorAll('[role=cell],[role=gridcell]')).map(
                function(c){ return (c.innerText||'').replace(/\s+/g,' ').trim(); });
            var sym = '';
            cells.forEach(function(c){
                var mm = OCC_RE.exec(c.replace(/\s/g,'').toUpperCase());
                if (mm && !sym) sym = mm[1];
            });
            if (sym) { rows.push({s:'B', sym:sym, cells:cells}); strategy='B'; }
        });
    }

    // 策略 C: innerText 逐行扫描（SPA 懒加载兜底）
    if (!rows.length) {
        var lines = (document.body ? document.body.innerText : '').split('\n');
        lines.forEach(function(ln, i) {
            var clean = ln.replace(/\s/g,'').toUpperCase();
            var mm = OCC_RE.exec(clean);
            if (mm) {
                var ctx2 = lines.slice(Math.max(0,i-3), i+4).map(function(l){return l.trim();});
                rows.push({s:'C', sym:mm[1], cells:ctx2});
                strategy = 'C';
            }
        });
    }

    // 调试：含期权关键词的页面文本样本
    var body_text = document.body ? document.body.innerText : '';
    var opt_sample = body_text.split('\n').filter(function(l){
        return /call|put|期权|option|strike|expir|行权|到期/i.test(l);
    }).slice(0,20).map(function(l){return l.trim();});

    return {
        url:        window.location.href,
        title:      document.title,
        body_len:   body_text.length,
        strategy:   strategy,
        rows:       rows,
        opt_sample: opt_sample
    };
})()
"""

def _parse_scraped_rows(raw: dict, acct_id: str) -> list[dict]:
    """JS 抓取结果 → 标准 options_positions 行列表（不含 Greeks）。"""
    results: list[dict] = []
    seen: set[str] = set()
    for row in raw.get("rows", []):
        sym = (row.get("sym") or "").upper().replace(" ", "")
        if not sym or sym in seen:
            continue
        occ = _parse_occ_sym(sym)
        if not occ:
            continue

        strategy = row.get("s", "?")

        if strategy == "SK":
            # SvelteKit 内联数据：字段直接可用
            qty_raw       = row.get("qty")
            unit_cost     = row.get("cost")
            current_price = row.get("last")
            total_pnl     = row.get("gainloss")
            qty           = int(qty_raw) if qty_raw is not None else None
        else:
            # 旧策略（A/B/C）：从 cells 文本中解析数字
            cells = row.get("cells", [])
            nums: list[float] = []
            for c in cells:
                c2 = re.sub(r'[A-Z]{1,6}\d{6}[CP]\d{8}', '', c, flags=re.I)
                for tok in re.findall(r'-?\d+\.?\d*', c2.replace(",", "")):
                    try:
                        v = float(tok)
                        if 0.001 < abs(v) < 1e7:
                            nums.append(v)
                    except ValueError:
                        pass
            qty = None
            for n in nums:
                if n == int(n) and 1 <= abs(n) <= 999:
                    qty = int(n); break
            price_cands   = [n for n in nums if n != qty and 0.01 <= abs(n) < 999]
            unit_cost     = price_cands[0] if len(price_cands) > 0 else None
            current_price = price_cands[1] if len(price_cands) > 1 else None
            total_pnl     = None

        direction = "short" if (qty is not None and qty < 0) else "long"
        results.append({
            "account_id":    acct_id,
            "symbol":        sym,
            "direction":     direction,
            "strike":        occ["strike"],
            "expiry":        occ["expiry"],
            "underlying":    occ["underlying"],
            "quantity":      abs(qty) if qty is not None else None,
            "unit_cost":     unit_cost,
            "current_price": current_price,
            "total_pnl":     total_pnl,
            "_strategy":     strategy,
            "_raw_cells":    row.get("cells", [])[:6],
        })
        seen.add(sym)
    return results


def _parse_positions_xlsx(xlsx_path: pathlib.Path, acct_id: str) -> list[dict]:
    """
    解析 Firstrade positions Excel 的期权 sheet，返回 options_positions 行列表。
    中文表头列：代号(0) 详细说明(1) 数量(2) 价格(3) ... 单位成本(18) 益损$(20) ... 到期日(28)
    数量为负表示空头。
    """
    import openpyxl
    results: list[dict] = []
    try:
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=True)
        opt_sheet = None
        for name in wb.sheetnames:
            if "期权" in name or "option" in name.lower():
                opt_sheet = wb[name]
                break
        if opt_sheet is None:
            _log.warning(f"[pos-xl] 未找到期权 sheet，可用: {wb.sheetnames}")
            wb.close()
            return results

        rows = list(opt_sheet.iter_rows(values_only=True))
        wb.close()

        # 找表头行（含"代号"或"Symbol"的行）
        hdr: list[str] | None = None
        hdr_idx = 0
        for i, row in enumerate(rows):
            vals = [str(v or "").strip() for v in row]
            if "代号" in vals or "Symbol" in vals:
                hdr = vals
                hdr_idx = i
                break
        if hdr is None:
            _log.warning("[pos-xl] 未找到表头行")
            return results

        col: dict[str, int] = {v: i for i, v in enumerate(hdr)}

        def _get(row: tuple, *names: str):
            for n in names:
                idx = col.get(n)
                if idx is not None and idx < len(row):
                    v = row[idx]
                    if v is not None and str(v).strip():
                        return v
            return None

        def _num(v) -> float | None:
            if v is None:
                return None
            try:
                return float(str(v).replace(",", "").replace("$", "").strip())
            except (ValueError, TypeError):
                return None

        seen: set[str] = set()
        for row in rows[hdr_idx + 1:]:
            sym_raw = _get(row, "代号", "Symbol")
            if not sym_raw:
                continue
            sym = str(sym_raw).strip().upper().replace(" ", "")
            if not sym or sym in seen:
                continue
            occ = _parse_occ_sym(sym)
            if not occ:
                continue  # 跳过非期权行（股票等）

            qty       = _num(_get(row, "数量", "Quantity", "Qty"))
            price     = _num(_get(row, "Last Price", "价格", "Price", "Last"))
            unit_cost = _num(_get(row, "Unit Cost", "单位成本", "Avg Cost"))
            mv        = _num(_get(row, "Market Value", "市值", "MktVal"))
            day_pnl   = _num(_get(row, "Day Chg $", "$ Day Chg", "当日盈亏"))
            total_pnl = _num(_get(row, "$ Gain/Loss", "益损 $", "益损$", "Gain/Loss $", "P&L"))

            if qty is None:
                continue

            direction = "short" if qty < 0 else "long"
            results.append({
                "account_id":    acct_id,
                "symbol":        sym,
                "direction":     direction,
                "strike":        occ["strike"],
                "expiry":        occ["expiry"],
                "underlying":    occ["underlying"],
                "quantity":      int(qty),   # signed: negative = short position
                "unit_cost":     unit_cost,
                "current_price": price,
                "market_value":  mv,
                "day_pnl":       day_pnl,
                "total_pnl":     total_pnl,
                "_strategy":     "XLSX",
                "_raw_cells":    [],
            })
            seen.add(sym)

        _log.info(f"[pos-xl] 解析完成: {len(results)} 个期权持仓 from {xlsx_path.name}")
    except Exception as e:
        _log.error(f"[pos-xl] 解析出错: {e}")
    return results


def _scrape_and_diff_positions(driver, acct_id: str) -> dict:
    """
    步骤 1.5：优先通过 xlsx 下载同步持仓，失败则回退 JS 抓取。
    xlsx 模式做全量替换（DELETE+INSERT）；JS 模式做 UPSERT（不删除幻象行）。
    返回 diff 报告 dict 供 cascade 显示。
    """
    report: dict = {"ok": False, "summary": "", "changes": [],
                    "new_rows": [], "raw_url": "", "strategy": "none"}
    try:
        # ── 优先：xlsx 下载（_download_positions_xlsx 定义在本文件后段）──
        xlsx_path = _download_positions_xlsx(driver)
        new_rows: list[dict] = []

        if xlsx_path is not None:
            new_rows = _parse_positions_xlsx(xlsx_path, acct_id)
            report["strategy"] = "XLSX"
            report["raw_url"]  = str(xlsx_path)

        if not new_rows:
            # ── 回退：JS 抓取 ─────────────────────────────
            _log.info("[pos] xlsx 下载失败或无期权行，回退 JS 抓取")
            driver.switch_to.new_window("tab")
            driver.get(_FT_POSITIONS_URL)

            deadline = time.time() + 30
            while time.time() < deadline:
                body = driver.execute_script(
                    "return document.body ? document.body.innerText : '';")
                if len(body) > 800 and "Loading" not in body:
                    break
                time.sleep(1.0)
            time.sleep(1.0)

            raw = driver.execute_script(_POSITIONS_JS)
            _close_tab(driver)

            if not raw:
                report["summary"] = "持仓页 JS 返回空"
                return report

            report["raw_url"]  = raw.get("url", "")
            report["strategy"] = raw.get("strategy", "none")
            _log.info(f"[pos] JS url={raw['url']} body={raw['body_len']} "
                      f"strategy={raw['strategy']} rows={len(raw.get('rows', []))}")

            new_rows = _parse_scraped_rows(raw, acct_id)

            if not new_rows:
                sample = raw.get("opt_sample", [])
                report["summary"] = (
                    f"未识别到期权行 | 策略={raw['strategy']} "
                    f"body={raw['body_len']}字符 "
                    f"期权关键词样本={len(sample)}行")
                if sample:
                    report["changes"] = [f"页面样本: {s}" for s in sample[:5]]
                return report

        report["new_rows"] = new_rows

        # ── 读现有 DB 对比 ────────────────────────────────
        conn = _db()
        existing = {r[0]: {"quantity": r[1], "unit_cost": r[2],
                            "expiry": r[3], "direction": r[4]}
                    for r in conn.execute(
                        "SELECT symbol, quantity, unit_cost, expiry, direction "
                        "FROM options_positions WHERE account_id=?",
                        (acct_id,)).fetchall()}
        conn.close()

        changes: list[str] = []
        new_syms = {r["symbol"] for r in new_rows}
        for nr in new_rows:
            sym = nr["symbol"]
            old = existing.get(sym)
            if old is None:
                changes.append(
                    f"🆕 新增  {sym}  qty={nr['quantity']}  "
                    f"strike={nr['strike']}  expiry={nr['expiry']}")
            else:
                diffs: list[str] = []
                if nr["quantity"] is not None and nr["quantity"] != old["quantity"]:
                    diffs.append(f"qty {old['quantity']}→{nr['quantity']}")
                if (nr["unit_cost"] is not None and old["unit_cost"] is not None
                        and abs((nr["unit_cost"] or 0) - (old["unit_cost"] or 0)) > 0.005):
                    diffs.append(f"cost {old['unit_cost']:.4f}→{nr['unit_cost']:.4f}")
                if diffs:
                    changes.append(f"🔄 变更  {sym}  " + "  ".join(diffs))
                else:
                    changes.append(f"✅ 一致  {sym}  qty={nr['quantity']}")
        for sym in existing:
            if sym not in new_syms:
                label = "🗑️ 删除" if report["strategy"] == "XLSX" else "⚠️ DB有/未抓"
                changes.append(f"{label}  {sym}")

        n_new = sum(1 for c in changes if "🆕" in c)
        n_upd = sum(1 for c in changes if "🔄" in c)
        n_ok  = sum(1 for c in changes if "✅" in c)
        n_del = sum(1 for c in changes if "🗑️" in c)
        n_mis = sum(1 for c in changes if "⚠️" in c)
        report.update({
            "ok":      True,
            "summary": (f"[{report['strategy']}] {len(new_rows)} 行 | "
                        f"新增 {n_new} / 变更 {n_upd} / 一致 {n_ok} / "
                        f"删除 {n_del} / 缺失 {n_mis}"),
            "changes": changes,
        })
        for line in changes:
            _log.info(f"[pos] {line}")

        # ── 写入 DB ──────────────────────────────────────
        if _POSITIONS_WRITE and new_rows:
            now_ts = datetime.datetime.now(_ET).isoformat()
            conn = _db()

            if report["strategy"] == "XLSX":
                # xlsx 全量替换：删除旧持仓，重新插入（含价格/市值/盈亏）
                conn.execute(
                    "DELETE FROM options_positions WHERE account_id=?", (acct_id,))
                for nr in new_rows:
                    conn.execute("""
                        INSERT INTO options_positions
                          (account_id,symbol,direction,strike,expiry,
                           quantity,unit_cost,current_price,market_value,
                           day_pnl,total_pnl,last_updated)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                    """, (acct_id, nr["symbol"], nr["direction"], nr["strike"],
                          nr["expiry"], nr["quantity"], nr["unit_cost"],
                          nr["current_price"], nr.get("market_value"),
                          nr.get("day_pnl"), nr.get("total_pnl"), now_ts))
            else:
                # JS 回退：UPSERT（不删除，避免误删真实持仓）
                for nr in new_rows:
                    conn.execute("""
                        INSERT INTO options_positions
                          (account_id,symbol,direction,strike,expiry,quantity,
                           unit_cost,current_price,last_updated)
                        VALUES (?,?,?,?,?,?,?,?,?)
                        ON CONFLICT(account_id,symbol) DO UPDATE SET
                          direction     = excluded.direction,
                          strike        = excluded.strike,
                          expiry        = excluded.expiry,
                          quantity      = excluded.quantity,
                          unit_cost     = CASE WHEN excluded.unit_cost IS NOT NULL
                                               THEN excluded.unit_cost ELSE unit_cost END,
                          current_price = CASE WHEN excluded.current_price IS NOT NULL
                                               THEN excluded.current_price ELSE current_price END,
                          last_updated  = excluded.last_updated
                    """, (acct_id, nr["symbol"], nr["direction"], nr["strike"],
                          nr["expiry"], nr["quantity"], nr["unit_cost"],
                          nr["current_price"], now_ts))

            conn.commit()
            conn.close()
            report["summary"] += " | ✅ 已写入 DB"
            _log.info(f"[pos] wrote {len(new_rows)} rows "
                      f"(strategy={report['strategy']})")

    except Exception as e:
        report["summary"] = f"持仓抓取出错: {e}"
        _log.error(f"[pos] scrape error: {e}")
        try:
            _close_tab(driver)
        except Exception:
            pass

    return report


def _chrome_reachable() -> bool:
    try:
        import urllib.request
        with urllib.request.urlopen(f"http://{_CDP_ADDR}/json", timeout=2) as r:
            return r.status == 200
    except Exception:
        return False


def _get_driver():
    """附着到 CDP 9222 的已登录 Chrome，返回 driver 或 None。"""
    if not _SEL_OK:
        return None
    try:
        opts = _ChromeOpts()
        opts.add_experimental_option("debuggerAddress", _CDP_ADDR)
        try:
            return _wd.Chrome(options=opts)
        except Exception:
            from webdriver_manager.chrome import ChromeDriverManager
            from selenium.webdriver.chrome.service import Service as _Svc
            return _wd.Chrome(service=_Svc(ChromeDriverManager().install()), options=opts)
    except Exception as e:
        _log.warning(f"CDP attach failed: {e}")
        return None


def _nav_and_wait(driver, url: str, min_len: int = 300, timeout: float = 15) -> bool:
    """在新标签页打开 url，等待页面内容加载，返回是否成功。"""
    # switch_to.new_window 是原生 WebDriver 命令，不需要当前 tab 有 JS 执行上下文。
    # 替代 driver.execute_script("window.open('')") 以兼容渲染器崩溃场景。
    driver.switch_to.new_window("tab")
    driver.get(url)
    deadline = time.time() + timeout
    while time.time() < deadline:
        body = driver.execute_script(
            "return document.body ? (document.body.innerText||'') : '';")
        if body and len(body) > min_len:
            return True
        time.sleep(0.5)
    return False


def _close_tab(driver):
    """切回第一个窗口（不关闭当前标签）。
    Chrome 149 的 beforeunload 保护以及 CDP 关闭会使 WebDriver session 失效，
    因此放弃主动关闭——空标签累积远比 session 崩掉代价小。
    """
    try:
        driver.switch_to.window(driver.window_handles[0])
    except Exception:
        pass


def _scrape_balance(driver, acct_id: str) -> bool:
    ok = False
    try:
        if not _nav_and_wait(driver, _FT_BALANCE_URL):
            _log.warning("Balance page did not load in time")
            _close_tab(driver)
            return False

        data = driver.execute_script(_BALANCE_JS)
        _close_tab(driver)

        if not data or not any(v is not None for v in data.values()):
            _log.warning("Balance parse returned all None")
            return False

        mu = data.get("margin_used") or 0
        te = data.get("total_equity")
        if mu and te:
            # 融资余额占账户净值的比例（衡量杠杆程度）
            data["margin_usage_pct"] = mu / te * 100
        else:
            data["margin_usage_pct"] = 0.0

        _save_balance(acct_id, data)
        _log.info(f"Balance saved for {acct_id}: {data}")
        ok = True
    except Exception as e:
        _log.error(f"_scrape_balance error: {e}")
        _close_tab(driver)
    return ok


def _scrape_history_csv(driver) -> bool:
    """导航到历史页面，点击「下载」menuitem，触发 Chrome 下载 CSV。"""
    ok = False
    try:
        # Chrome 109+ 需要 Browser.setDownloadBehavior
        try:
            driver.execute_cdp_cmd("Browser.setDownloadBehavior", {
                "behavior": "allow",
                "downloadPath": str(_DOWNLOADS),
                "eventsEnabled": True,
            })
        except Exception:
            pass

        if not _nav_and_wait(driver, _FT_HISTORY_URL, min_len=200, timeout=12):
            _close_tab(driver)
            return False

        # Step 1: 点开菜单触发器（Manage Account History Menu）
        driver.execute_script(r"""
document.querySelectorAll('button,[role=button]').forEach(b => {
    if ((b.getAttribute('aria-label')||'').includes('Manage Account History Menu')) b.click();
});
""")
        time.sleep(1.2)

        # Step 2: 精准点击 role=menuitem 中文本为「下载」的项
        clicked = driver.execute_script(r"""
var items = document.querySelectorAll('[role=menuitem]');
for (var el of items) {
    if ((el.innerText||el.textContent||'').trim() === '下载') {
        el.dispatchEvent(new MouseEvent('click', {bubbles:true, cancelable:true}));
        return true;
    }
}
return false;
""")
        time.sleep(4)
        _close_tab(driver)
        ok = bool(clicked)
        _log.info(f"History CSV download triggered: {ok}")
    except Exception as e:
        _log.error(f"_scrape_history_csv error: {e}")
        _close_tab(driver)
    return ok


def _download_positions_xlsx(driver) -> "pathlib.Path | None":
    """
    在 Firstrade positions 页点击"下载"按钮，等待 xlsx 写完后返回文件路径。
    失败返回 None（调用方回退到 JS 抓取）。
    """
    try:
        try:
            driver.execute_cdp_cmd("Browser.setDownloadBehavior", {
                "behavior":      "allow",
                "downloadPath":  str(_DOWNLOADS),
                "eventsEnabled": True,
            })
        except Exception:
            pass

        # 快照现有 xlsx 的修改时间，用于识别新增或被覆盖的文件
        _t_start = time.time()
        before_mtimes: dict[pathlib.Path, float] = {
            f: f.stat().st_mtime for f in _DOWNLOADS.glob("*.xlsx")
        }

        driver.switch_to.new_window("tab")
        driver.get(_FT_POSITIONS_URL)

        # 等页面加载（body文字出现）+ 等 position-actions-menu 按钮渲染（SvelteKit异步）
        deadline = time.time() + 30
        while time.time() < deadline:
            body = driver.execute_script(
                "return document.body ? document.body.innerText : '';")
            if len(body) > 800 and "Loading" not in body:
                break
            time.sleep(1.0)
        # 再等 position-actions-menu 按钮出现（最多额外 12s）
        btn_deadline = time.time() + 12
        while time.time() < btn_deadline:
            has_btn = driver.execute_script(
                'return !!document.querySelector(\'[id^="position-actions"]\');')
            if has_btn:
                break
            time.sleep(0.8)

        # 策略 1：直接找 aria-label / title 含 Download/Export 的元素
        clicked = driver.execute_script(r"""
        var sels = [
            '[aria-label*="Download"]',
            '[aria-label*="download"]',
            '[aria-label*="Export"]',
            '[title*="Download"]',
            '[title*="Export"]'
        ];
        for (var i = 0; i < sels.length; i++) {
            var el = document.querySelector(sels[i]);
            if (el) {
                el.dispatchEvent(new MouseEvent('click', {bubbles:true, cancelable:true}));
                return 'direct:' + sels[i];
            }
        }
        // menuitem 文本 = 下载/Download/Export（已展开的菜单）
        var items = document.querySelectorAll('[role=menuitem]');
        for (var j = 0; j < items.length; j++) {
            var t = (items[j].innerText||items[j].textContent||'').trim();
            if (/^(下载|download|export)$/i.test(t)) {
                items[j].dispatchEvent(new MouseEvent('click', {bubbles:true, cancelable:true}));
                return 'menuitem:' + t;
            }
        }
        return null;
        """)

        if not clicked:
            # 策略 2：按钮文字 = 下载/Download（直接可见的文字按钮）
            clicked = driver.execute_script(r"""
            var btns = document.querySelectorAll('button,[role=button],a');
            for (var i = 0; i < btns.length; i++) {
                var t = (btns[i].innerText||btns[i].textContent||'').trim();
                if (/^(下载|download|export to excel|export excel)$/i.test(t)) {
                    btns[i].dispatchEvent(new MouseEvent('click', {bubbles:true, cancelable:true}));
                    return 'btn-text:' + t;
                }
            }
            return null;
            """)

        if not clicked:
            # 策略 3：找页面级 ⋮ 按钮（id^="position-actions"，有 data-dropdown-menu 属性）
            # 诊断确认：位于 Stocks/ETFs 上方，菜单含 Compact View/Company Logos/Download
            _trigger_clicked = driver.execute_script(r"""
            // 1. 优先：id前缀匹配（最精确）
            var trigger = document.querySelector('[id^="position-actions"]');
            // 2. 备选：data-dropdown-menu + aria-haspopup
            if (!trigger) {
                trigger = document.querySelector('[data-dropdown-menu][aria-haspopup=menu]');
            }
            // 3. 备选：aria-haspopup=menu 且无文字的空按钮
            if (!trigger) {
                var candidates = Array.from(document.querySelectorAll(
                    'button[aria-haspopup=menu]'));
                trigger = candidates.find(function(b){
                    return !(b.innerText||b.textContent||'').trim();
                }) || null;
            }
            if (!trigger) {
                // 调试：打印所有 aria-haspopup=menu 元素
                var dbg = Array.from(document.querySelectorAll('[aria-haspopup=menu]'))
                    .map(function(e){ return e.id+' '+e.className.slice(0,30); });
                return 'no-trigger:dbg=' + JSON.stringify(dbg);
            }
            trigger.click();
            return 'trigger-clicked:' + (trigger.id || trigger.className.slice(0,30));
            """)
            _log.info(f"[pos-dl] 策略3 触发器: {_trigger_clicked}")

            if _trigger_clicked and _trigger_clicked.startswith('trigger-clicked'):
                time.sleep(1.5)
                _menu_result = driver.execute_script(r"""
                var menu = document.querySelector('[role=menu]');
                if (!menu) return null;
                var walker = document.createTreeWalker(
                    menu, NodeFilter.SHOW_TEXT, null, false);
                var node;
                while ((node = walker.nextNode())) {
                    var v = node.nodeValue.trim();
                    if (/^(下载|Download|Export)$/i.test(v)) {
                        var p = node.parentElement;
                        p.click();
                        return 'dl:' + v + ':' + p.tagName;
                    }
                }
                return 'menu-no-dl:' + menu.textContent.trim().slice(0, 200);
                """)
                _log.info(f"[pos-dl] 策略3 菜单结果: {_menu_result}")
                if _menu_result and _menu_result.startswith('dl:'):
                    clicked = _menu_result

        if not clicked:
            # 全部策略失败：打印 Stocks/ETFs 区域 HTML 帮助调试
            section_html = driver.execute_script(r"""
            var allBtns = Array.from(document.querySelectorAll('button'));
            var etfEl = allBtns.find(b => (b.innerText||'').trim() === 'Stocks/ETFs');
            if (!etfEl) return 'Stocks/ETFs button not found';
            var container = etfEl.parentElement;
            for (var i = 0; i < 5; i++) {
                if (!container.parentElement) break;
                if (container.parentElement.querySelectorAll('button').length > 5) {
                    container = container.parentElement; break;
                }
                container = container.parentElement;
            }
            return container.outerHTML.slice(0, 800);
            """) or ''
            _log.warning(f"[pos-dl] 未找到下载按钮，区域HTML:\n{section_html}")
            _close_tab(driver)
            return None

        _log.info(f"[pos-dl] 下载按钮已点击: {clicked}")

        # 等待新/覆盖的 xlsx 文件出现（30秒超时，文件大小稳定 + >5 KB）
        # 同时检测新文件（不在 before_mtimes 中）和被覆盖的文件（mtime > _t_start）
        deadline = time.time() + 30
        found: "pathlib.Path | None" = None
        while time.time() < deadline:
            time.sleep(0.5)
            candidates: list[pathlib.Path] = []
            for f in _DOWNLOADS.glob("*.xlsx"):
                try:
                    mtime = f.stat().st_mtime
                    old_mtime = before_mtimes.get(f)
                    if old_mtime is None or mtime > _t_start:
                        candidates.append(f)
                except OSError:
                    pass
            pos_files = [f for f in candidates if "position" in f.name.lower()]
            hits = pos_files or candidates
            if hits:
                f = hits[0]
                try:
                    sz1 = f.stat().st_size
                    time.sleep(0.8)
                    if f.exists() and f.stat().st_size == sz1 and sz1 > 5000:
                        found = f
                        break
                except OSError:
                    pass

        _close_tab(driver)
        if found:
            _log.info(f"[pos-dl] ✅ {found.name}  {found.stat().st_size} bytes")
        else:
            _log.warning("[pos-dl] 等待 xlsx 超时（30s）")
        return found

    except Exception as e:
        _log.error(f"[pos-dl] 错误: {e}")
        _close_tab(driver)
        return None


# ── Chrome 自动启动 ───────────────────────────────────────
_CHROME_EXE_CANDIDATES = [
    r"C:\Program Files\Google\Chrome\Application\chrome.exe",
    r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
    str(pathlib.Path.home() / "AppData/Local/Google/Chrome/Application/chrome.exe"),
]
_CHROME_PROFILE_DIR = str(_ROOT / "data" / "chrome_profile")


def _find_chrome_exe() -> str | None:
    for p in _CHROME_EXE_CANDIDATES:
        if pathlib.Path(p).exists():
            return p
    return None


def _launch_chrome_cdp() -> bool:
    """用独立 Profile 启动调试模式 Chrome，等待 CDP 就绪，返回是否成功。"""
    import subprocess
    exe = _find_chrome_exe()
    if not exe:
        _log.error("[chrome] 未找到 Chrome 可执行文件")
        return False
    pathlib.Path(_CHROME_PROFILE_DIR).mkdir(parents=True, exist_ok=True)
    subprocess.Popen(
        [
            exe,
            "--remote-debugging-port=9222",
            f"--user-data-dir={_CHROME_PROFILE_DIR}",
            "--no-first-run",
            "--no-default-browser-check",
            "--disable-extensions",
        ],
        # Windows: 不弹出额外控制台窗口
        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
    )
    deadline = time.time() + 15
    while time.time() < deadline:
        if _chrome_reachable():
            return True
        time.sleep(0.8)
    _log.error("[chrome] CDP 15s 内未就绪")
    return False


def _check_ft_login() -> bool:
    """
    检测是否已登录 Firstrade。
    通过 CDP JSON 检查现有标签页，无需 WebDriver。
    已登录 → True；未登录 → False。
    """
    try:
        import urllib.request, json as _json
        with urllib.request.urlopen(f"http://{_CDP_ADDR}/json", timeout=3) as _r:
            tabs = _json.loads(_r.read())
        for tab in tabs:
            if "invest.firstrade.com/app" in tab.get("url", ""):
                _log.info(f"[chrome] login OK (existing tab): {tab['url'][:60]}")
                return True
    except Exception as _e:
        _log.debug(f"[chrome] CDP tab list check failed: {_e}")
    return False


def _ensure_chrome(step=None) -> str:
    """
    确保 Chrome CDP 可用且已登录 Firstrade。
    返回值: "ready" / "needs_login" / "no_chrome"

    调用方根据返回值决定是否继续同步：
      "ready"       → 直接继续
      "needs_login" → 提示用户在弹出窗口登录后再次点击
      "no_chrome"   → Chrome 未安装或启动失败
    """
    def _s(msg: str) -> None:
        _log.info(msg)
        if step:
            step(msg)

    if not _chrome_reachable():
        _s("🔍 Chrome CDP 未连接，自动启动...")
        if not _launch_chrome_cdp():
            _s("❌ Chrome 启动失败（未找到可执行文件或 15s 超时）")
            return "no_chrome"
        _s("✅ Chrome 已启动")
    else:
        _s("✅ Chrome CDP 已连接")

    # _check_ft_login 使用 CDP JSON，不需要 WebDriver driver
    if _check_ft_login():
        _s("✅ Firstrade 已登录，准备同步")
        return "ready"
    else:
        _s("🔐 Firstrade 需要登录")
        _s("   → 请在弹出的 Chrome 窗口完成登录，然后再次点击「⚡ 同步账户」")
        return "needs_login"


def _verify_session(driver) -> bool:
    """
    真实导航验证 Firstrade session 是否有效（不信任 CDP tab URL 缓存）。
    开新标签 → 导航到 balance 页 → 检查是否被重定向到 login。
    """
    try:
        driver.switch_to.new_window("tab")
        driver.get(_FT_BALANCE_URL)
        time.sleep(3)
        ok = "cgi-bin/login" not in driver.current_url
        _close_tab(driver)
        return ok
    except Exception as _e:
        _log.debug(f"[session] verify error: {_e}")
        try:
            _close_tab(driver)
        except Exception:
            pass
        return False


def _do_ft_login(driver) -> bool:
    """
    用 .env 里的 FIRSTRADE_USER_1 / FIRSTRADE_PASS_1 自动填表登录。
    成功 → True；失败（未找到表单 / 超时 / 无凭据）→ False。
    """
    from selenium.webdriver.common.by import By
    user = os.environ.get("FIRSTRADE_USER_1", "")
    pwd  = os.environ.get("FIRSTRADE_PASS_1", "")
    if not user or not pwd:
        _log.warning("[login] .env 未配置 FIRSTRADE_USER_1/PASS_1，无法自动登录")
        return False
    try:
        driver.switch_to.new_window("tab")
        driver.get("https://invest.firstrade.com/cgi-bin/login")
        time.sleep(3)

        # 用 JS 找所有可见 text/password 输入框（兼容不同表单结构）
        fields = driver.execute_script(r"""
        var inputs = Array.from(document.querySelectorAll('input'));
        return inputs.filter(function(i){
            var t = i.type.toLowerCase();
            return (t==='text'||t==='email'||t==='password') &&
                   i.offsetWidth > 0 && i.offsetHeight > 0;
        }).map(function(i){
            return {type: i.type, name: i.name, id: i.id};
        });
        """) or []
        _log.info(f"[login] 可见输入框: {fields}")

        u_sel = next(
            (f for f in fields if f["type"].lower() in ("text", "email")), None)
        p_sel = next(
            (f for f in fields if f["type"].lower() == "password"), None)

        if not u_sel or not p_sel:
            _log.warning(f"[login] 表单字段未找到: {fields}")
            _close_tab(driver)
            return False

        def _pick(info):
            if info["id"]:
                return f'#{info["id"]}'
            if info["name"]:
                return f'input[name="{info["name"]}"]'
            return f'input[type="{info["type"]}"]'

        u_el = driver.find_element(By.CSS_SELECTOR, _pick(u_sel))
        p_el = driver.find_element(By.CSS_SELECTOR, _pick(p_sel))
        u_el.clear(); u_el.send_keys(user)
        p_el.clear(); p_el.send_keys(pwd)

        # 点提交按钮，找不到则用 form.submit()
        submitted = driver.execute_script(r"""
        var btn = document.querySelector('button[type="submit"],input[type="submit"]');
        if (btn) { btn.click(); return true; }
        var form = document.querySelector('form');
        if (form) { form.submit(); return true; }
        return false;
        """)
        if not submitted:
            p_el.submit()

        # 等待跳转到 /app/（最多 20s）
        deadline = time.time() + 20
        while time.time() < deadline:
            time.sleep(1)
            if "invest.firstrade.com/app" in driver.current_url:
                _log.info(f"[login] ✅ 自动重登成功: {driver.current_url[:60]}")
                _close_tab(driver)
                return True

        _log.warning(f"[login] 登录超时，当前URL: {driver.current_url[:80]}")
        _close_tab(driver)
        return False
    except Exception as e:
        _log.error(f"[login] 自动登录异常: {e}")
        try:
            _close_tab(driver)
        except Exception:
            pass
        return False


def _auto_sync(acct_id: str = "account_1"):
    """调度器 / 手动触发入口（后台线程中运行，禁止调用 st.*）。"""
    ss = _sync_state()
    ss["last_status"] = "running"
    ss["chrome_ok"]   = False

    if not _chrome_reachable():
        ss["last_status"] = "no_chrome"
        ss["last_error"]  = "Chrome CDP 9222 未响应"
        _log.info("Auto sync skipped: Chrome not reachable")
        return

    driver = _get_driver()
    if driver is None:
        ss["last_status"] = "error"
        ss["last_error"]  = "无法获取 ChromeDriver"
        return

    ss["chrome_ok"] = True

    # ── Session 验证 + 自动重登（最多重试1次）──────────────────────
    if not _verify_session(driver):
        _log.info("[sync] Session 已过期，尝试自动重登...")
        if not _do_ft_login(driver):
            ss["last_status"] = "needs_login"
            ss["last_error"]  = "Session 过期且自动重登失败，请手动在 Chrome 登录后重试"
            _log.warning("[sync] 自动重登失败，终止同步")
            return
        _log.info("[sync] 自动重登成功，继续同步")

    try:
        bal_ok = _scrape_balance(driver, acct_id)

        # 步骤 1.5：下载持仓 xlsx → 对比写 DB
        ss["positions_diff"] = None
        try:
            diff = _scrape_and_diff_positions(driver, acct_id)
            ss["positions_diff"] = diff
        except Exception as _pe:
            _log.warning(f"[pos] step 1.5 error: {_pe}")
            ss["positions_diff"] = {"ok": False, "summary": str(_pe), "changes": []}

        csv_ok = _scrape_history_csv(driver)
        ss["bal_ok"]      = bal_ok
        ss["csv_ok"]      = csv_ok
        ss["last_time"]   = datetime.datetime.now(_ET)
        ss["last_status"] = "ok" if bal_ok else "partial"
        ss["last_error"]  = "" if bal_ok else "余额页面数据解析失败，请检查页面结构"
    except Exception as e:
        ss["last_status"] = "error"
        ss["last_error"]  = str(e)
        _log.error(f"Auto sync exception: {e}")
    finally:
        pass  # 不调用 driver.quit()，避免关闭用户的 Chrome


def _auto_greeks_check():
    """触发二：每个交易日 09:35 全组合 Greeks 快照 + Delta 漂移检测。"""
    for cfg in ACCT_CFG:
        try:
            _compute_portfolio_greeks(cfg["id"])
            _log.info(f"[Greeks] 09:35 snapshot done for {cfg['id']}")
        except Exception as _exc:
            _log.exception(f"[Greeks] 09:35 snapshot error for {cfg['id']}: {_exc}")


def _auto_price_refresh():
    """
    不依赖 Chrome — 仅用 MarketData.app + yfinance 刷新所有账户的
    期权现价 / Greeks 和股票现价。每日交易时段自动执行 4 次。
    """
    for cfg in ACCT_CFG:
        aid = cfg["id"]
        try:
            _refresh_options_prices(aid)
            _log.info(f"[sched] option prices refreshed: {aid}")
        except Exception as e:
            _log.warning(f"[sched] option price refresh {aid}: {e}")
        try:
            _refresh_stock_prices(aid)
            _log.info(f"[sched] stock prices refreshed: {aid}")
        except Exception as e:
            _log.warning(f"[sched] stock price refresh {aid}: {e}")


@st.cache_resource
def _start_scheduler():
    if not _SCHED_OK:
        return None
    sched = BackgroundScheduler(timezone="America/New_York")
    # Balance + CSV sync (requires Chrome CDP)
    for h, m in [(9, 35), (12, 30), (15, 30), (16, 30)]:
        sched.add_job(_auto_sync, CronTrigger(hour=h, minute=m,
                      timezone="America/New_York"),
                      id=f"sync_{h:02d}{m:02d}", replace_existing=True)
    # Price-only refresh (no Chrome needed)
    for h, m in [(9, 40), (12, 35), (15, 35), (16, 35)]:
        sched.add_job(_auto_price_refresh, CronTrigger(hour=h, minute=m,
                      timezone="America/New_York"),
                      id=f"price_{h:02d}{m:02d}", replace_existing=True)
    # Greeks daily check — Trigger 2
    sched.add_job(_auto_greeks_check, CronTrigger(hour=9, minute=35,
                  timezone="America/New_York"),
                  id="greeks_0935", replace_existing=True)
    # Daily briefing — generate at 09:35 after sync completes
    for _cfg in ACCT_CFG:
        _aid = _cfg["id"]
        sched.add_job(
            _generate_and_save_daily_briefing,
            CronTrigger(hour=9, minute=35, timezone="America/New_York"),
            args=[_aid],
            id=f"briefing_0935_{_aid}",
            replace_existing=True,
        )
    # Beta 每周日 02:00 ET 自动刷新
    sched.add_job(_refresh_beta_spy,
                  CronTrigger(day_of_week="sun", hour=2, minute=0,
                              timezone="America/New_York"),
                  id="beta_refresh_weekly", replace_existing=True)
    sched.start()
    _log.info("Scheduler started: sync 09:35/12:30/15:30/16:30 · prices 09:40/12:35/15:35/16:35 · greeks 09:35 · beta-refresh Sun02:00 ET")
    return sched


# ════════════════════════════════════════════════════════
# Streamlit 页面
# ════════════════════════════════════════════════════════
st.set_page_config(page_title="ENERGREX · 账户监控", page_icon="🏦",
                   layout="wide", initial_sidebar_state="expanded")

_BG, _SURF, _BORDER = "#0A1628", "#0F1923", "#1E2D3D"
_TEXT, _MUTED       = "#E2E8F0", "#8B9BB4"
_GREEN, _RED, _AMB  = "#00D4AA", "#FF4B6E", "#FFB347"
_BLUE, _PURP        = "#4FC3F7", "#A78BFA"

st.markdown(f"""<style>
footer{{visibility:hidden;}} #MainMenu{{visibility:hidden;}}
.acct-card{{background:{_SURF};border:1px solid {_BORDER};border-radius:10px;padding:16px 18px;}}
.acct-title{{font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:1.5px;
             color:{_MUTED};margin-bottom:12px;}}
.mrow{{display:flex;justify-content:space-between;margin-bottom:6px;align-items:baseline;}}
.mlbl{{font-size:11px;color:{_MUTED};}}
.mval{{font-size:13px;font-weight:700;color:{_TEXT};}}
.mval.g{{color:{_GREEN};}} .mval.r{{color:{_RED};}} .mval.a{{color:{_AMB};}}
.sync-log{{background:{_BG};border:1px solid {_BORDER};border-radius:6px;
           padding:10px 14px;font-size:11px;line-height:1.9;max-height:200px;
           overflow-y:auto;font-family:monospace;color:{_MUTED};}}
.login-banner{{background:#1a2e1a;border:1px solid {_GREEN};border-radius:10px;
               padding:20px 24px;margin:12px 0;}}
/* ── 全局 st.metric 字体缩小，防止数字截断 ── */
[data-testid="stMetricValue"]{{
    font-size:1.0rem !important;
    line-height:1.4 !important;
    white-space:normal !important;
    word-break:break-word !important;
}}
[data-testid="stMetricLabel"]{{font-size:0.72rem !important;}}
[data-testid="stMetricDelta"]{{font-size:0.72rem !important;}}
</style>""", unsafe_allow_html=True)

# 启动 watchdog + 调度器（session_state 防止重复启动）
_start_watcher()
ws = _watch_state()
if not st.session_state.get("_scheduler_started"):
    _start_scheduler()
    st.session_state["_scheduler_started"] = True

# ── 共享侧边栏（导航 / 数据更新 / 风险状态）────────────────
import _sidebar as _sb
_sb.render()

# ════════════════════════════════════════════════════════
# 侧边栏（账户监控专属：文件上传 / 账户选择）
# ════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown(f"<div style='font-size:20px;font-weight:800;letter-spacing:2px;color:{_GREEN}'>"
                f"⚡ ENERGREX</div><div style='font-size:11px;color:{_MUTED};margin-bottom:8px'>"
                f"账户持仓监控</div>", unsafe_allow_html=True)
    st.divider()

    # 监控状态
    if ws["last_time"]:
        st.success(f"✓ 最近导入：{ws['last_file']}")
        st.caption(f"{ws['last_time'].strftime('%H:%M:%S')} ET · {ws['last_rows']} 行")
    else:
        st.info("⏳ 监控中，等待 Firstrade 导出文件…")

    # 手动导入按钮（选文件）
    st.divider()
    up = st.file_uploader("手动上传 CSV", type="csv", label_visibility="collapsed",
                          help="也可直接从 Firstrade 下载到 Downloads 文件夹，自动检测")
    if up:
        tmp = _FT_DIR / up.name
        tmp.write_bytes(up.read())
        result = _process_csv_file(tmp)
        if result["ok"]:
            ws["last_file"] = result["file"]
            ws["last_time"] = datetime.datetime.now(_ET)
            ws["last_type"] = result["type"]
            ws["last_rows"] = result["rows"]
            ws["new_data"]  = True
            st.success(f"✓ 导入 {result['rows']} 行")
            st.rerun()
        else:
            st.error(result.get("reason", "解析失败"))

    st.divider()
    sel_acct_label = st.selectbox("账户", [c["label"] for c in ACCT_CFG],
                                  key="sb_acct", label_visibility="collapsed")
    st.markdown(f"<div style='color:{_MUTED};font-size:11px;line-height:2'>"
                f"{'🟢' if _WATCHDOG_OK else '🔴'} watchdog 文件监控<br>"
                f"📂 ~/Downloads/export*.csv"
                f"</div>", unsafe_allow_html=True)

# ════════════════════════════════════════════════════════
# 主区域
# ════════════════════════════════════════════════════════
# ── 新数据 toast ──────────────────────────────────────────
if ws["new_data"]:
    st.toast(f"✅ 检测到新数据，已自动更新（{ws['last_file']}，{ws['last_rows']} 行）",
             icon="📥")
    ws["new_data"] = False

# ════════════════════════════════════════════════════════
# 每日作战室
# ════════════════════════════════════════════════════════
_war_acct_id = next((c["id"] for c in ACCT_CFG if c["label"] == sel_acct_label),
                    ACCT_CFG[0]["id"])
_now_et   = datetime.datetime.now(_ET)
_today_str = _now_et.strftime("%Y年%m月%d日  %H:%M ET")

st.markdown(
    f"<div style='display:flex;align-items:center;justify-content:space-between;"
    f"margin-bottom:4px'>"
    f"<span style='font-size:22px;font-weight:800;color:{_GREEN};letter-spacing:1px'>"
    f"⚡ ENERGREX 作战室</span>"
    f"<span style='font-size:12px;color:{_MUTED}'>{_today_str}</span>"
    f"</div>",
    unsafe_allow_html=True,
)

# ── A. 实时风险状态条 ────────────────────────────────────
with st.spinner("载入风险快照…"):
    _war_snap = _compute_risk_snapshot(_war_acct_id)

_war_risk  = _war_snap.get("risk_status", "GREEN")
_RISK_BAR  = {
    "RED_HARD_STOP":  (_RED,  "🔴 高度警戒 — 压力损失 ≥15%，须立即减仓"),
    "ORANGE_DE_RISK": (_AMB,  "🟠 风险偏高 — 压力损失 ≥12%，建议减仓审查"),
    "YELLOW_WARNING": ("#FFD700", "🟡 轻度警示 — 压力损失 ≥8%，关注风险"),
    "GREEN":          (_GREEN, "🟢 风险正常 — 压力测试通过"),
}
_risk_color, _risk_label = _RISK_BAR.get(_war_risk, (_MUTED, _war_risk))
_bdr_now     = (_war_snap.get("beta_delta_ratio") or 0) * 100
_s10_now     = (_war_snap.get("stress_10_ratio")  or 0) * 100
_s10_amt_now = _war_snap.get("stress_10", 0) or 0
_s20_now     = (_war_snap.get("stress_20_ratio")  or 0) * 100
_s20_amt_now = _war_snap.get("stress_20", 0) or 0
_theta_now   = _war_snap.get("theta_per_day", 0)
_equity_now  = _war_snap.get("equity", 0)
_sc10_now = _RED if abs(_s10_now) > 20 else "#FFD700" if abs(_s10_now) > 10 else _GREEN
_sc20_now = _RED if abs(_s20_now) > 20 else "#FFD700" if abs(_s20_now) > 10 else _GREEN

st.markdown(
    f"<div style='background:{_risk_color}18;border:2px solid {_risk_color};"
    f"border-radius:10px;padding:10px 18px;margin:6px 0;"
    f"display:flex;align-items:center;justify-content:space-between'>"
    f"<span style='font-size:15px;font-weight:700;color:{_risk_color}'>{_risk_label}</span>"
    f"<div style='font-size:11px;color:{_MUTED};text-align:right;line-height:1.7'>"
    f"BD {_bdr_now:.1f}%&nbsp;&nbsp;·&nbsp;&nbsp;Theta ${_theta_now:+,.0f}/天<br>"
    f"<span style='color:{_sc10_now}'>压力 -10%&nbsp;&nbsp;"
    f"${_s10_amt_now:+,.0f}&nbsp;({_s10_now:+.1f}% 净值)</span>"
    f"&nbsp;&nbsp;·&nbsp;&nbsp;"
    f"<span style='color:{_sc20_now}'>压力 -20%&nbsp;&nbsp;"
    f"${_s20_amt_now:+,.0f}&nbsp;({_s20_now:+.1f}% 净值)</span>"
    f"</div></div>",
    unsafe_allow_html=True,
)

# Beta-Delta 额外警告
if abs(_bdr_now) > _RISK_LIMITS["max_beta_delta_ratio"] * 100:
    st.warning(f"🟠 **Beta-Delta {_bdr_now:.1f}%** 超过限额 "
               f"{_RISK_LIMITS['max_beta_delta_ratio']*100:.0f}%，建议执行对冲")

# ── B. 卖Call触发提醒（实时）────────────────────────────
with st.spinner("检查 Long Call 触发条件…"):
    _call_triggers = _check_sell_call_triggers(_war_acct_id)

if _call_triggers:
    st.markdown(
        f"<div style='background:{_RED}18;border:2px solid {_RED};"
        f"border-radius:8px;padding:10px 16px;margin:6px 0'>"
        f"<div style='font-size:13px;font-weight:700;color:{_RED};"
        f"margin-bottom:6px'>🔔 Long Call 触发预警（{len(_call_triggers)} 个）</div>",
        unsafe_allow_html=True,
    )
    for _ct in _call_triggers:
        _trig_str = " | ".join(_ct["triggers"])
        _pnl_s = f"盈亏 {_ct['pnl_pct']:+.0f}%" if _ct["pnl_pct"] is not None else ""
        _dte_s = f"DTE={_ct['dte']}" if _ct["dte"] is not None else ""
        st.markdown(
            f"<div style='background:{_RED}10;border:1px solid {_RED}33;"
            f"border-radius:6px;padding:8px 12px;margin:4px 0'>"
            f"<span style='color:{_RED};font-weight:700'>{_ct['underlying']} "
            f"${_ct['strike']:.0f}C {_ct['expiry']}</span>"
            f"<span style='color:{_MUTED};font-size:11px;margin-left:10px'>"
            f"{_dte_s}  {_pnl_s}  [{_trig_str}]</span><br>"
            f"<span style='color:{_TEXT};font-size:12px'>{_ct['suggestion']}</span>"
            f"</div>",
            unsafe_allow_html=True,
        )
    st.markdown("</div>", unsafe_allow_html=True)

# ── C. 价外程度预警（价差接近归零）────────────────────────
with st.spinner("检查价差组合 OTM 程度…"):
    _otm_alerts = _check_otm_spread_alerts(_war_acct_id)

if _otm_alerts:
    st.markdown(
        f"<div style='background:#FF8C0018;border:2px solid #FF8C00;"
        f"border-radius:8px;padding:10px 16px;margin:6px 0'>"
        f"<div style='font-size:13px;font-weight:700;color:#FF8C00;"
        f"margin-bottom:6px'>⚠️ 价差组合接近归零预警（{len(_otm_alerts)} 个）</div>",
        unsafe_allow_html=True,
    )
    for _oa in _otm_alerts:
        _dte_s = f" DTE={_oa['dte']}" if _oa["dte"] is not None else ""
        st.markdown(
            f"<div style='background:#FF8C0010;border:1px solid #FF8C0033;"
            f"border-radius:6px;padding:8px 12px;margin:4px 0'>"
            f"<span style='color:#FF8C00;font-weight:700'>{_oa['underlying']} "
            f"{_oa['spread_type']} ${_oa['low_strike']:.0f}/{_oa['high_strike']:.0f}"
            f"{'P' if _oa['opt_type']=='P' else 'C'} {_oa['expiry']}</span>"
            f"<span style='color:{_MUTED};font-size:11px;margin-left:10px'>"
            f"当前价值 ${_oa['current_value']:.0f} / 原始成本 ${_oa['original_cost']:.0f}"
            f"  剩余 {_oa['pct_remaining']:.1f}%{_dte_s}</span><br>"
            f"<span style='color:{_TEXT};font-size:12px'>{_oa['message']}</span>"
            f"</div>",
            unsafe_allow_html=True,
        )
    st.markdown("</div>", unsafe_allow_html=True)

# ── D. 标的单日异动预警（±5%）─────────────────────────────
with st.spinner("检查标的单日异动…"):
    _move_alerts = _check_underlying_move_alerts(_war_acct_id)

if _move_alerts:
    st.markdown(
        f"<div style='background:#FFD70018;border:2px solid #FFD700;"
        f"border-radius:8px;padding:10px 16px;margin:6px 0'>"
        f"<div style='font-size:13px;font-weight:700;color:#B8860B;"
        f"margin-bottom:6px'>📊 标的单日异动预警（{len(_move_alerts)} 个）</div>",
        unsafe_allow_html=True,
    )
    for _ma in _move_alerts:
        _chg_color = _RED if _ma["chg_pct"] < 0 else _GREEN
        st.markdown(
            f"<div style='background:#FFD70010;border:1px solid #FFD70033;"
            f"border-radius:6px;padding:8px 12px;margin:4px 0'>"
            f"<span style='color:#B8860B;font-weight:700'>{_ma['underlying']}</span>"
            f"<span style='color:{_chg_color};font-weight:700;margin-left:10px'>"
            f"{_ma['chg_pct']:+.1f}%</span>"
            f"<span style='color:{_MUTED};font-size:11px;margin-left:10px'>"
            f"${_ma['prev_close']:.2f} → ${_ma['price']:.2f}</span><br>"
            f"<span style='color:{_TEXT};font-size:12px'>{_ma['message']}</span>"
            f"</div>",
            unsafe_allow_html=True,
        )
    st.markdown("</div>", unsafe_allow_html=True)

# ── E. 今日操作简报 ──────────────────────────────────────
_briefing = _load_today_briefing(_war_acct_id)

_brief_hdr_col, _brief_btn_col = st.columns([4, 1])
with _brief_hdr_col:
    if _briefing:
        _gen_time_disp = (_briefing["gen_time"] or "")[:16].replace("T", " ")
        st.markdown(
            f"<div style='font-size:13px;font-weight:700;color:{_TEXT};"
            f"margin:8px 0 4px'>⚡ 今日操作简报</div>"
            f"<div style='font-size:11px;color:{_MUTED}'>"
            f"生成时间：{_gen_time_disp} ET</div>",
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            f"<div style='font-size:13px;font-weight:700;color:{_TEXT};"
            f"margin:8px 0 4px'>⚡ 今日操作简报</div>"
            f"<div style='font-size:11px;color:{_AMB}'>"
            f"尚未生成（09:35 ET 自动生成，或手动触发）</div>",
            unsafe_allow_html=True,
        )

with _brief_btn_col:
    if st.button("⚡ 立即生成", key="gen_briefing", use_container_width=True,
                 help="手动触发生成今日简报（等同于 09:35 自动任务）"):
        with st.spinner("生成作战简报…"):
            _generate_and_save_daily_briefing(_war_acct_id)
        st.rerun()

if _briefing:
    _b_recs   = _briefing.get("recs", [])
    _b_alerts = _briefing.get("alerts", [])
    _b_snap   = _briefing.get("snap", {})

    # 4 KPI
    _kw1, _kw2, _kw3, _kw4 = st.columns(4)
    _b_eq      = _b_snap.get("equity", 0) or 0
    _b_bdr     = (_b_snap.get("beta_delta_ratio") or 0) * 100
    _b_s10     = (_b_snap.get("stress_10_ratio")  or 0) * 100
    _b_s10_amt = _b_snap.get("stress_10", 0) or 0
    _b_s20     = (_b_snap.get("stress_20_ratio")  or 0) * 100
    _b_s20_amt = _b_snap.get("stress_20", 0) or 0
    _b_th      = _b_snap.get("theta_per_day", 0) or 0
    _b_nexp_s  = _b_snap.get("nearest_expiry_sym", "")
    _b_nexp_d  = _b_snap.get("nearest_expiry_date")
    _sc10_b = _RED if abs(_b_s10) > 20 else "#FFD700" if abs(_b_s10) > 10 else _GREEN
    _sc20_b = _RED if abs(_b_s20) > 20 else "#FFD700" if abs(_b_s20) > 10 else _GREEN
    with _kw1:
        st.metric("账户净值",  f"${_b_eq:,.0f}")
    with _kw2:
        _bdr_col = "normal" if abs(_b_bdr) <= _RISK_LIMITS["max_beta_delta_ratio"]*100 else "inverse"
        st.metric("Beta-Delta", f"{_b_bdr:.1f}%", delta_color=_bdr_col)
    with _kw3:
        st.markdown(
            f"<div style='padding:4px 0'>"
            f"<div style='font-size:11px;color:{_MUTED};font-weight:600;"
            f"text-transform:uppercase;letter-spacing:.5px;margin-bottom:6px'>压力测试</div>"
            f"<div style='font-size:13px;color:{_sc10_b};line-height:1.8'>"
            f"压力 -10%&nbsp;&nbsp;"
            f"<b>${_b_s10_amt:+,.0f}</b>&nbsp;({_b_s10:+.1f}% 净值)</div>"
            f"<div style='font-size:13px;color:{_sc20_b};line-height:1.8'>"
            f"压力 -20%&nbsp;&nbsp;"
            f"<b>${_b_s20_amt:+,.0f}</b>&nbsp;({_b_s20:+.1f}% 净值)</div>"
            f"</div>",
            unsafe_allow_html=True,
        )
    with _kw4:
        st.metric("每日 Theta", f"${_b_th:+,.0f}")

    # 紧急建议列表（最多5条）
    _urgent = [r for r in _b_recs if r.get("优先级", "").startswith(("🔴", "🟠"))]
    _other  = [r for r in _b_recs if not r.get("优先级", "").startswith(("🔴", "🟠"))]
    _show_recs = (_urgent + _other)[:5]

    if _show_recs:
        for _br in _show_recs:
            _bp_color = (_RED  if _br["优先级"].startswith("🔴") else
                         _AMB  if _br["优先级"].startswith("🟠") else
                         _GREEN if _br["优先级"].startswith("🟢") else _MUTED)
            st.markdown(
                f"<div style='background:{_SURF};border-left:4px solid {_bp_color};"
                f"border-radius:0 6px 6px 0;padding:8px 14px;margin:4px 0'>"
                f"<span style='color:{_bp_color};font-weight:700'>{_br['优先级']}</span>"
                f"<span style='color:{_TEXT};margin:0 8px;font-weight:600'>{_br['标的']}</span>"
                f"<span style='color:{_MUTED};font-size:11px'>{_br.get('组合','—')} "
                f"· DTE {_br.get('DTE','—')}</span><br>"
                f"<span style='color:{_TEXT};font-size:12px'>{_br['行动建议']}</span>"
                f"</div>",
                unsafe_allow_html=True,
            )
        if len(_b_recs) > 5:
            st.caption(f"另有 {len(_b_recs) - 5} 条建议 → 见「💡 交易建议」标签页")

    # 简报中的 call alerts（保存时生成）
    if _b_alerts:
        st.caption(f"📌 简报生成时检测到 {len(_b_alerts)} 个 Long Call 触发（实时数据见上方红框）")

# ── D. 到期预警 ──────────────────────────────────────────
_nexp_date_war = _war_snap.get("nearest_expiry_date")
_nexp_sym_war  = _war_snap.get("nearest_expiry_sym", "")
if _nexp_date_war:
    _dte_war = (_nexp_date_war - datetime.date.today()).days
    if _dte_war <= 7:
        st.error(f"🔴 **到期紧急** — {_nexp_sym_war} 仅剩 **{_dte_war} 天**到期，须今日决策")
    elif _dte_war <= 14:
        st.error(f"🔴 **到期预警** — {_nexp_sym_war} 还有 **{_dte_war} 天**，须本周处理")
    elif _dte_war <= 21:
        st.warning(f"🟠 **即将到期** — {_nexp_sym_war} 还有 **{_dte_war} 天**，开始减仓评估")

st.divider()

# ════════════════════════════════════════════════════════
# 账户卡片
# ════════════════════════════════════════════════════════
st.markdown("#### 账户总览")

# ── CDP 同步状态（只读，同步入口在左侧边栏）─────────────────
_ss           = _sync_state()
_chrome_alive = _chrome_reachable()
_ss["chrome_ok"] = _chrome_alive
_sel_cfg = next(c for c in ACCT_CFG if c["label"] == sel_acct_label)

if _chrome_alive:
    if _ss["last_time"]:
        _ts_disp = _ss["last_time"].strftime("%m-%d %H:%M") + " ET"
        if _ss["last_status"] in ("ok", "partial"):
            st.success(f"✅ Chrome 已连接 · 上次同步: {_ts_disp}")
        elif _ss["last_status"] == "running":
            st.info("🔄 同步中…")
        else:
            st.error(f"❌ 同步失败 · {_ss.get('last_error','')}")
    else:
        st.info("🟢 Chrome 已连接，等待首次自动同步（09:35 / 12:30 / 15:30 / 16:30 ET）")
else:
    st.warning("🔴 Chrome 未连接 — 请运行 `start_chrome.bat` 并登录 Firstrade，或从左侧边栏点击「⚡ 同步账户」")


def _fmt_m(v, d="—"):
    return f"${v:,.2f}" if v is not None else d

def _fmt_p(v, d="—"):
    return f"{v:.1f}%" if v is not None else d

def _pnl_cls(v):
    if v is None: return "mval"
    return "mval g" if v >= 0 else "mval r"

def _margin_cls(v):
    if v is None: return "mval"
    if v < 50: return "mval g"
    if v < 75: return "mval a"
    return "mval r"


cols = st.columns(2)
for col, cfg in zip(cols, ACCT_CFG):
    bal  = _load_latest_balance(cfg["id"])
    te   = bal.get("total_equity")
    cb   = bal.get("cash_balance")
    mu   = bal.get("margin_used")
    ma   = bal.get("margin_available")
    mp   = bal.get("margin_usage_pct")
    dp   = bal.get("day_pnl")
    ts   = (bal.get("sync_time") or "")[:16]
    with col:
        # ── 余额卡片 ──
        st.markdown(f"""
<div class='acct-card'>
  <div class='acct-title'>{cfg['label']}</div>
  <div class='mrow'><span class='mlbl'>总资产净值</span>
    <span class='mval' style='font-size:16px'>{_fmt_m(te)}</span></div>
  <div class='mrow'><span class='mlbl'>现金余额</span>
    <span class='mval'>{_fmt_m(cb)}</span></div>
  <div class='mrow'><span class='mlbl'>保证金已用</span>
    <span class='mval'>{_fmt_m(mu)}</span></div>
  <div class='mrow'><span class='mlbl'>保证金可用</span>
    <span class='mval'>{_fmt_m(ma)}</span></div>
  <div class='mrow'><span class='mlbl'>保证金使用率</span>
    <span class='{_margin_cls(mp)}'>{_fmt_p(mp)}</span></div>
  <div class='mrow'><span class='mlbl'>当日盈亏</span>
    <span class='{_pnl_cls(dp)}'>{_fmt_m(dp)}</span></div>
  <div style='color:{_MUTED};font-size:10px;margin-top:8px'>
    {'最后更新: ' + ts + ' ET' if ts else '⏳ 等待自动同步（运行 start_chrome.bat 并登录）'}</div>
</div>""", unsafe_allow_html=True)

# 保证金使用率所需数据
df_bal = _load_balance_history(730)
if not df_bal.empty:
    df_bal["sync_time"] = pd.to_datetime(df_bal["sync_time"], format="mixed", utc=True)
clr = {ACCT_CFG[0]["id"]: _GREEN, ACCT_CFG[1]["id"]: _BLUE}
lbl = {c["id"]: c["label"] for c in ACCT_CFG}

# 保证金使用率
st.markdown("#### 保证金使用率趋势")
if not df_bal.empty:
    df_mg = df_bal.dropna(subset=["margin_usage_pct"])
    if not df_mg.empty:
        fig2 = go.Figure()
        for aid, grp in df_mg.groupby("account_id"):
            grp = grp.sort_values("sync_time")
            fig2.add_trace(go.Scatter(
                x=grp["sync_time"], y=grp["margin_usage_pct"],
                mode="lines+markers", name=lbl.get(aid, aid),
                line=dict(color=clr.get(aid, _AMB), width=2.5), marker=dict(size=5),
                hovertemplate="%{x|%m-%d %H:%M}<br>%{y:.1f}%<extra></extra>"))
        fig2.add_hline(y=50, line=dict(color=_AMB, width=1, dash="dot"),
                       annotation_text="  50%", annotation_font_color=_AMB)
        fig2.add_hline(y=75, line=dict(color=_RED, width=1, dash="dot"),
                       annotation_text="  75%", annotation_font_color=_RED)
        fig2.update_layout(
            paper_bgcolor=_BG, plot_bgcolor=_BG, height=220,
            xaxis=dict(showgrid=False, tickfont=dict(color=_MUTED)),
            yaxis=dict(showgrid=True, gridcolor=_BORDER, tickfont=dict(color=_MUTED),
                       ticksuffix="%", range=[0,105]),
            legend=dict(font=dict(color=_TEXT), bgcolor=_BG, bordercolor=_BORDER,
                        orientation="h", y=1.1, x=1, xanchor="right"),
            margin=dict(l=10, r=10, t=20, b=10))
        st.plotly_chart(fig2, use_container_width=True)
    else:
        st.info("暂无保证金数据。")
else:
    st.info("暂无历史数据。")

# ════════════════════════════════════════════════════════
# 持仓明细 + 期权持仓管理
# ════════════════════════════════════════════════════════
st.divider()
st.markdown("#### 持仓明细")

_pos_tabs = st.tabs(["🏦 期权持仓", "📊 交易绩效", "📅 事件日历", "💡 交易建议", "📋 交易历史",
                     "🔍 数据核查"]
                    + [c["label"] for c in ACCT_CFG])

# ── 期权持仓 Tab ──────────────────────────────────────
with _pos_tabs[0]:
    _opt_acct_label = st.selectbox("账户", [c["label"] for c in ACCT_CFG], key="opt_acct")
    _opt_acct_id    = next((c["id"] for c in ACCT_CFG if c["label"] == _opt_acct_label), ACCT_CFG[0]["id"])

    df_opts = _load_options_positions(_opt_acct_id)

    # ── IV Regime 横幅 ──────────────────────────────────────────────
    _iv_regime = _compute_iv_regime(_opt_acct_id)
    _iv_status = _iv_regime["status"]
    _IV_BANNER = {
        "EXTREME_IV":           ("🔴", "极端 IV（≥95th pct）— 停止卖出新 vega", "error"),
        "HIGH_IV":              ("🟠", "高 IV（≥85th pct）— 适合卖权，注意尾部风险", "warning"),
        "NORMAL":               ("🟡", "IV 正常区间", "info"),
        "LOW_IV":               ("🟢", "低 IV（<30th pct）— 适合买权", "info"),
        "INSUFFICIENT_HISTORY": ("⏳", "IV 历史数据不足（需 ≥20 次观测）", "info"),
        "NO_DATA":              ("⚪", "暂无 IV 数据（更新报价后自动积累）", "info"),
    }
    _iv_icon, _iv_msg, _iv_cls = _IV_BANNER.get(_iv_status, ("⚪", _iv_status, "info"))
    _max_p = _iv_regime.get("max_piv")
    if _max_p:
        _iv_msg += f"  ·  最高 PIV: {_max_p['symbol']} {_max_p['piv']*100:.0f}%"
    getattr(st, _iv_cls)(f"{_iv_icon} **IV Regime** — {_iv_msg}")

    # ── 汇总指标（只统计有报价的合约，避免 None 污染总数）──
    if not df_opts.empty:
        _priced = df_opts[df_opts["current_price"].notna()]
        _no_price_cnt = df_opts["current_price"].isna().sum()
        _total_mv   = _priced["market_value"].fillna(0).sum()
        _total_dpnl = _priced["day_pnl"].fillna(0).sum()
        _total_pnl  = df_opts["total_pnl"].fillna(0).sum()
        _total_rz   = df_opts["realized_pnl"].fillna(0).sum() if "realized_pnl" in df_opts.columns else 0.0
        _mc1, _mc2, _mc3, _mc4, _mc5 = st.columns(5)
        _mc1.metric("期权总市值",   f"${_total_mv:,.0f}")
        _mc2.metric("当日盈亏",     f"${_total_dpnl:,.2f}", delta=f"{_total_dpnl:+,.2f}")
        _mc3.metric("未实现盈亏",   f"${_total_pnl:,.2f}",  delta=f"{_total_pnl:+,.2f}")
        _mc4.metric("已实现盈亏",   f"${_total_rz:,.2f}",   delta=f"{_total_rz:+,.2f}")
        _mc5.metric("合约数",       f"{len(_priced)}/{len(df_opts)} 有报价",
                    delta=f"-{_no_price_cnt} 无报价" if _no_price_cnt else None)
        st.markdown("")

    # ── 组合视图（Portfolio View）───────────────────────────────────
    if not df_opts.empty:
        _spread_data = _build_spread_portfolios(_opt_acct_id)
        if _spread_data:
            st.markdown("##### 📊 组合结构识别")

            # ── 聚合辅助 ──────────────────────────────────────────
            _by_und: dict = {}
            for _port in _spread_data:
                _by_und.setdefault(_port["underlying"], []).append(_port)

            _RISK_RANK = {"LOW": 0, "MEDIUM": 1, "HIGH": 2, "CRITICAL": 3}
            _RISK_ICON = {"CRITICAL": "🔴", "HIGH": "🟠", "MEDIUM": "🟡", "LOW": "🟢"}
            _STYPE_SHORT = {
                "Bull Call Spread": "Bull Spread",
                "Bear Call Spread": "Bear Spread",
                "Bear Put Spread":  "Bear Spread",
                "Bull Put Spread":  "Bull Spread",
                "Calendar Spread":  "Calendar",
                "Diagonal Spread (LEAPS)": "LEAPS Diag",
                "Naked Long Call":  "Long Call",
                "Naked Long Put":   "Long Put",
                "Naked Short Call": "Short Call ⚠️",
                "Naked Short Put":  "Short Put ⚠️",
                "Reversed Diagonal (⚠️ 买腿先到期)": "⚠️ Rev Diag",
            }

            def _worst_rl(ports):
                return max(ports,
                           key=lambda p: _RISK_RANK.get(p.get("risk_level", "LOW"), 0)
                           ).get("risk_level", "LOW")

            # ── 区块1：总览指标 ──────────────────────────────────
            _total_exposure = sum(abs(_p.get("net_total") or 0) for _p in _spread_data)
            _ml_list = [_p.get("max_loss")   for _p in _spread_data]
            _mp_list = [_p.get("max_profit") for _p in _spread_data]
            _total_ml      = sum(x for x in _ml_list if x is not None)
            _total_mp      = sum(x for x in _mp_list if x is not None)
            _has_unlimited = any(x is None   for x in _mp_list)
            _total_pnl_sum = sum(_p.get("current_pnl") or 0 for _p in _spread_data)

            _k1, _k2, _k3, _k4 = st.columns(4)
            _k1.metric("📊 总净权利金", f"${_total_exposure:,.0f}")
            _k2.metric("⬇️ 总最大亏损", f"${_total_ml:,.0f}")
            _k3.metric("⬆️ 总最大盈利",
                       "∞ 无限" if _has_unlimited else f"${_total_mp:,.0f}")
            _k4.metric("💰 当前总盈亏", f"${_total_pnl_sum:+,.0f}")

            st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

            # ── 区块2：按标的汇总（折叠行）──────────────────────────
            for _und in sorted(_by_und):
                _ports = _by_und[_und]
                _und_ml      = sum(_p.get("max_loss") or 0    for _p in _ports)
                _und_mp_list = [_p.get("max_profit")           for _p in _ports]
                _und_unl     = any(_x is None for _x in _und_mp_list)
                _und_mp      = sum(_x for _x in _und_mp_list  if _x is not None)
                _und_pnl     = sum(_p.get("current_pnl") or 0 for _p in _ports)
                _und_dte     = min(_p.get("dte") or 9999       for _p in _ports)
                _wrl         = _worst_rl(_ports)
                _ri          = _RISK_ICON.get(_wrl, "⚪")

                _stypes = list(dict.fromkeys(
                    _STYPE_SHORT.get(_p["type"], _p["type"]) for _p in _ports))
                _strat_str = " + ".join(_stypes)
                _ml_str = f"${_und_ml:,.0f}" if _und_ml else "—"
                _mp_str = "无限" if _und_unl else (f"${_und_mp:,.0f}" if _und_mp else "—")

                _exp_label = (
                    f"{_ri} **{_und}** — {_strat_str}"
                    f"  |  ML {_ml_str}  MP {_mp_str}"
                    f"  PnL ${_und_pnl:+,.0f}  DTE {_und_dte}"
                )
                with st.expander(_exp_label, expanded=(_wrl in ("CRITICAL", "HIGH"))):
                    for _port in _ports:
                        st.caption(f"**{_port['type']}** — {_port['recommendation']}")
                        _legs = _port.get("legs", [])
                        if _legs:
                            _leg_lines = []
                            for _lg in _legs:
                                _side = "买入" if _lg.get("qty", 0) > 0 else "卖出"
                                _leg_lines.append(
                                    f"  {_side} {_lg.get('direction','').capitalize()} "
                                    f"${_lg.get('strike', 0):.1f} "
                                    f"到期 {_lg.get('expiry', '—')} "
                                    f"× {abs(_lg.get('qty', 0))} "
                                    f"成本 ${_lg.get('unit_cost', 0):.2f}"
                                )
                            st.markdown("  \n".join(_leg_lines))

            # ── 区块3：风险集中度 ─────────────────────────────────
            if _total_exposure > 0:
                st.markdown(
                    f"<div style='font-size:10px;color:{_MUTED};text-transform:uppercase;"
                    f"letter-spacing:1px;margin:10px 0 4px'>⚠️ 风险集中度（净权利金占比）</div>",
                    unsafe_allow_html=True,
                )
                _und_exp = {
                    _u: sum(abs(_p.get("net_total") or 0) for _p in _ps)
                    for _u, _ps in _by_und.items()
                }
                for _u, _ux in sorted(_und_exp.items(), key=lambda kv: -kv[1]):
                    _pct = _ux / _total_exposure
                    _bc  = (_RED if _worst_rl(_by_und[_u]) == "CRITICAL" else
                            _AMB if _worst_rl(_by_und[_u]) == "HIGH"     else _GREEN)
                    st.markdown(
                        f"<div style='display:flex;align-items:center;gap:10px;"
                        f"margin-bottom:2px'>"
                        f"<span style='width:52px;font-size:12px;font-weight:700;"
                        f"color:{_TEXT}'>{_u}</span>"
                        f"<span style='font-size:11px;color:{_MUTED}'>{_pct*100:.0f}%</span>"
                        f"<span style='font-size:11px;color:{_bc}'>${_ux:,.0f}</span>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )
                    st.progress(_pct)

            st.markdown("---")

    # ── 持仓明细（只读，带状态 + 到期预警 + AI 评分）──
    if not df_opts.empty:
        _ai_scores = _load_ai_scores()
        _disp = df_opts[["symbol","direction","strike","expiry",
                          "quantity","unit_cost","current_price",
                          "market_value","day_pnl","total_pnl",
                          "realized_pnl"] +
                         (["iv","delta","theta"] if all(c in df_opts.columns
                          for c in ["iv","delta","theta"]) else [])].copy()

        # AI 评分列（根据标的）
        def _get_score_label(sym):
            mo = _OCC_RE.match(str(sym).upper())
            und = mo.group(1) if mo else str(sym).upper()
            s = _ai_scores.get(und)
            return _score_label(s) if s is not None else ""
        _disp["AI评分"] = _disp["symbol"].apply(_get_score_label)

        _disp["状态"]    = _disp["current_price"].apply(
            lambda v: "" if (v is not None and not pd.isna(v)) else "❓ 无报价")
        _disp["到期预警"] = _disp["expiry"].apply(_expiry_badge)
        _disp = _disp.rename(columns={
            "symbol":"代号","direction":"方向","strike":"行权价","expiry":"到期日",
            "quantity":"数量","unit_cost":"单位成本","current_price":"现价",
            "market_value":"市值","day_pnl":"当日盈亏$","total_pnl":"未实现$",
            "realized_pnl":"已实现$",
            "iv":"IV","delta":"Delta","theta":"Theta",
        })
        _base_cols = ["代号","方向","行权价","到期日","数量","单位成本",
                      "现价","状态","市值","当日盈亏$","未实现$","已实现$","AI评分","到期预警"]
        _greek_cols = ["IV","Delta","Theta"]
        _col_order = _base_cols + [c for c in _greek_cols if c in _disp.columns]
        _cfg = {
            "行权价":    st.column_config.NumberColumn(format="$%.3f"),
            "单位成本":  st.column_config.NumberColumn(format="$%.2f"),
            "现价":      st.column_config.NumberColumn(format="$%.3f"),
            "市值":      st.column_config.NumberColumn(format="$%.0f"),
            "当日盈亏$": st.column_config.NumberColumn(format="$%.2f"),
            "未实现$":   st.column_config.NumberColumn(format="$%.2f"),
            "已实现$":   st.column_config.NumberColumn(format="$%.2f"),
            "IV":        st.column_config.NumberColumn(format="%.1f%%"),
            "Delta":     st.column_config.NumberColumn(format="%.3f"),
            "Theta":     st.column_config.NumberColumn(format="%.4f"),
        }
        st.dataframe(
            _disp[_col_order], use_container_width=True, hide_index=True,
            height=min(45 + len(_disp) * 38, 520),
            column_config=_cfg)

    st.markdown("---")
    st.caption(f"📡 期权行情更新请使用左侧边栏「📈 更新行情」  ·  API Key: {'✅ 已配置' if _MD_KEY else '❌ 未配置'}")

    # ── 风险快照 ────────────────────────────────────────────────────
    st.markdown("#### ⚡ 风险快照（Beta-Delta / 压力测试）")
    with st.container():
        _snap = _compute_risk_snapshot(_opt_acct_id)
        if "error" in _snap:
            st.info("暂无账户净值数据，请先完成同步。")
        else:
            # IV alert check — runs first so alerts appear at top
            _iv_data = _fetch_iv_monitor_batch(_IV_WATCH_SYMS)
            _iv_hot  = [(s, d["ivr"]) for s, d in _iv_data.items()
                        if (d.get("ivr") or 0) >= 75]
            for _s, _ivr in sorted(_iv_hot, key=lambda x: -x[1]):
                st.error(f"🔴 **{_s}** IV极高（IVR {_ivr:.0f}%），当前是卖权对冲最佳时机")

            # ── 第一行：账户净值 / 毛敞口 / 杠杆率 / 每日Theta ──────
            _rs1, _rs2, _rs3, _rs4 = st.columns(4)
            with _rs1:
                st.metric("账户净值", f"${_snap['equity']:,.0f}")
                st.caption("当前账户总资产净值")
            with _rs2:
                _dn = _snap.get('delta_notional') or 0
                st.metric("Delta 敞口", f"${_dn:,.0f}")
                st.caption(f"Delta 口径名义风险（价差不双计）｜毛敞口 ${_snap['gross_notional']:,.0f}")
            with _rs3:
                _lev_d    = _snap.get('leverage_delta') or 0
                _lev_over = _lev_d > _RISK_LIMITS["max_leverage"]
                st.metric("杠杆率（Δ口径）",
                          f"{_lev_d:.1f}x" if _lev_d else "—",
                          delta="⚠️ 超限" if _lev_over else None)
                _lev_raw = _snap.get('leverage') or 0
                st.caption(f"Delta 敞口 ÷ 净值｜毛口径 {_lev_raw:.1f}x（价差两腿各自计入）")
            with _rs4:
                _th = _snap['theta_per_day']
                st.metric("每日 Theta", f"${_th:+,.0f}")
                st.caption(f"时间流逝每天{'损失' if _th < 0 else '收取'}\\${abs(_th):,.0f}，月约\\${abs(_th)*30:,.0f}")

            # ── 第二行：Beta加权Delta / Vega / 压力-10% / 压力-20% ──
            _rs5, _rs6, _rs7, _rs8 = st.columns(4)
            _bdr = _snap.get('beta_delta_ratio') or 0
            _bdr_over = abs(_bdr) > _RISK_LIMITS["max_beta_delta_ratio"]
            with _rs5:
                st.metric("Beta 加权 Delta",
                          f"{_bdr*100:.1f}%" if _bdr else "—",
                          delta="⚠️ 超限" if _bdr_over else None)
                _dir = "偏多方向" if _bdr >= 0 else "偏空方向"
                st.caption(f"相当于持有{abs(_bdr):.2f}倍市场风险，{_dir}")
            with _rs6:
                st.metric("Vega 敞口", f"${_snap['vega_per_pt']:+,.0f}")
                st.caption("IV每变动1个百分点，组合盈亏变化该金额")

            _stress10_r = _snap.get("stress_10_ratio") or 0
            _stress20_r = _snap.get("stress_20_ratio") or 0
            _s10 = _snap['stress_10']
            _s20 = _snap['stress_20']
            with _rs7:
                st.metric("压力 -10%",
                          f"{_stress10_r*100:+.1f}%",
                          delta=f"${_s10:+,.0f}" + (" ⚠️" if abs(_stress10_r) >= _RISK_LIMITS["stress_warning"] else ""))
                st.caption(f"市场跌10%时，账户约{'亏损' if _s10 < 0 else '盈利'}${abs(_s10):,.0f}")
            with _rs8:
                st.metric("压力 -20%",
                          f"{_stress20_r*100:+.1f}%",
                          delta=f"${_s20:+,.0f}")
                if _s20 >= 0:
                    st.caption(f"市场跌20%，账户反而盈利${_s20:,.0f}（对冲生效）")
                else:
                    st.caption(f"市场跌20%时，账户约亏损${abs(_s20):,.0f}")

            # ── 第三行：Gamma / 最近到期 / 最大回撤(TWR) ──────────
            _rs9, _rs10, _rs11, _rs12 = st.columns(4)
            _gamma = _snap.get("gamma_total", 0)
            with _rs9:
                st.metric("总 Gamma", f"{_gamma:.3f}")
                st.caption(f"标的每变动$1，Delta约变化{_gamma:.1f}；接近到期时急剧加速")
            _nexp_d = _snap.get("nearest_expiry_date")
            _nexp_s = _snap.get("nearest_expiry_sym", "")
            with _rs10:
                if _nexp_d:
                    _days_to_exp = (_nexp_d - datetime.date.today()).days
                    st.metric("最近到期",
                              f"{_nexp_s} {_nexp_d.strftime('%m/%d')}",
                              delta=f"还有 {_days_to_exp} 天")
                else:
                    st.metric("最近到期", "—")
                st.caption("DTE≤21开始减仓评估，DTE≤14立即处理")
            _dd_pct_disp = _snap.get("drawdown", 0) * 100
            with _rs11:
                st.metric("历史最大回撤", f"{_dd_pct_disp:.1f}%")
                st.caption("历史业绩数据（仅供参考），不反映当前持仓风险")

            # ── 第四行：期权成本/净值比率 ──────────────────────────
            _ocr = _compute_options_cost_ratio(_opt_acct_id)
            _ocr_light = {"red": "🔴", "amber": "🟡", "green": "🟢", "gray": "⚪"}.get(_ocr["status"], "⚪")
            _ocr_cols = st.columns(4)
            if _ocr["ratio"] is not None:
                _ocr_pct   = f"{_ocr['ratio']*100:.1f}%"
                _ocr_delta = (f"${_ocr['total_cost']:,.0f} / ${_ocr['nav']:,.0f}  "
                              f"{_ocr_light} {'超限' if _ocr['status'] in ('red','amber') else '正常'}")
            else:
                _ocr_pct   = "—"
                _ocr_delta = "暂无数据"
            with _ocr_cols[0]:
                st.metric("期权成本/净值比率", _ocr_pct, delta=_ocr_delta)
                st.caption("期权总建仓成本占账户净值的比例，超限则风险过于集中")
            with _ocr_cols[1]:
                st.metric("成本比率上限", f"{_ocr['limit']*100:.0f}%",
                          delta="来自 Portfolio_Config")
                st.caption("在 Portfolio_Config.json 的 options_cost_ratio_limit 字段修改")

            # ── 当前持仓风险警告（仅基于当前希腊值/持仓状态）────────
            _risk_st = _snap["risk_status"]
            _RISK_MSG = {
                "RED_HARD_STOP":  ("🔴", "压力损失 ≥15% — 强制减仓", "error"),
                "ORANGE_DE_RISK": ("🟠", "压力损失 ≥12% — 建议减仓审查", "warning"),
                "YELLOW_WARNING": ("🟡", "压力损失 ≥8% — 关注风险", "warning"),
                "GREEN":          None,
            }
            _risk_info = _RISK_MSG.get(_risk_st)
            if _risk_info:
                _ri, _rm, _rc = _risk_info
                getattr(st, _rc)(f"{_ri} **压力测试警告** — {_rm}")

            if _bdr_over:
                _bdr_pct = abs(_bdr) * 100
                st.warning(f"🟠 **Beta-Delta 超限** — 当前 {_bdr_pct:.1f}%，"
                           f"超过限额 {_RISK_LIMITS['max_beta_delta_ratio']*100:.0f}%，建议对冲")

            if _ocr["ratio"] is not None and _ocr["status"] in ("red", "amber"):
                _ocr_limit_pct = _ocr["limit"] * 100
                st.warning(f"🟠 **期权成本超限** — 当前 {_ocr['ratio']*100:.1f}%"
                           f"（上限 {_ocr_limit_pct:.0f}%），集中度过高")

            if _nexp_d:
                _dte_warn = (_nexp_d - datetime.date.today()).days
                if _dte_warn <= 14:
                    st.error(f"🔴 **临近到期** — {_nexp_s} 还有 {_dte_warn} 天到期，须立即处理")
                elif _dte_warn <= 21:
                    st.warning(f"🟠 **即将到期** — {_nexp_s} 还有 {_dte_warn} 天，开始减仓评估")

            # ── IV 监控状态 ──────────────────────────────────────────
            st.markdown("---")
            st.markdown("**📡 IV 监控状态**（15分钟缓存）")
            _iv_cols = st.columns(len(_IV_WATCH_SYMS))
            for _ic, _isym in zip(_iv_cols, _IV_WATCH_SYMS):
                _idat = _iv_data.get(_isym, {})
                _iv_v  = _idat.get("iv")
                _ivr_v = _idat.get("ivr")
                with _ic:
                    _err  = _idat.get("error")
                    _days = _idat.get("iv_days", 0)
                    if _iv_v is None:
                        _err_short = (_err or "获取中")[:22]
                        st.metric(_isym, "—", delta=_err_short)
                        if _err:
                            st.caption(_err[:60])
                    else:
                        if _ivr_v is not None and _ivr_v >= 75:
                            _iv_tag = f"IVR {_ivr_v:.0f}% 🔴"
                        elif _ivr_v is not None and _ivr_v >= 50:
                            _iv_tag = f"IVR {_ivr_v:.0f}% 🟡"
                        elif _ivr_v is not None:
                            _iv_tag = f"IVR {_ivr_v:.0f}% ⚪"
                        else:
                            _note = _idat.get("note") or ""
                            _iv_tag = f"IVR积累中({_days}天)" if _days < 30 else "IVR —"
                        st.metric(_isym, f"IV {_iv_v:.1f}%", delta=_iv_tag)

    # ── 已实现期权交易记录 ─────────────────────────────────────────
    _df_realized = _load_realized_trades(_opt_acct_id)
    if not _df_realized.empty:
        with st.expander(f"📋 已实现期权交易 ({len(_df_realized)} 笔，合计 "
                         f"${_df_realized['realized_pnl'].sum():+,.2f})", expanded=False):
            _wins = (_df_realized["win_loss"] == "win").sum()
            _tot  = len(_df_realized)
            st.caption(f"胜率 {_wins}/{_tot} = {_wins/_tot*100:.0f}%  ·  "
                       f"平均持仓 {_df_realized['holding_days'].mean():.0f} 天")
            _rz_disp = _df_realized[["underlying","strategy_type","open_date","close_date",
                                      "holding_days","quantity","open_cash","close_cash",
                                      "realized_pnl","return_on_risk","win_loss"]].copy()
            _rz_disp = _rz_disp.rename(columns={
                "underlying":"标的","strategy_type":"策略","open_date":"开仓日",
                "close_date":"平仓日","holding_days":"天数","quantity":"手数",
                "open_cash":"开仓价值","close_cash":"平仓价值",
                "realized_pnl":"已实现$","return_on_risk":"回报/风险","win_loss":"结果"
            })
            st.dataframe(_rz_disp, use_container_width=True, hide_index=True,
                         height=min(60 + len(_rz_disp) * 36, 400),
                         column_config={
                             "开仓价值":   st.column_config.NumberColumn(format="$%.2f"),
                             "平仓价值":   st.column_config.NumberColumn(format="$%.2f"),
                             "已实现$":    st.column_config.NumberColumn(format="$%.2f"),
                             "回报/风险":  st.column_config.NumberColumn(format="%.2%"),
                         })

    # ── 编辑区 ──
    with st.expander("✏️ 手动录入 / 编辑期权持仓", expanded=df_opts.empty):
        _bc1, _bc2 = st.columns([1, 1])
        with _bc1:
            if st.button("🔄 从交易记录自动导入", key="opt_derive",
                         help="根据交易历史净算未平仓合约，已有数据不覆盖"):
                _derived = _derive_open_options(_opt_acct_id)
                _existing_syms = set(df_opts["symbol"].tolist()) if not df_opts.empty else set()
                _new = [r for r in _derived if r["symbol"] not in _existing_syms]
                if _new:
                    _save_options_positions(_opt_acct_id, _new)
                    st.success(f"导入 {len(_new)} 个未平仓合约")
                    st.rerun()
                else:
                    st.info("没有新的未平仓合约（已全部存在）")

        # 编辑表格
        _edit_cols = {
            "symbol":        st.column_config.TextColumn("代号", required=True),
            "direction":     st.column_config.SelectboxColumn("方向", options=["Call","Put"], required=True),
            "strike":        st.column_config.NumberColumn("行权价", format="$%.2f", min_value=0.0),
            "expiry":        st.column_config.TextColumn("到期日(YYYY-MM-DD)", max_chars=10),
            "quantity":      st.column_config.NumberColumn("数量(负=卖出)", step=1),
            "unit_cost":     st.column_config.NumberColumn("单位成本", format="$%.2f", min_value=0.0),
            "current_price": st.column_config.NumberColumn("现价", format="$%.2f", min_value=0.0),
            "market_value":  st.column_config.NumberColumn("市值$", format="$%.0f"),
            "day_pnl":       st.column_config.NumberColumn("当日盈亏$", format="$%.2f"),
            "total_pnl":     st.column_config.NumberColumn("总盈亏$", format="$%.2f"),
        }
        _edit_df = df_opts[list(_edit_cols.keys())].copy() if not df_opts.empty else pd.DataFrame(columns=list(_edit_cols.keys()))
        _edited = st.data_editor(
            _edit_df,
            column_config=_edit_cols,
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            key="opt_editor",
            height=min(80 + len(_edit_df) * 38, 420),
        )
        if st.button("💾 保存持仓", key="opt_save", type="primary"):
            _rows = _edited.dropna(subset=["symbol"]).to_dict("records")
            if _rows:
                _save_options_positions(_opt_acct_id, _rows)
                st.success(f"已保存 {len(_rows)} 条期权持仓")
                st.rerun()
            else:
                st.warning("请至少填写一行数据（代号为必填）")

# ── 📊 交易绩效 Tab ───────────────────────────────────────────────
with _pos_tabs[1]:
    _perf_acct_label = st.selectbox("账户", [c["label"] for c in ACCT_CFG], key="perf_acct")
    _perf_acct_id    = next((c["id"] for c in ACCT_CFG if c["label"] == _perf_acct_label),
                            ACCT_CFG[0]["id"])
    _pf_col1, _pf_col2 = st.columns([1, 3])
    with _pf_col1:
        if st.button("🔄 运行 FIFO 分析", key="perf_fifo",
                     help="重新计算 FIFO 成本匹配和已实现盈亏（不需要 API Key）"):
            with st.spinner("FIFO 成本匹配…"):
                _fifo_r = _fifo_match_options(_perf_acct_id)
            st.success(f"FIFO 完成：{_fifo_r['realized_count']} 笔已实现，"
                       f"已实现 ${_fifo_r['total_realized_pnl']:+,.2f}，"
                       f"胜率 {_fifo_r['wins']}/{_fifo_r['realized_count']}")
            st.rerun()

    _stats = _compute_performance_stats(_perf_acct_id)

    if _stats is None:
        st.info("暂无已实现交易数据。点击上方「🔄 运行 FIFO 分析」或到期权持仓点「📡 更新现价 + 成本」。")
    else:
        # ── 汇总指标 ──
        def _kpi(col, label, value, color=_TEXT):
            col.markdown(
                f"<div style='background:{_SURF};border:1px solid {_BORDER};border-radius:8px;"
                f"padding:6px 4px;text-align:center;min-width:0;overflow:hidden'>"
                f"<div style='font-size:9px;color:{_MUTED};margin-bottom:3px;white-space:nowrap;"
                f"overflow:hidden;text-overflow:ellipsis'>{label}</div>"
                f"<div style='font-size:0.85rem;font-weight:700;color:{color};"
                f"line-height:1.3;word-break:break-word'>{value}</div>"
                f"</div>",
                unsafe_allow_html=True,
            )

        _pnl_color = _GREEN if _stats["total_pnl"] >= 0 else _RED
        _loss_color = _RED if _stats["max_loss"] < 0 else _TEXT
        _pf_val = (f"{_stats['profit_factor']:.2f}x"
                   if _stats["profit_factor"] else "—")

        _p1, _p2, _p3, _p4, _p5, _p6, _p7 = st.columns(7)
        _kpi(_p1, "总已实现盈亏",
             f"${_stats['total_pnl']:+,.0f}", _pnl_color)
        _kpi(_p2, "交易次数",
             str(_stats["count"]))
        _kpi(_p3, "整体胜率",
             f"{_stats['win_rate']*100:.1f}%")
        _kpi(_p4, "盈亏比", _pf_val,
             _GREEN if _stats["profit_factor"] and _stats["profit_factor"] >= 1 else _RED)
        _kpi(_p5, "最大单笔亏损",
             f"${_stats['max_loss']:,.0f}", _loss_color)
        _kpi(_p6, "近20笔胜率",
             f"{_stats['latest_wr20']*100:.1f}%")
        _kpi(_p7, "EWMA胜率(λ.94)",
             f"{_stats['ewma_wr']*100:.1f}%",
             _GREEN if _stats["ewma_wr"] >= 0.5 else _RED)
        st.markdown("")

        _df_c = _stats["df"]

        # ── 累计 P&L ──
        _fig_cum = go.Figure()
        _fig_cum.add_trace(go.Scatter(
            x=_df_c["trade_num"], y=_df_c["cumulative_pnl"],
            mode="lines+markers", name="累计盈亏",
            line=dict(color=_GREEN if _stats["total_pnl"] >= 0 else _RED, width=2.5),
            marker=dict(size=5),
            customdata=_df_c[["underlying", "strategy_type"]],
            hovertemplate="第%{x}笔 %{customdata[0]} %{customdata[1]}<br>累计 $%{y:+,.2f}<extra></extra>",
        ))
        _fig_cum.add_hline(y=0, line=dict(color=_BORDER, width=1, dash="dot"))
        _fig_cum.update_layout(
            title="累计 P&L 曲线", paper_bgcolor=_BG, plot_bgcolor=_BG, height=270,
            xaxis=dict(showgrid=False, tickfont=dict(color=_MUTED), title="交易序号"),
            yaxis=dict(showgrid=True, gridcolor=_BORDER, tickfont=dict(color=_MUTED),
                       tickprefix="$", tickformat=",.0f"),
            margin=dict(l=10, r=10, t=36, b=10))

        # ── 滚动胜率 ──
        _fig_wr = go.Figure()
        _fig_wr.add_trace(go.Scatter(
            x=_df_c["trade_num"], y=_df_c["rolling20_wr"] * 100,
            mode="lines", name="滚动20笔",
            line=dict(color=_GREEN, width=2),
            hovertemplate="第%{x}笔  %{y:.1f}%<extra></extra>"))
        _fig_wr.add_trace(go.Scatter(
            x=_df_c["trade_num"], y=_df_c["rolling50_wr"] * 100,
            mode="lines", name="滚动50笔",
            line=dict(color=_AMB, width=1.5, dash="dash"),
            hovertemplate="第%{x}笔  %{y:.1f}%<extra></extra>"))
        _fig_wr.add_hline(y=50, line=dict(color=_BORDER, width=1, dash="dot"),
                          annotation_text="  50%", annotation_font_color=_MUTED)
        _fig_wr.update_layout(
            title="滚动胜率", paper_bgcolor=_BG, plot_bgcolor=_BG, height=270,
            xaxis=dict(showgrid=False, tickfont=dict(color=_MUTED)),
            yaxis=dict(showgrid=True, gridcolor=_BORDER, tickfont=dict(color=_MUTED),
                       ticksuffix="%", range=[0, 105]),
            legend=dict(font=dict(color=_TEXT), bgcolor=_BG, orientation="h", y=1.15),
            margin=dict(l=10, r=10, t=36, b=10))

        # ── 回撤 ──
        _fig_dd = go.Figure()
        _fig_dd.add_trace(go.Scatter(
            x=_df_c["trade_num"], y=_df_c["drawdown"],
            mode="lines", fill="tozeroy", name="回撤",
            line=dict(color=_RED, width=1.5),
            fillcolor="rgba(255,75,110,0.12)",
            hovertemplate="第%{x}笔  $%{y:,.2f}<extra></extra>"))
        _fig_dd.update_layout(
            title="P&L 峰值回撤", paper_bgcolor=_BG, plot_bgcolor=_BG, height=270,
            xaxis=dict(showgrid=False, tickfont=dict(color=_MUTED)),
            yaxis=dict(showgrid=True, gridcolor=_BORDER, tickfont=dict(color=_MUTED),
                       tickprefix="$", tickformat=",.0f"),
            margin=dict(l=10, r=10, t=36, b=10))

        _lc1, _lc2 = st.columns(2)
        with _lc1:
            st.plotly_chart(_fig_cum, use_container_width=True)
        with _lc2:
            st.plotly_chart(_fig_wr,  use_container_width=True)

        _lc3, _lc4 = st.columns(2)
        with _lc3:
            st.plotly_chart(_fig_dd, use_container_width=True)
        with _lc4:
            st.markdown("**按策略分组**")
            if _stats["by_strat"]:
                _strat_rows = pd.DataFrame([
                    {"策略": k, "次数": v["count"],
                     "胜率": f"{v['win_rate']*100:.0f}%",
                     "均盈亏": f"${v['avg_pnl']:+,.2f}",
                     "合计":   f"${v['total']:+,.2f}"}
                    for k, v in sorted(_stats["by_strat"].items(), key=lambda x: -x[1]["total"])
                ])
                st.dataframe(_strat_rows, use_container_width=True, hide_index=True, height=240)

        st.markdown("**按标的分组**")
        if _stats["by_und"]:
            _und_rows = pd.DataFrame([
                {"标的": k, "次数": v["count"],
                 "胜率": f"{v['win_rate']*100:.0f}%",
                 "均盈亏": f"${v['avg_pnl']:+,.2f}",
                 "合计":   f"${v['total']:+,.2f}"}
                for k, v in sorted(_stats["by_und"].items(), key=lambda x: -x[1]["total"])
            ])
            st.dataframe(_und_rows, use_container_width=True, hide_index=True, height=220)

        st.caption(f"平均盈利: ${_stats['avg_win']:+,.2f}  ·  "
                   f"平均亏损: ${_stats['avg_loss']:+,.2f}  ·  "
                   f"最大单笔盈利: ${_stats['max_gain']:+,.2f}  ·  "
                   f"最大回撤: ${_stats['max_drawdown']:+,.2f}")

        # ── 组合策略分析 ──
        st.markdown("**按组合策略分组**")
        if _stats.get("by_combo"):
            _combo_rows = pd.DataFrame([
                {"组合策略": k, "次数": v["count"],
                 "胜率": f"{v['win_rate']*100:.0f}%",
                 "均盈亏": f"${v['avg_pnl']:+,.2f}",
                 "合计":   f"${v['total']:+,.2f}"}
                for k, v in sorted(_stats["by_combo"].items(), key=lambda x: -x[1]["total"])
            ])
            st.dataframe(_combo_rows, use_container_width=True, hide_index=True,
                         height=min(60 + len(_combo_rows) * 38, 400))

        # ── 交易明细 ──
        st.divider()
        with st.expander(f"📋 逐笔交易明细（共 {_stats['count']} 笔）", expanded=False):
            _df_detail = _stats["df"].copy()
            # 筛选器
            _flt_col1, _flt_col2, _flt_col3 = st.columns(3)
            _und_opts = ["全部"] + sorted(_df_detail["underlying"].dropna().unique().tolist())
            _strat_opts = ["全部"] + sorted(_df_detail["strategy_type"].dropna().unique().tolist())
            _sel_und   = _flt_col1.selectbox("标的", _und_opts,   key="det_und")
            _sel_strat = _flt_col2.selectbox("策略", _strat_opts, key="det_strat")
            _sel_wl    = _flt_col3.selectbox("盈亏", ["全部", "盈利", "亏损"], key="det_wl")

            if _sel_und   != "全部": _df_detail = _df_detail[_df_detail["underlying"] == _sel_und]
            if _sel_strat != "全部": _df_detail = _df_detail[_df_detail["strategy_type"] == _sel_strat]
            if _sel_wl == "盈利":   _df_detail = _df_detail[_df_detail["pnl"] > 0]
            elif _sel_wl == "亏损": _df_detail = _df_detail[_df_detail["pnl"] < 0]

            _show_cols = ["close_date", "underlying", "strategy_type", "combo_strategy",
                          "quantity", "open_cash", "close_cash", "pnl",
                          "holding_days", "win_loss"]
            _show_cols = [c for c in _show_cols if c in _df_detail.columns]
            _col_labels = {
                "close_date": "平仓日", "underlying": "标的",
                "strategy_type": "策略类型", "combo_strategy": "组合策略",
                "quantity": "数量", "open_cash": "开仓权利金",
                "close_cash": "平仓权利金", "pnl": "盈亏$",
                "holding_days": "持仓天", "win_loss": "胜负",
            }
            _df_show = _df_detail[_show_cols].rename(columns=_col_labels).sort_values(
                "平仓日", ascending=False)
            st.dataframe(
                _df_show, use_container_width=True, hide_index=True,
                height=min(60 + len(_df_show) * 38, 480),
                column_config={
                    "盈亏$": st.column_config.NumberColumn("盈亏$", format="$%.2f"),
                    "开仓权利金": st.column_config.NumberColumn("开仓权利金", format="$%.2f"),
                    "平仓权利金": st.column_config.NumberColumn("平仓权利金", format="$%.2f"),
                }
            )
            _flt_total = _df_detail["pnl"].sum()
            _flt_wr    = (_df_detail["pnl"] > 0).mean() * 100 if len(_df_detail) else 0
            st.caption(f"筛选结果: {len(_df_detail)} 笔  合计盈亏 ${_flt_total:+,.2f}  "
                       f"胜率 {_flt_wr:.0f}%")

        # ── QQQ 对比（切换视图）──
        st.divider()
        _bm_sel_col, _bm_alpha_col = st.columns([3, 1])
        with _bm_sel_col:
            _bm_view = st.selectbox(
                "对比视图",
                ["📈 历史表现（过去1年）", "🤖 AI赋能对比（2026-06-16起）"],
                key="bm_view_sel", label_visibility="collapsed")

        # ── 视图一：历史表现（TWR）──
        if _bm_view.startswith("📈"):
            with st.spinner("计算 TWR 收益率 / 拉取 QQQ 历史数据…"):
                _h_acct, _h_cashflows = _compute_twr_series(_perf_acct_id, _HIST_START)
                _today_plus1 = (datetime.datetime.now(_ET) +
                                datetime.timedelta(days=1)).strftime("%Y-%m-%d")
                _h_qqq_map = _fetch_qqq_range(_HIST_START, _today_plus1)

            _qqq_rows_h = [{"date": d, "close": _h_qqq_map[d]}
                           for d in sorted(_h_qqq_map) if d >= _HIST_START]
            _h_qqq = pd.DataFrame(_qqq_rows_h)
            if not _h_qqq.empty:
                _base_qqq_h = float(_h_qqq["close"].iloc[0])
                _h_qqq["rel_pct"] = (_h_qqq["close"] / _base_qqq_h - 1) * 100

            _acct_last = float(_h_acct["twr_pct"].iloc[-1]) if not _h_acct.empty else 0.0
            _qqq_last  = float(_h_qqq["rel_pct"].iloc[-1])  if not _h_qqq.empty  else 0.0
            _alpha_val = _acct_last - _qqq_last
            with _bm_alpha_col:
                _kpi(st, "超额Alpha",
                     f"{_alpha_val:+.1f}%",
                     _GREEN if _alpha_val >= 0 else _RED)

            if _h_acct.empty:
                st.info("暂无历史 NAV 数据（需先运行一次 CDP 同步）。")
            else:
                # 最大回撤（TWR growth-factor相对计算，非百分点差）
                _h_growth  = 1 + pd.Series(_h_acct["twr_pct"].values) / 100
                _h_gmax    = _h_growth.cummax()
                _h_dd_raw  = _h_growth / _h_gmax - 1   # 相对跌幅，负数
                _h_dd_idx  = int(_h_dd_raw.idxmin())
                _h_dd_pct  = float(_h_dd_raw.iloc[_h_dd_idx]) * 100  # 如 -12.3
                _h_dd_date = _h_acct["date"].iloc[_h_dd_idx]
                _h_last_pct  = float(_h_acct["twr_pct"].iloc[-1])
                _h_last_date = _h_acct["date"].iloc[-1]
                _h_color = _GREEN if _h_last_pct >= 0 else _RED

                _fig_bm = go.Figure()
                _fig_bm.add_trace(go.Scatter(
                    x=_h_acct["date"], y=_h_acct["twr_pct"],
                    mode="lines", name="ENERGREX (TWR)",
                    line=dict(color=_h_color, width=2.5),
                    hovertemplate="%{x}  %{y:+.2f}%<extra>ENERGREX</extra>"))
                if not _h_qqq.empty:
                    _fig_bm.add_trace(go.Scatter(
                        x=_h_qqq["date"], y=_h_qqq["rel_pct"],
                        mode="lines", name="QQQ",
                        line=dict(color="#4A9EFF", width=2, dash="dot"),
                        hovertemplate="%{x}  %{y:+.2f}%<extra>QQQ</extra>"))
                _fig_bm.add_hline(y=0, line=dict(color=_BORDER, width=1, dash="dot"))
                # 最大回撤标注
                if abs(_h_dd_pct) > 0.5:
                    _fig_bm.add_annotation(
                        x=_h_dd_date, y=float(_h_acct["twr_pct"].iloc[_h_dd_idx]),
                        text=f"最大回撤 {_h_dd_pct:+.1f}%",
                        showarrow=True, arrowhead=2, arrowcolor=_RED,
                        font=dict(color=_RED, size=11), ax=30, ay=30)
                # 当前收益标注
                _fig_bm.add_annotation(
                    x=_h_last_date, y=_h_last_pct,
                    text=f"当前 {_h_last_pct:+.1f}%",
                    showarrow=True, arrowhead=2, arrowcolor=_MUTED,
                    font=dict(color=_TEXT, size=11), ax=-40, ay=-30)
                # 出入金竖线
                for _cf_date, _cf_amt in sorted(_h_cashflows.items()):
                    _cf_label = (f"出金 ${abs(_cf_amt):,.0f}" if _cf_amt < 0
                                 else f"入金 ${_cf_amt:,.0f}")
                    _fig_bm.add_vline(
                        x=_cf_date,
                        line=dict(color=_MUTED, width=1, dash="dash"),
                        annotation_text=_cf_label,
                        annotation_position="top left",
                        annotation_font=dict(color=_MUTED, size=10))
                _fig_bm.update_layout(
                    title=f"投资收益率 vs QQQ（TWR，排除出入金）{_HIST_START} 至今",
                    paper_bgcolor=_BG, plot_bgcolor=_BG, height=360,
                    xaxis=dict(showgrid=False, tickfont=dict(color=_MUTED),
                               title="日期", type="date"),
                    yaxis=dict(showgrid=True, gridcolor=_BORDER,
                               tickfont=dict(color=_MUTED), ticksuffix="%"),
                    legend=dict(font=dict(color=_TEXT), bgcolor=_BG,
                                orientation="h", y=1.12),
                    margin=dict(l=10, r=10, t=44, b=10))
                st.plotly_chart(_fig_bm, use_container_width=True)
                st.caption(
                    f"TWR 排除出入金影响  ·  "
                    f"QQQ 同期: {_qqq_last:+.1f}%  ·  超额Alpha: {_alpha_val:+.1f}%")

        # ── 视图二：AI赋能对比 ──
        else:
            _nav_df, _qqq_df, _base_date, _base_nav = _get_benchmark_chart_data(_perf_acct_id)

            _ai_acct_last = float(_nav_df["rel_pct"].iloc[-1]) if not _nav_df.empty else 0.0
            _ai_qqq_last  = float(_qqq_df["rel_pct"].iloc[-1]) if not _qqq_df.empty  else 0.0
            _ai_alpha     = _ai_acct_last - _ai_qqq_last
            with _bm_alpha_col:
                _kpi(st, "超额Alpha",
                     f"{_ai_alpha:+.1f}%",
                     _GREEN if _ai_alpha >= 0 else _RED)

            if _nav_df.empty:
                st.info(f"暂无 AI赋能数据（起点 {_AI_START}）。"
                        "每次 CDP 同步余额后自动写入。")
            else:
                _fig_bm = go.Figure()
                _fig_bm.add_trace(go.Scatter(
                    x=_nav_df["date"], y=_nav_df["rel_pct"],
                    mode="lines+markers", name="ENERGREX",
                    line=dict(color=_GREEN, width=2.5),
                    marker=dict(size=8, symbol="circle"),
                    hovertemplate="%{x}  %{y:+.2f}%<extra>ENERGREX</extra>"))
                if not _qqq_df.empty:
                    _fig_bm.add_trace(go.Scatter(
                        x=_qqq_df["date"], y=_qqq_df["rel_pct"],
                        mode="lines+markers", name="QQQ",
                        line=dict(color="#4A9EFF", width=2),
                        marker=dict(size=6),
                        hovertemplate="%{x}  %{y:+.2f}%<extra>QQQ</extra>"))
                _fig_bm.add_hline(y=0, line=dict(color=_BORDER, width=1, dash="dot"))
                _fig_bm.add_annotation(
                    x=_base_date, y=0,
                    text=f"AI赋能起点 {_AI_START}  ${_base_nav:,.2f}",
                    showarrow=True, arrowhead=2, arrowcolor=_MUTED,
                    font=dict(color=_MUTED, size=11), ax=50, ay=-35)
                _fig_bm.update_layout(
                    title=f"AI赋能对比 vs QQQ（{_AI_START} 起）",
                    paper_bgcolor=_BG, plot_bgcolor=_BG, height=360,
                    xaxis=dict(showgrid=False, tickfont=dict(color=_MUTED),
                               title="日期", type="date"),
                    yaxis=dict(showgrid=True, gridcolor=_BORDER,
                               tickfont=dict(color=_MUTED), ticksuffix="%"),
                    legend=dict(font=dict(color=_TEXT), bgcolor=_BG,
                                orientation="h", y=1.12),
                    margin=dict(l=10, r=10, t=44, b=10))
                st.plotly_chart(_fig_bm, use_container_width=True)
                st.caption("从今天起每日追踪 ENERGREX vs QQQ 表现对比")

# ── 📅 事件日历 Tab ───────────────────────────────────────────────
with _pos_tabs[2]:
    _cal_acct_label = st.selectbox("账户", [c["label"] for c in ACCT_CFG], key="cal_acct")
    _cal_acct_id    = next((c["id"] for c in ACCT_CFG if c["label"] == _cal_acct_label),
                           ACCT_CFG[0]["id"])
    with st.spinner("加载财报日 / 到期日历…"):
        _events = _get_event_calendar(_cal_acct_id, window_days=30)

    _LEVEL_ICON = {"CRITICAL": "🔴", "HIGH": "🟠", "MEDIUM": "🟡", "LOW": "⚪"}
    _TYPE_LABEL = {"earnings": "📢 财报", "option_expiry": "⏰ 到期"}

    if not _events:
        st.info("未来30天内暂无持仓相关事件（财报数据来自 yfinance，每小时更新）。")
    else:
        _crit = sum(1 for e in _events if e["risk_level"] == "CRITICAL")
        _high = sum(1 for e in _events if e["risk_level"] == "HIGH")
        _mid  = sum(1 for e in _events if e["risk_level"] == "MEDIUM")
        if _crit:
            st.error(f"🔴 **{_crit} 个紧急事件** — 期权到期 DTE ≤ 14，须立即处理")
        if _high:
            st.warning(f"🟠 **{_high} 个高风险事件** — 到期 DTE ≤ 21 或财报 ≤ 7天")
        if _mid:
            st.info(f"🟡 **{_mid} 个中风险事件** — 关注")

        _evt_df = pd.DataFrame([{
            "风险":     _LEVEL_ICON.get(e["risk_level"], "⚪"),
            "日期":     e["date"],
            "剩余天数": e["days_left"],
            "类型":     _TYPE_LABEL.get(e["event_type"], e["event_type"]),
            "标的":     e["symbol"],
            "说明":     e["title"],
            "来源":     e["source"],
        } for e in _events])

        st.dataframe(
            _evt_df, use_container_width=True, hide_index=True,
            height=min(60 + len(_evt_df) * 38, 500),
            column_config={
                "风险":     st.column_config.TextColumn("风险", width="small"),
                "剩余天数": st.column_config.NumberColumn("剩余天数", format="%d 天"),
            }
        )
        st.caption("财报日来自 yfinance（1小时缓存）· 期权到期来自当前持仓表 · "
                   "财报前7天自动标记为 🟠 高风险")

# ── 💡 交易建议 Tab ──────────────────────────────────────────────
with _pos_tabs[3]:
    _rec_acct_label = st.selectbox("账户", [c["label"] for c in ACCT_CFG], key="rec_acct")
    _rec_acct_id    = next((c["id"] for c in ACCT_CFG if c["label"] == _rec_acct_label),
                           ACCT_CFG[0]["id"])

    # ════════════════════════════════════════════════════════════════
    # 出场策略评估
    # ════════════════════════════════════════════════════════════════
    st.markdown("#### 出场策略评估")
    st.caption("止盈 >+50% · 止损 <-50% 且 DTE<30 · 滚仓：Short腿 DTE≤21 · 按紧迫度排序")

    with st.spinner("计算出场分析…"):
        _ea_result = _compute_exit_analysis(_rec_acct_id)

    _ea_ports  = _ea_result.get("portfolios", [])
    _ea_summ   = _ea_result.get("summary", {})

    if not _ea_ports:
        st.info("暂无持仓数据。")
    else:
        # ── 组合层面汇总诊断 ─────────────────────────────────
        _ea_cost_pct  = _ea_summ.get("cost_pct", 0)
        _ea_neq       = _ea_summ.get("net_equity", 0)
        _ea_tot_cost  = _ea_summ.get("total_cost", 0)
        _ea_top_unds  = _ea_summ.get("top_unds", [])
        _ea_n_broken  = _ea_summ.get("n_broken", 0)

        _ea_conc_str = "  ·  ".join(
            f"{u} ({v/_ea_neq*100:.0f}%净值)" if _ea_neq > 0 else u
            for u, v in _ea_top_unds
        )
        _ea_warn_color = _RED if _ea_cost_pct > 100 else (_AMB if _ea_cost_pct > 60 else _GREEN)
        _ea_broken_note = f"  ·  {_ea_n_broken} 个组合方向假设已被推翻" if _ea_n_broken else ""

        st.markdown(
            f"<div style='background:{_ea_warn_color}15;border:1px solid {_ea_warn_color}44;"
            f"border-radius:8px;padding:10px 16px;margin-bottom:12px'>"
            f"<span style='font-size:13px;font-weight:700;color:{_ea_warn_color}'>"
            f"期权成本占净值 {_ea_cost_pct:.0f}%（${_ea_tot_cost:,.0f} / ${_ea_neq:,.0f}）"
            f"{_ea_broken_note}</span><br>"
            f"<span style='font-size:12px;color:{_MUTED}'>"
            f"集中度最高：{_ea_conc_str}</span>"
            f"</div>",
            unsafe_allow_html=True)

        # ── 逐组合卡片（按紧迫度排序）────────────────────────
        for _ep in _ea_ports:
            _border_c = _ep["action_color"]
            _pnl_str  = (f"${_ep['current_pnl']:+,.0f}"
                         + (f" ({_ep['pnl_pct']:+.0f}%)" if _ep["pnl_pct"] is not None else ""))
            _dte_str  = (f"DTE={_ep['min_short_dte']}天" if _ep["has_short"] and _ep["min_short_dte"]
                         else (f"DTE={_ep['min_dte']}天" if _ep["min_dte"] else ""))
            _eq_str   = f"占净值 {_ep['equity_pct']:.0f}%"
            _leg_strs = []
            for _l in _ep["legs"]:
                _ldir = "多" if (_l.get("qty") or 0) > 0 else "空"
                _lq   = abs(_l.get("qty") or 0)
                _leg_strs.append(
                    f"{_ldir}{_lq}×${_l.get('strike', 0):.0f}"
                    f"{'C' if _l.get('direction','').lower()=='call' else 'P'}"
                )
            _legs_label = " + ".join(_leg_strs)
            _und_price_s = (f"  ·  现价 ${_ep['und_price']:.2f}"
                            if _ep.get("und_price") else "")
            _thesis_icon = " ⚠️方向反转" if _ep["thesis_broken"] else ""

            st.markdown(
                f"<div style='border:1px solid {_border_c}55;border-left:3px solid {_border_c};"
                f"border-radius:8px;padding:10px 14px;margin-bottom:8px;background:{_SURF}'>"

                # Header row: action + underlying + type
                f"<div style='display:flex;align-items:center;gap:10px;margin-bottom:6px'>"
                f"<span style='color:{_border_c};font-weight:700;font-size:14px'>"
                f"{_ep['action']}</span>"
                f"<span style='font-size:14px;font-weight:700'>{_ep['underlying']}</span>"
                f"<span style='font-size:12px;color:{_MUTED}'>{_ep['type']}"
                f"  {_legs_label}</span>"
                f"<span style='margin-left:auto;font-size:11px;color:{_MUTED}'>"
                f"{_ep['expiry']}{_thesis_icon}</span>"
                f"</div>"

                # Metric row: PnL | DTE | Equity% | Underlying price
                f"<div style='display:flex;gap:18px;font-size:12px;margin-bottom:6px'>"
                f"<span style='color:{'#00C853' if (_ep['pnl_pct'] or 0)>=0 else _RED};font-weight:600'>"
                f"盈亏 {_pnl_str}</span>"
                f"<span style='color:{_MUTED}'>{_dte_str}</span>"
                f"<span style='color:{'#FF4B4B' if _ep['equity_pct']>25 else _MUTED}'>{_eq_str}</span>"
                f"<span style='color:{_MUTED}'>{_und_price_s}</span>"
                f"</div>"

                # Why text
                f"<div style='font-size:11px;color:{_TEXT};line-height:1.5'>{_ep['why']}</div>"
                f"</div>",
                unsafe_allow_html=True,
            )

    # ════════════════════════════════════════════════════════════════
    # PLTR 对冲监控（IV Rank + 卖Call信号）
    # ════════════════════════════════════════════════════════════════
    st.divider()
    st.markdown("#### PLTR 对冲监控")
    with st.spinner("拉取 PLTR IV Rank（MarketData.app）…"):
        _pltr_snap = _fetch_ivrank_md("PLTR")

    if _pltr_snap["error"]:
        st.warning(f"MarketData.app 数据获取失败：{_pltr_snap['error']}")
    else:
        _piv  = _pltr_snap.get("iv")           # float % 或 None
        _pivr = _pltr_snap.get("ivr")          # float 0-100 或 None（需30天历史）
        _pivp = _pltr_snap.get("iv_percentile")  # 旧API字段，链接口无此数据
        _plo  = _pltr_snap.get("iv_low")
        _phi  = _pltr_snap.get("iv_high")
        _psig, _prec, _pcol = _pltr_ivr_signal(_piv or 0, _pivr)

        # ── 四个指标格子 ──────────────────────────────────────────
        _pa, _pb, _pc, _pd = st.columns(4)
        _kpi(_pa, "当前 ATM IV",
             f"{_piv:.1f}%" if _piv else "—",
             _GREEN if _piv and _piv > 60 else (_AMB if _piv and _piv > 40 else _MUTED))
        _kpi(_pb, "IV Rank（52周）",
             f"{_pivr:.0f}%" if _pivr is not None else "—",
             _GREEN if _pivr and _pivr > 60 else (_AMB if _pivr and _pivr > 40 else _MUTED))
        _kpi(_pc, "IV Percentile",
             f"{_pivp:.0f}%" if _pivp is not None else "—",
             _GREEN if _pivp and _pivp > 60 else (_AMB if _pivp and _pivp > 40 else _MUTED))
        _kpi(_pd, "综合评级", _psig, _pcol)

        # ── 建议卡片 ─────────────────────────────────────────────
        _range_note = (f"52周 IV 区间：{_plo:.0f}% – {_phi:.0f}%"
                       if _plo and _phi else "52周区间：MarketData.app 提供")
        st.markdown(
            f"<div style='background:{_SURF};border:1px solid {_BORDER};"
            f"border-radius:8px;padding:12px 16px;margin-top:8px'>"
            f"<span style='font-size:14px;font-weight:700;color:{_pcol}'>{_psig} {_prec}</span>"
            f"<br><span style='font-size:11px;color:{_MUTED}'>{_range_note}</span>"
            f"</div>",
            unsafe_allow_html=True)

        # ── 触发条件说明 ─────────────────────────────────────────
        with st.expander("查看触发条件"):
            st.markdown(
                "| IVR | IV绝对值 | 信号 |\n"
                "|-----|---------|------|\n"
                "| < 25% | 任意 | ⚪ 等待，历史偏低 |\n"
                "| 25–40% | 任意 | ⚪ 等待 |\n"
                "| 40–60% | > 50% | 🟡 可卖1张 |\n"
                "| 60–80% | > 60% | 🟠 建议卖2张 |\n"
                "| > 80% | > 70% | 🔴 最佳时机，卖2–3张 |"
            )

    # ════════════════════════════════════════════════════════════════
    # 机构 Greeks 仪表板 (Module D)
    # ════════════════════════════════════════════════════════════════
    st.divider()

    # ── 标题行：仪表板名 + VIX 内联 ──────────────────────────────
    _gh_left, _gh_right = st.columns([3, 1])
    with _gh_left:
        st.markdown("#### 机构 Greeks 仪表板")
        st.caption("三重触发：Delta漂移±0.10 · 09:35日检 · 财报前7天/VIX涨>15%升级为高优先级")
    with _gh_right:
        _vix_pre = _get_vix_snapshot()
        if _vix_pre["vix"]:
            _vc = (_RED if (_vix_pre["change_pct"] or 0) > 15
                   else (_AMB if (_vix_pre["change_pct"] or 0) > 5 else _MUTED))
            _vix_sign = "+" if (_vix_pre["change_pct"] or 0) >= 0 else ""
            st.markdown(
                f"<div style='text-align:right;padding-top:10px'>"
                f"<span style='font-size:11px;color:{_MUTED}'>VIX&nbsp;</span>"
                f"<span style='font-size:14px;font-weight:700;color:{_vc}'>"
                f"{_vix_pre['vix']:.1f}</span>"
                f"<span style='font-size:11px;color:{_vc}'>"
                f"&nbsp;({_vix_sign}{_vix_pre['change_pct']:.1f}%)</span>"
                f"</div>",
                unsafe_allow_html=True)

    # ════════════════════════════════════════════════════════════════
    # QQQ 对冲建议（数量计算）
    # ════════════════════════════════════════════════════════════════
    st.divider()
    st.markdown("#### 🛡️ QQQ 对冲建议")

    # 目标 Beta-Delta 滑块（默认 150%）
    _qqq_target_pct = st.slider(
        "目标 Beta-Delta（% 净值）",
        min_value=100, max_value=200, value=150, step=5,
        key="qqq_hedge_target",
        help="把组合 Beta-Delta 从当前值降到这个目标。来自 Portfolio_Config 的建议值为 150%。")
    _qqq_target_ratio = _qqq_target_pct / 100.0

    with st.spinner("计算 QQQ 对冲方案…"):
        _hplan = _compute_qqq_hedge_plan(_rec_acct_id, _qqq_target_ratio)

    if "error" in _hplan:
        st.warning(f"对冲计算失败：{_hplan['error']}")
    else:
        # ── 状态行 ──────────────────────────────────────────────
        _h_cur   = _hplan["current_bd_ratio"]
        _h_tgt   = _hplan["target_bd_ratio"]
        _h_qqq   = _hplan["qqq_price"]
        _h_iv    = _hplan["qqq_iv"]
        _h_bqq   = _hplan["b_qqq"]
        _h_btoh  = _hplan["bd_to_hedge"]
        _h_ebd   = _hplan["existing_bd"]
        _h_n_ex  = _hplan["n_existing"]
        _h_leggs = _hplan["existing_legs"]

        _hsc1, _hsc2, _hsc3, _hsc4 = st.columns(4)
        _hsc1.metric("当前 Beta-Delta", f"{_h_cur:.1f}%",
                     delta="需要对冲" if _h_cur > _h_tgt else "无需对冲")
        _hsc2.metric("目标 Beta-Delta", f"{_h_tgt:.0f}%")
        _hsc3.metric("QQQ 现价", f"${_h_qqq:,.2f}",
                     delta=f"IV {_h_iv:.1f}%  β={_h_bqq}")
        _hsc4.metric("需对冲量（$）", f"${_h_btoh:+,.0f}",
                     delta=f"现有对冲 ${_h_ebd:+,.0f}")

        # ── 现有 QQQ 持仓明细 ───────────────────────────────────
        if _h_leggs:
            _h_legs_str = "  ·  ".join(
                f"{l['sym']} ×{l['qty']:+.0f}  Δ{l['delta']:+.3f}  @${l['price']:.2f}"
                for l in _h_leggs)
            st.caption(f"现有 QQQ 期权持仓：{_h_legs_str}")
            if _h_btoh <= 0:
                st.success(f"✅ 现有对冲已足够（BD {_h_cur:.1f}% ≤ 目标 {_h_tgt:.0f}%），无需加仓")

        # ── 需要对冲时才显示方案 ────────────────────────────────
        if _h_btoh > 0:
            st.caption(
                f"需要减少 Beta-Delta ${_h_btoh:,.0f}（{_h_cur:.1f}% → {_h_tgt:.0f}%）"
                f"  ·  每张 Put Spread（β={_h_bqq}）对冲量见各方案")
            st.markdown("---")

            # Plan helper: render one plan card
            def _hedge_plan_card(col, title, plan: dict, exp_str: str, label: str, accent: str):
                buy_k  = plan["buy_strike"]
                sell_k = plan["sell_strike"]
                n      = plan["n_total"]
                cost   = plan["total_cost"]
                cps    = plan["cost_per_spread"]
                post   = plan["post_bd_ratio"]
                th_chg = plan["theta_change"]
                new_ocr = plan["new_ocr"]
                bd_ps  = plan["bd_per_spread"]
                occ_buy  = f"QQQ{exp_str}{int(buy_k*1000):08d}P"
                occ_sell = f"QQQ{exp_str}{int(sell_k*1000):08d}P"
                _tgt_bd  = _hplan['target_bd_ratio']
                _tgt_col = _GREEN if post <= _tgt_bd else _AMB
                _tgt_lbl = "✓达标" if post <= _tgt_bd else "⚠未达标"
                with col:
                    st.markdown(
                        f"<div style='background:{_SURF};border:2px solid {accent};"
                        f"border-radius:10px;padding:14px 16px;height:100%'>"
                        f"<div style='font-size:13px;font-weight:700;color:{accent}'>"
                        f"{title} {label}</div>"
                        f"<div style='font-size:11px;color:{_MUTED};margin:4px 0 8px'>spread 宽度 "
                        f"${buy_k - sell_k:.0f}  ·  DTE≈{_hplan['plan_dte']}天</div>"
                        f"<div style='font-size:12px;line-height:2;color:{_TEXT}'>"
                        f"🟢 买入 {occ_buy} × <b>{n}</b> 张<br>"
                        f"🔴 卖出 {occ_sell} × <b>{n}</b> 张<br>"
                        f"<span style='color:{_MUTED}'>每张净成本：<b>${cps:.2f}</b>  "
                        f"长腿Δ {plan['d_long']:+.3f} / 短腿Δ {plan['d_short']:+.3f}<br>"
                        f"每张对冲 BD ${-bd_ps:,.0f}</span>"
                        f"</div>"
                        f"<hr style='border-color:{_BORDER};margin:8px 0'>"
                        f"<div style='font-size:12px;line-height:1.9;color:{_TEXT}'>"
                        f"💰 总成本：<b>${cost:,.0f}</b><br>"
                        f"📊 执行后 Beta-Delta：<b>{post:.1f}%</b>  "
                        f"<span style='font-size:11px;color:{_tgt_col}'>"
                        f"{_tgt_lbl}</span><br>"
                        f"⏱ Theta 变化：{th_chg:+.2f} $/天<br>"
                        f"📋 执行后成本率：{new_ocr:.1f}%"
                        f"</div>"
                        f"</div>",
                        unsafe_allow_html=True)

            _ha, _hb, _hc = st.columns(3)
            _pa = _hplan["plan_a"]
            _pb = _hplan["plan_b"]
            _pc = _hplan["plan_c"]
            _occ_exp = _hplan["plan_occ_exp"]

            _hedge_plan_card(_ha, "方案A", _pa, _occ_exp, "（标准$35宽）", _BLUE)
            _hedge_plan_card(_hb, "方案B", _pb, _occ_exp, "（宽幅$60宽）", _AMB)

            # Plan C card (keep existing + add)
            with _hc:
                _n_add = _pc["n_additional"]
                _n_ex  = _pc["n_existing"]
                _buy_c = _pc["buy_strike"]
                _sell_c = _pc["sell_strike"]
                _cps_c  = _pc["cost_per_spread"]
                _acost  = _pc["additional_cost"]
                _post_c = _pc["post_bd_ratio"]
                _occ_b  = f"QQQ{_occ_exp}{int(_buy_c*1000):08d}P"
                _occ_s  = f"QQQ{_occ_exp}{int(_sell_c*1000):08d}P"
                _pc_tgt = _hplan['target_bd_ratio']
                _pc_col = _GREEN if _post_c <= _pc_tgt else _AMB
                _pc_lbl = "✓达标" if _post_c <= _pc_tgt else "⚠未达标"
                st.markdown(
                    f"<div style='background:{_SURF};border:2px solid {_GREEN};"
                    f"border-radius:10px;padding:14px 16px;height:100%'>"
                    f"<div style='font-size:13px;font-weight:700;color:{_GREEN}'>"
                    f"方案C （保留现有+补充）</div>"
                    f"<div style='font-size:11px;color:{_MUTED};margin:4px 0 8px'>"
                    f"现有 {_n_ex} 张  ·  已对冲 BD ${_pc['existing_bd']:+,.0f}</div>"
                    f"<div style='font-size:12px;line-height:2;color:{_TEXT}'>"
                    f"🟢 追加买入 {_occ_b} × <b>{_n_add}</b> 张<br>"
                    f"🔴 追加卖出 {_occ_s} × <b>{_n_add}</b> 张<br>"
                    f"<span style='color:{_MUTED}'>每张净成本：<b>${_cps_c:.2f}</b></span>"
                    f"</div>"
                    f"<hr style='border-color:{_BORDER};margin:8px 0'>"
                    f"<div style='font-size:12px;line-height:1.9;color:{_TEXT}'>"
                    f"💰 追加成本：<b>${_acost:,.0f}</b><br>"
                    f"📊 执行后 Beta-Delta：<b>{_post_c:.1f}%</b>  "
                    f"<span style='font-size:11px;color:{_pc_col}'>{_pc_lbl}</span><br>"
                    f"ℹ️ 在方案A基础上复用现有持仓"
                    f"</div>"
                    f"</div>",
                    unsafe_allow_html=True)

            # ── 免责声明 ─────────────────────────────────────────
            st.caption(
                f"⚠️ 以上为理论计算（BS定价，IV={_h_iv:.1f}%，DTE={_hplan['plan_dte']}天，β={_h_bqq}）。"
                "执行前须核实实时 Bid/Ask、流动性、到期日是否有标准月份。不构成交易指令。")

    # ── 控制行：计算按钮 + 目标 Delta 滑块 ───────────────────────
    _gc1, _gc2 = st.columns([1, 3])
    with _gc1:
        _run_greeks = st.button("🔬 计算 Greeks", key="run_greeks_btn", type="primary")
    with _gc2:
        _delta_lo, _delta_hi = st.slider(
            "目标 Delta 范围（均值/合约）",
            min_value=-1.0, max_value=2.0,
            value=(0.30, 0.60), step=0.05,
            key="delta_target_range")

    if _run_greeks or st.session_state.get("_greeks_cache"):
        with st.spinner("从 MarketData.app + yfinance 拉取 ATM IV，BS 计算全组合 Greeks…"):
            _gdata = _compute_portfolio_greeks(_rec_acct_id)
            st.session_state["_greeks_cache"] = _gdata   # 缓存结果避免重复计算
    elif isinstance(st.session_state.get("_greeks_cache"), dict):
        _gdata = st.session_state["_greeks_cache"]
    else:
        _gdata = None

    if _gdata:
        # ── 触发器警报 ─────────────────────────────────────────────
        for _trig in _gdata["triggers"]:
            if _trig["level"] == "CRITICAL":
                st.error(_trig["msg"])
            else:
                st.warning(_trig["msg"])
        if not _gdata["triggers"]:
            st.success("✅ 三重触发均未激活 — 组合风险在监控范围内")

        # ── 4 KPI 卡片 ────────────────────────────────────────────
        _tot = _gdata["totals"]
        _n   = _gdata["n_contracts"]
        _avg_d = _tot["avg_delta"]

        _delta_in_range = _delta_lo <= _avg_d <= _delta_hi
        _delta_card_color = _GREEN if _delta_in_range else _RED

        _gk1, _gk2, _gk3, _gk4 = st.columns(4)
        _kpi(_gk1, f"总 Delta（{_n}张）",
             f"{_tot['delta']:+.3f}",
             _delta_card_color)
        _kpi(_gk2, "均值 Delta/合约",
             f"{_avg_d:+.3f}",
             _delta_card_color)
        _kpi(_gk3, "日 Theta 损耗",
             f"${_tot['theta']:+.0f}/天",
             _GREEN if _tot["theta"] > 0 else _RED)
        _kpi(_gk4, "Vega 敞口",
             f"${_tot['vega']:+.0f}/1%IV",
             _AMB if abs(_tot["vega"]) > 200 else _MUTED)

        # ── 智能对冲建议 ──────────────────────────────────────────
        _top_long  = _gdata.get("top_long")
        _top_short = _gdata.get("top_short")
        _by_und    = _gdata.get("by_und", {})

        if _avg_d < _delta_lo:
            _gap = _delta_lo - _avg_d
            _n_adj = max(1, round(_gap / 0.35))
            _suggest_und = _top_short or (_top_long or "标的")
            st.info(
                f"📉 **均值 Delta {_avg_d:+.3f} 低于目标下限 {_delta_lo:+.2f}**\n\n"
                f"需增加约 **+{_gap:.2f} Delta/合约**（约 {_n_adj} 张合约）。\n"
                f"建议：买入 **{_suggest_und}** 看涨期权，选 Delta ≈ +0.35–0.45，"
                f"DTE ≥ 45 天（避免 Gamma 爆炸）。"
                f"\n\n目标 Delta 贡献最小标的：{_suggest_und}（{_by_und.get(_suggest_und, 0):+.3f}）")
        elif _avg_d > _delta_hi:
            _gap = _avg_d - _delta_hi
            _n_adj = max(1, round(_gap / 0.35))
            _suggest_und = _top_long or "标的"
            st.warning(
                f"📈 **均值 Delta {_avg_d:+.3f} 超出目标上限 {_delta_hi:+.2f}**\n\n"
                f"需减少约 **{_gap:.2f} Delta/合约**（约 {_n_adj} 张合约）。\n"
                f"建议：卖出 **{_suggest_und}** 看涨期权（Delta ≈ 0.30–0.40，DTE ≥ 45）"
                f"或买入保护性看跌期权。"
                f"\n\n目标 Delta 最大贡献标的：{_suggest_und}（{_by_und.get(_suggest_und, 0):+.3f}）")
        else:
            st.success(
                f"✅ 均值 Delta **{_avg_d:+.3f}** 在目标范围 "
                f"[{_delta_lo:+.2f}, {_delta_hi:+.2f}]，无需调仓。")

        # ── 高 Gamma 风险警告横幅 ─────────────────────────────────
        if _gdata["rows"]:
            _hg = [r["symbol"] for r in _gdata["rows"] if r["high_gamma"]]
            if _hg:
                st.error(f"⚡ **高 Gamma 风险（DTE < 21）：{', '.join(_hg)}** — 须优先处理")

        # ── 每持仓 Greeks 明细表 ──────────────────────────────────
        if _gdata["rows"]:
            _gdf = pd.DataFrame(_gdata["rows"])
            _disp = [
                "status", "symbol", "opt_type", "qty",
                "strike", "expiry", "dte", "spot", "iv_pct",
                "bs_delta", "pos_delta", "pos_gamma", "pos_theta", "pos_vega"]
            _lmap = {
                "status":    "状态",
                "symbol":    "合约代码",
                "opt_type":  "方向",
                "qty":       "手数",
                "strike":    "行权价",
                "expiry":    "到期日",
                "dte":       "DTE",
                "spot":      "现价",
                "iv_pct":    "IV%",
                "bs_delta":  "单合约Δ",
                "pos_delta": "持仓Δ",
                "pos_gamma": "持仓Γ",
                "pos_theta": "日Θ($)",
                "pos_vega":  "Vega$/1%",
            }
            _gshow = _gdf[[c for c in _disp if c in _gdf.columns]].copy()
            _gshow.rename(columns=_lmap, inplace=True)

            st.dataframe(
                _gshow,
                use_container_width=True,
                hide_index=True,
                height=min(60 + len(_gshow) * 42, 600),
                column_config={
                    "状态":     st.column_config.TextColumn("状态", width="small"),
                    "手数":     st.column_config.NumberColumn("手数", width="small"),
                    "DTE":      st.column_config.NumberColumn("DTE",  width="small"),
                    "日Θ($)":  st.column_config.NumberColumn("日Θ($)",    format="%.2f"),
                    "Vega$/1%": st.column_config.NumberColumn("Vega$/1%", format="%.2f"),
                    "持仓Δ":   st.column_config.NumberColumn("持仓Δ",    format="%+.4f"),
                    "持仓Γ":   st.column_config.NumberColumn("持仓Γ",    format="%.6f"),
                    "单合约Δ": st.column_config.NumberColumn("单合约Δ",  format="%+.4f"),
                })

            # ── 组合合计行 ──────────────────────────────────────────
            _dc = _delta_card_color
            st.markdown(
                f"<div style='background:{_SURF};border:1px solid {_BORDER};"
                f"border-radius:6px;padding:10px 16px;margin-top:4px;"
                f"font-size:12px;color:{_TEXT};line-height:1.8'>"
                f"📊 <b>组合合计</b>&nbsp;&nbsp;|&nbsp;&nbsp;"
                f"总 Δ <b style='color:{_dc}'>{_tot['delta']:+.4f}</b>"
                f"&nbsp;&nbsp;|&nbsp;&nbsp;"
                f"总 Γ <b>{_tot['gamma']:+.6f}</b>"
                f"&nbsp;&nbsp;|&nbsp;&nbsp;"
                f"日 Θ <b style='color:{_GREEN if _tot['theta']>0 else _RED}'>"
                f"${_tot['theta']:+.2f}/天</b>"
                f"&nbsp;&nbsp;|&nbsp;&nbsp;"
                f"Vega <b>${_tot['vega']:+.2f}/1%IV</b>"
                f"</div>",
                unsafe_allow_html=True)

            # ── IV 来源说明 ─────────────────────────────────────────
            _src = _gdata.get("iv_src_counts", {})
            _src_parts = []
            if _src.get("md"):    _src_parts.append(f"MarketData.app {_src['md']}只")
            if _src.get("yf"):    _src_parts.append(f"yfinance {_src['yf']}只")
            if _src.get("db"):    _src_parts.append(f"数据库缓存 {_src['db']}只")
            if _src.get("default"): _src_parts.append(f"默认 50% {_src['default']}只")
            if _src_parts:
                st.caption(f"IV 来源：{'  ·  '.join(_src_parts)}")
        else:
            st.info("暂无可计算 Greeks 的持仓（需要 IV、到期日、行权价数据）。")

    st.divider()
    st.markdown("#### ⚡ 操作建议 + 策略模拟器")
    st.caption(
        "⚠️ 模拟数据 — 不构成真实交易。"
        "模拟用价格：当前市价中间价估算（实际成交可能有滑点）。")

    with st.spinner("生成交易建议…"):
        _recs = _generate_recommendations(_rec_acct_id)

    if not _recs:
        st.success("✅ 当前无紧急持仓事项，风险在控制范围内。")
    else:
        _cnt_crit = sum(1 for r in _recs if r["优先级"].startswith("🔴"))
        _cnt_high = sum(1 for r in _recs if r["优先级"].startswith("🟠"))
        _cnt_opp  = sum(1 for r in _recs if r["优先级"].startswith("🟢"))
        if _cnt_crit:
            st.error(f"🔴 **{_cnt_crit} 条紧急建议** — 须尽快处理")
        if _cnt_high:
            st.warning(f"🟠 **{_cnt_high} 条高优先建议**")
        if _cnt_opp:
            st.info(f"🟢 **{_cnt_opp} 条新机会候选**（有风险余量，高分标的）")

        # ── 可选择的建议表（含勾选框）────────────────────────────
        _rec_sel_df = pd.DataFrame([{
            "模拟": False,
            "#":   r["序号"],
            "优先级": r["优先级"],
            "标的": r["标的"],
            "组合": r["组合"],
            "DTE":  r["DTE"],
            "当前盈亏": r["当前盈亏"],
            "行动建议": r["行动建议"],
        } for r in _recs])

        _edited_recs = st.data_editor(
            _rec_sel_df,
            column_config={
                "模拟": st.column_config.CheckboxColumn("模拟", default=False, width="small"),
                "#":   st.column_config.NumberColumn("#", width="small"),
                "优先级": st.column_config.TextColumn("优先级"),
                "标的": st.column_config.TextColumn("标的", width="small"),
                "DTE":  st.column_config.TextColumn("DTE", width="small"),
            },
            use_container_width=True,
            hide_index=True,
            disabled=["#", "优先级", "标的", "组合", "DTE", "当前盈亏", "行动建议"],
            height=min(60 + len(_rec_sel_df) * 42, 500),
            key="rec_editor",
        )

        # ── 模拟执行按钮 ─────────────────────────────────────────
        _sim_col1, _sim_col2 = st.columns([1, 4])
        with _sim_col1:
            _do_sim = st.button("⚡ 模拟执行", key="do_sim_btn", type="primary")
        with _sim_col2:
            _selected_rows = _edited_recs[_edited_recs["模拟"] == True]
            if not _selected_rows.empty:
                st.caption(f"已选 {len(_selected_rows)} 条：" +
                           "、".join(_selected_rows["标的"].tolist()))
            else:
                st.caption("勾选左侧 ✓ 选择要模拟的建议，再点击「⚡ 模拟执行」")

        if _do_sim:
            if _selected_rows.empty:
                st.warning("请先勾选至少一条建议再模拟。")
            else:
                _sel_idx  = list(_selected_rows.index)
                _sel_recs = [_recs[i] for i in _sel_idx]
                _sim_acts = [r.get("_sim_action", {"type": "no_change"}) for r in _sel_recs]

                _base_snap = _compute_risk_snapshot(_rec_acct_id)
                _hplan_sim = _compute_qqq_hedge_plan(_rec_acct_id)
                _after     = _compute_sim_impact(_rec_acct_id, _sim_acts, _base_snap, _hplan_sim)
                _sc_before = _run_scenarios(_base_snap)
                # Build after-snap dict for scenarios
                _after_snap_mini = {
                    "equity":      _base_snap.get("equity", 1),
                    "beta_delta":  _after["beta_delta"],
                    "theta_per_day": _after["theta_per_day"],
                    "stress_10":   _after["stress_10"],
                    "stress_20":   _after["stress_20"],
                    "vega_per_pt": _base_snap.get("vega_per_pt", 0),
                }
                _sc_after = _run_scenarios(_after_snap_mini)

                _sim_result = {
                    "actions":    _after["actions"],
                    "base_snap":  _base_snap,
                    "after":      _after,
                    "sc_before":  _sc_before,
                    "sc_after":   _sc_after,
                }
                st.session_state[f"_sim_{_rec_acct_id}"] = _sim_result
                _rid = _save_sim_record(
                    _rec_acct_id,
                    _after["actions"],
                    {k: _base_snap.get(k) for k in
                     ("equity","beta_delta","beta_delta_ratio","theta_per_day",
                      "stress_10_ratio","stress_20_ratio")},
                    {k: _after.get(k) for k in
                     ("beta_delta","beta_delta_ratio","theta_per_day",
                      "stress_10","stress_10_ratio","stress_20","stress_20_ratio",
                      "cash_delta")},
                    _sc_after,
                )
                st.success(f"✅ 模拟完成（已保存记录 #{_rid}）")

        # ── 模拟结果显示 ─────────────────────────────────────────
        _sim_res = st.session_state.get(f"_sim_{_rec_acct_id}")
        if _sim_res:
            _aft  = _sim_res["after"]
            _base = _sim_res["base_snap"]
            _scb  = _sim_res["sc_before"]
            _sca  = _sim_res["sc_after"]
            equity_now = _base.get("equity", 1) or 1

            st.markdown(
                "<div style='border:2px dashed #4FC3F7;border-radius:10px;"
                "padding:14px 18px;margin:10px 0;background:#071423'>"
                "<div style='font-size:13px;font-weight:700;color:#4FC3F7;margin-bottom:8px'>"
                "📊 模拟结果对比（虚线框 = 模拟数据）</div>",
                unsafe_allow_html=True)

            # 操作说明
            for _ad in _aft.get("actions", []):
                st.markdown(f"  → {_ad}")

            # Before / After 对比
            _ba1, _ba2 = st.columns(2)
            def _cmp_row(label, bval, aval, fmt="{:.1f}%", better="lower"):
                _bstr = fmt.format(bval) if bval is not None else "—"
                _astr = fmt.format(aval) if aval is not None else "—"
                try:
                    _delta = float(aval) - float(bval)
                    _good  = (_delta < 0) if better == "lower" else (_delta > 0)
                    _col   = "#00D4AA" if _good else "#FF4B6E"
                    _arrow = f"↓{abs(_delta):.1f}" if _delta < 0 else f"↑{abs(_delta):.1f}"
                except Exception:
                    _col = "#8B9BB4"; _arrow = "—"
                return f"**{label}**", _bstr, _astr, f"<span style='color:{_col}'>{_arrow}</span>"

            with _ba1:
                st.markdown("**执行前**")
                st.metric("Beta-Delta",  f"{(_base.get('beta_delta_ratio') or 0)*100:.1f}%")
                st.metric("每日 Theta",  f"${_base.get('theta_per_day', 0):+,.0f}")
                st.metric("压力 -10%",   f"${_base.get('stress_10', 0):+,.0f}")
                st.metric("压力 -20%",   f"${_base.get('stress_20', 0):+,.0f}")
            with _ba2:
                st.markdown("**执行后（模拟）**")
                _bdr_a  = _aft.get("beta_delta_ratio", 0)
                _s10_a  = _aft.get("stress_10_ratio", 0)
                st.metric("Beta-Delta",  f"{_bdr_a:.1f}%",
                          delta=f"{_bdr_a - (_base.get('beta_delta_ratio') or 0)*100:.1f}%")
                st.metric("每日 Theta",  f"${_aft.get('theta_per_day',0):+,.0f}",
                          delta=f"${_aft.get('theta_delta',0):+,.2f}/天")
                st.metric("压力 -10%",   f"${_aft.get('stress_10',0):+,.0f}",
                          delta=f"${_aft.get('s10_delta',0):+,.0f}")
                st.metric("压力 -20%",   f"${_aft.get('stress_20',0):+,.0f}",
                          delta=f"${_aft.get('s20_delta',0):+,.0f}")

            # 资金变化
            cash_delta_value = _aft.get("cash_delta", 0)
            st.caption(
                f"现金变化：{'+收入 $' if cash_delta_value >= 0 else '-支出 $'}{abs(cash_delta_value):,.0f}"
                f"（{'关仓收回' if cash_delta_value >= 0 else '新建头寸成本'}）")

            st.markdown("</div>", unsafe_allow_html=True)

            # ── 情景压力测试 ─────────────────────────────────────
            st.markdown("**情景压力测试**")
            _sc1, _sc2, _sc3 = st.columns(3)
            _scn = [
                (_sc1, "情景一：平静±3%（30天）", _scb["calm"], _sca["calm"]),
                (_sc2, "情景二：QQQ跌10%",         _scb["drop10"], _sca["drop10"]),
                (_sc3, "情景三：QQQ涨10%",          _scb["rally10"], _sca["rally10"]),
            ]
            for _col, _sname, _sb_pnl, _sa_pnl in _scn:
                with _col:
                    _diff = _sa_pnl - _sb_pnl
                    _col_diff = _GREEN if _diff >= 0 else _RED
                    st.markdown(
                        f"<div style='background:{_SURF};border:1px solid {_BORDER};"
                        f"border-radius:8px;padding:10px 12px'>"
                        f"<div style='font-size:11px;color:{_MUTED}'>{_sname}</div>"
                        f"<div style='font-size:13px;margin-top:4px'>"
                        f"执行前：<b>${_sb_pnl:+,.0f}</b><br>"
                        f"执行后：<b>${_sa_pnl:+,.0f}</b><br>"
                        f"<span style='color:{_col_diff}'>差异：${_diff:+,.0f}</span>"
                        f"</div></div>",
                        unsafe_allow_html=True)

            st.caption("情景三为近似估算（Δ线性+Gamma/Vega调整），仅供参考。")

            if st.button("🗑️ 清除模拟结果", key="clear_sim"):
                del st.session_state[f"_sim_{_rec_acct_id}"]
                st.rerun()

        # ── 建议详情（可展开）────────────────────────────────────
        st.markdown("---")
        st.markdown("#### 组合建议详情（点击展开）")
        _urgent_recs = [r for r in _recs if r["优先级"].startswith(("🔴", "🟠"))]
        _other_recs  = [r for r in _recs if not r["优先级"].startswith(("🔴", "🟠"))]
        for _r in (_urgent_recs + _other_recs):
            _expand = _r["优先级"].startswith("🔴")
            _header = (f"{_r['优先级']} **{_r['标的']}** — {_r.get('组合','—')} "
                       f"（DTE: {_r['DTE']}，手数: {_r.get('手数',0)}，{_r['当前盈亏']}）")
            with st.expander(_header, expanded=_expand):
                st.markdown(f"**行动建议**: {_r['行动建议']}")
                st.markdown(f"**触发原因**: {_r['触发原因']}")
                _ml = _r.get("最大亏损", "—")
                _mp = _r.get("最大盈利", "—")
                if _ml != "—" or _mp != "—":
                    st.markdown(f"**风险/回报**: 最大亏损 {_ml}  ·  最大盈利 {_mp}")
                if _r.get("AI评分") and _r["AI评分"] != "—":
                    st.markdown(f"**AI 评分**: {_r['AI评分']}")
                _sa = _r.get("_sim_action", {})
                st.caption(f"模拟类型：{_sa.get('type','—')}  —  {_sa.get('label','')}")

        # ── 模拟历史 ─────────────────────────────────────────────
        st.markdown("---")
        st.markdown("#### 📅 模拟历史（最近8次）")
        _sim_hist = _load_sim_records(_rec_acct_id)
        if not _sim_hist:
            st.caption("暂无模拟记录。勾选建议并点击「⚡ 模拟执行」会自动保存。")
        else:
            _hist_rows = []
            for _hr in _sim_hist:
                _hb = _hr.get("before", {})
                _ha = _hr.get("after",  {})
                _hs = _hr.get("scenarios", {})
                _hist_rows.append({
                    "时间":      _hr["sim_ts"],
                    "操作":      " | ".join(_hr.get("actions", ["—"])[:2]),
                    "BD前":      f"{(_hb.get('beta_delta_ratio') or 0)*100:.1f}%",
                    "BD后":      f"{(_ha.get('beta_delta_ratio') or 0):.1f}%",
                    "Theta变化": f"${(_ha.get('theta_per_day',0) - _hb.get('theta_per_day',0)):+,.2f}",
                    "模拟跌10%": f"${_hs.get('drop10', 0):+,.0f}",
                    "现金±":     f"${_ha.get('cash_delta', 0):+,.0f}",
                })
            st.dataframe(pd.DataFrame(_hist_rows), use_container_width=True,
                         hide_index=True)

# ── 📋 交易历史 Tab ───────────────────────────────────────────────
with _pos_tabs[4]:
    _hist_acct_label = st.selectbox("账户", [c["label"] for c in ACCT_CFG], key="hist_acct")
    _hist_acct_id    = next((c["id"] for c in ACCT_CFG if c["label"] == _hist_acct_label),
                            ACCT_CFG[0]["id"])

    _df_hist = _load_realized_trades(_hist_acct_id)

    if _df_hist.empty:
        st.info("暂无已实现交易记录。请到「📊 交易绩效」Tab 点击「🔄 运行 FIFO 分析」生成数据。")
    else:
        # ── 汇总指标 ─────────────────────────────────────────────
        _ht = len(_df_hist)
        _hw = (_df_hist["win_loss"] == "win").sum()
        _hp = pd.to_numeric(_df_hist["realized_pnl"], errors="coerce").fillna(0)
        _htotal = _hp.sum()
        _havg_w = _hp[_hp > 0].mean() if (_hp > 0).any() else 0.0
        _havg_l = _hp[_hp < 0].mean() if (_hp < 0).any() else 0.0
        _hpf    = abs(_havg_w / _havg_l) if _havg_l else None

        _hc1, _hc2, _hc3, _hc4, _hc5 = st.columns(5)
        _hc1.metric("总笔数",  str(_ht))
        _hc2.metric("胜率",    f"{_hw/_ht*100:.1f}%  ({_hw}/{_ht})")
        _hc3.metric("累计盈亏", f"${_htotal:+,.2f}",
                    delta=f"{_htotal:+,.2f}",
                    delta_color="normal")
        _hc4.metric("平均盈利", f"${_havg_w:+,.2f}")
        _hc5.metric("盈亏比",   f"{_hpf:.2f}x" if _hpf else "—")

        st.markdown("")

        # ── 明细表 ───────────────────────────────────────────────
        _hist_disp = _df_hist[[
            "close_date", "underlying", "strategy_type",
            "realized_pnl", "win_loss", "open_date", "holding_days",
        ]].copy()
        _hist_disp.columns = ["平仓日", "标的", "策略类型", "盈亏$", "胜负", "开仓日", "持仓天数"]
        _hist_disp["盈亏$"]   = pd.to_numeric(_hist_disp["盈亏$"], errors="coerce")
        _hist_disp["胜负"]    = _hist_disp["胜负"].map({"win": "✅ 盈", "loss": "❌ 亏"})
        _hist_disp = _hist_disp.sort_values("平仓日", ascending=False)

        st.dataframe(
            _hist_disp,
            use_container_width=True,
            hide_index=True,
            height=min(45 + _ht * 35, 600),
            column_config={
                "盈亏$": st.column_config.NumberColumn(
                    "盈亏 $", format="$%.2f"),
                "持仓天数": st.column_config.NumberColumn(
                    "持仓天", width="small"),
                "胜负": st.column_config.TextColumn("胜负", width="small"),
            }
        )

        st.caption(
            f"共 {_ht} 笔  ·  胜率 {_hw/_ht*100:.1f}%  ·  "
            f"平均盈利 ${_havg_w:+,.2f}  ·  平均亏损 ${_havg_l:+,.2f}  ·  "
            f"盈亏比 {f'{_hpf:.2f}x' if _hpf else '—'}"
        )

# ── 🔍 数据核查 Tab ──────────────────────────────────────────────
with _pos_tabs[5]:
    _dq_acct_label = st.selectbox("账户", [c["label"] for c in ACCT_CFG], key="dq_acct")
    _dq_acct_id    = next((c["id"] for c in ACCT_CFG if c["label"] == _dq_acct_label),
                          ACCT_CFG[0]["id"])

    _df_dq = _load_options_positions(_dq_acct_id)

    if _df_dq.empty:
        st.info("暂无期权持仓数据。请先在「🏦 期权持仓」Tab 录入持仓或导入 CSV。")
    else:
        # ── 构建审计行 ─────────────────────────────────────────
        _dq_rows = []
        for _, _ar in _df_dq.iterrows():
            _sym  = str(_ar.get("symbol")        or "").upper()
            _dir  = str(_ar.get("direction")     or "")
            _uc   = _ar.get("unit_cost")
            _cp   = _ar.get("current_price")
            _qty  = _ar.get("quantity")
            _mv   = _ar.get("market_value")
            _tp   = _ar.get("total_pnl")
            _strk = _ar.get("strike")
            _exp  = _ar.get("expiry")

            _qty_n = int(_qty) if _qty is not None else 0

            # 重新计算 total_pnl
            _calc_tp = (round((_cp - _uc) * _qty_n * 100, 2)
                        if (_cp is not None and _uc is not None) else None)

            # 异常标记
            _flags = []
            if _uc is not None and _uc > 100:
                _flags.append("⚠️ unit_cost>100（单位疑似错误）")
            if _uc is not None and _uc <= 0:
                _flags.append("⚠️ unit_cost≤0（成本缺失）")
            if _cp is None:
                _flags.append("⚠️ 无现价")
            if _calc_tp is not None and _tp is not None:
                _diff = abs(_calc_tp - float(_tp))
                if _diff > 1:
                    _flags.append(f"⚠️ pnl差 ${_diff:.0f}（DB:{_tp:.0f} vs 算:{_calc_tp:.0f}）")
            if _strk is None or _strk == 0:
                _flags.append("⚠️ strike 缺失")
            if not _exp:
                _flags.append("⚠️ expiry 缺失")

            _dq_rows.append({
                "代号":            _sym,
                "方向":            _dir,
                "数量":            _qty_n,
                "strike":          _strk,
                "expiry":          _exp,
                "unit_cost(DB)":   _uc,
                "现价(DB)":        _cp,
                "市值(DB)":        _mv,
                "pnl(DB)":         _tp,
                "pnl(算)":         _calc_tp,
                "异常标记":        "  ".join(_flags) if _flags else "✅",
            })

        _df_audit = pd.DataFrame(_dq_rows)

        # ── 汇总 ───────────────────────────────────────────────
        _n_total = len(_df_audit)
        _n_ok    = (_df_audit["异常标记"] == "✅").sum()
        _n_warn  = _n_total - _n_ok
        _ac1, _ac2, _ac3 = st.columns(3)
        _ac1.metric("合约总数", str(_n_total))
        _ac2.metric("正常 ✅",  str(_n_ok))
        _ac3.metric("有异常 ⚠️", str(_n_warn),
                    delta=f"-{_n_warn}" if _n_warn else None,
                    delta_color="inverse" if _n_warn else "off")

        # ── 异常过滤 ───────────────────────────────────────────
        _show_all = st.toggle("显示全部（含正常）", value=False, key="dq_show_all")
        _df_show  = _df_audit if _show_all else _df_audit[_df_audit["异常标记"] != "✅"]
        if _df_show.empty:
            st.success("✅ 所有持仓数据通过核查，无异常")
        else:
            st.dataframe(
                _df_show,
                use_container_width=True,
                hide_index=True,
                height=min(60 + len(_df_show) * 38, 600),
                column_config={
                    "unit_cost(DB)":  st.column_config.NumberColumn("unit_cost",  format="%.4f"),
                    "现价(DB)":       st.column_config.NumberColumn("现价",       format="%.4f"),
                    "市值(DB)":       st.column_config.NumberColumn("市值",       format="$%.2f"),
                    "pnl(DB)":        st.column_config.NumberColumn("pnl(DB)",   format="$%.2f"),
                    "pnl(算)":        st.column_config.NumberColumn("pnl(算)",   format="$%.2f"),
                    "strike":         st.column_config.NumberColumn("strike",     format="%.2f"),
                    "异常标记":       st.column_config.TextColumn("异常标记",    width="large"),
                },
            )

        st.caption(
            "异常类型：unit_cost>100 → 成本单位可能是总价而非每股价；"
            "pnl差 → DB 存的 total_pnl 与 (现价-成本)×数量×100 不一致；"
            "strike/expiry 缺失 → 影响 Greeks 计算。"
            "修正方式：在「🏦 期权持仓」Tab 编辑器中手动更正 unit_cost 后保存。"
        )

# ── 股票持仓 Tabs ──────────────────────────────────────────────
for _ptab, _pcfg in zip(_pos_tabs[6:], ACCT_CFG):
    with _ptab:
        _df_all   = _load_positions(_pcfg["id"])
        _stocks   = (_df_all[_df_all["position_type"] == "stock"].copy()
                     if not _df_all.empty else pd.DataFrame())

        # ── 汇总指标 ────────────────────────────────────────────────
        if not _stocks.empty:
            _smv  = pd.to_numeric(_stocks["market_value"],   errors="coerce").fillna(0).sum()
            _spnl = pd.to_numeric(_stocks["unrealized_pnl"], errors="coerce").fillna(0).sum()
            _sm1, _sm2, _sm3, _sm4 = st.columns(4)
            _sm1.metric("股票总市值", f"${_smv:,.0f}")
            _sm2.metric("总浮动盈亏", f"${_spnl:,.2f}", delta=f"{_spnl:+,.2f}")
            _sm3.metric("持仓只数",   f"{len(_stocks)}")
            _updated_at = _stocks["sync_time"].max() if "sync_time" in _stocks.columns else None
            if _updated_at:
                try:
                    _dt = datetime.datetime.fromisoformat(str(_updated_at))
                    _sm4.metric("数据时间", _dt.strftime("%m/%d %H:%M"))
                except Exception:
                    pass
            st.markdown("")

        # 股票行情更新请使用左侧边栏「📈 更新行情」

        # ── 持仓只读展示 ─────────────────────────────────────────
        if not _stocks.empty:
            _sd = _stocks.copy()
            # 补算 unit_cost（若列为空则从 cost_basis / qty 推算）
            _sd["unit_cost"] = pd.to_numeric(_sd.get("unit_cost"), errors="coerce")
            _mask_uc = _sd["unit_cost"].isna()
            if _mask_uc.any():
                _sd.loc[_mask_uc, "unit_cost"] = (
                    pd.to_numeric(_sd.loc[_mask_uc, "cost_basis"], errors="coerce") /
                    pd.to_numeric(_sd.loc[_mask_uc, "quantity"],   errors="coerce")
                ).round(2)
            _show_cols = ["symbol","description","quantity","unit_cost",
                          "current_price","market_value","unrealized_pnl","unrealized_pnl_pct"]
            _avail = [c for c in _show_cols if c in _sd.columns]
            st.dataframe(
                _sd[_avail].rename(columns={
                    "symbol":"代码","description":"名称","quantity":"数量",
                    "unit_cost":"单价成本","current_price":"现价",
                    "market_value":"市值","unrealized_pnl":"浮盈亏",
                    "unrealized_pnl_pct":"盈亏%",
                }),
                use_container_width=True, hide_index=True,
                height=min(45 + len(_sd) * 36, 420),
                column_config={
                    "单价成本": st.column_config.NumberColumn(format="$%.2f"),
                    "现价":     st.column_config.NumberColumn(format="$%.2f"),
                    "市值":     st.column_config.NumberColumn(format="$%.2f"),
                    "浮盈亏":   st.column_config.NumberColumn(format="$%.2f"),
                    "盈亏%":    st.column_config.NumberColumn(format="%.2f%%"),
                })
            st.markdown("")

        # ── 手动录入 / 编辑持仓 ──────────────────────────────────
        with st.expander("✏️ 手动录入 / 编辑持仓", expanded=_stocks.empty):
            st.caption("数量/单价成本为权威数据，现价可留空（由刷新按钮自动填入）。"
                       "直接粘贴 Excel 数据或逐行填写，完成后点「💾 保存」。")

            # 准备编辑用 DataFrame
            if not _stocks.empty:
                _edit_src = _stocks.reindex(
                    columns=["symbol","description","quantity","unit_cost","current_price"]
                ).copy()
                _edit_src["unit_cost"] = pd.to_numeric(_edit_src.get("unit_cost"), errors="coerce")
                _mask_uc2 = _edit_src["unit_cost"].isna()
                if _mask_uc2.any():
                    _edit_src.loc[_mask_uc2, "unit_cost"] = (
                        pd.to_numeric(_stocks.loc[_mask_uc2, "cost_basis"], errors="coerce") /
                        pd.to_numeric(_stocks.loc[_mask_uc2, "quantity"],   errors="coerce")
                    ).round(2)
            else:
                _edit_src = pd.DataFrame({
                    "symbol":        ["", "", "", "", ""],
                    "description":   ["", "", "", "", ""],
                    "quantity":      [0,  0,  0,  0,  0 ],
                    "unit_cost":     [0.0,0.0,0.0,0.0,0.0],
                    "current_price": [None,None,None,None,None],
                })

            _edited_stocks = st.data_editor(
                _edit_src,
                num_rows="dynamic",
                use_container_width=True,
                key=f"stock_editor_{_pcfg['id']}",
                column_config={
                    "symbol":        st.column_config.TextColumn("代号", help="如 ARM、AVGO", width="small"),
                    "description":   st.column_config.TextColumn("名称/备注", width="medium"),
                    "quantity":      st.column_config.NumberColumn("数量", min_value=0, step=1),
                    "unit_cost":     st.column_config.NumberColumn("单价成本", format="$%.4f",
                                                                   help="每股买入成本"),
                    "current_price": st.column_config.NumberColumn("现价（可选）", format="$%.4f"),
                },
            )

            _sc1, _sc2 = st.columns([1, 3])
            with _sc1:
                if st.button("💾 保存持仓", key=f"save_stk_{_pcfg['id']}", type="primary"):
                    _save_rows = [
                        r for r in _edited_stocks.to_dict("records")
                        if str(r.get("symbol", "")).strip()
                    ]
                    if _save_rows:
                        _n = _save_stock_positions(_pcfg["id"], _save_rows)
                        st.success(f"已保存 {_n} 只股票持仓")
                        st.rerun()
                    else:
                        st.warning("没有有效数据（代号不能为空）")
            with _sc2:
                # 可选：从 Firstrade xlsx 批量导入
                _dl_dir  = pathlib.Path.home() / "Downloads"
                _xl_list = sorted(_dl_dir.glob("*positions*.xlsx"),
                                  key=lambda f: f.stat().st_mtime, reverse=True)
                if _xl_list:
                    if st.button("📂 从 xlsx 批量导入",
                                 key=f"bulk_import_{_pcfg['id']}",
                                 help=f"从 ~/Downloads/{_xl_list[0].name} 导入，覆盖现有数据"):
                        _sn, _on = _import_from_xlsx_file(_xl_list[0], _pcfg["id"])
                        st.success(f"已导入 {_sn} 只股票 + {_on} 张期权")
                        st.rerun()

# 页脚
st.divider()
_footer_col1, _footer_col2 = st.columns([3, 1])
with _footer_col1:
    st.markdown(
        f"<div style='color:{_MUTED};font-size:11px;padding:4px 0'>"
        f"ENERGREX · watchdog CSV 自动导入 · SQLite data/energrex.db · "
        f"{datetime.datetime.now(_ET).strftime('%Y-%m-%d %H:%M ET')}</div>",
        unsafe_allow_html=True)
with _footer_col2:
    if st.button("🔄 刷新页面", key="manual_refresh", use_container_width=True):
        st.rerun()

# ── watchdog 新文件通知（不再 sleep，只 toast 提示）──────────
_ws = _watch_state()
if _ws.get("new_data"):
    _ws["new_data"] = False
    st.toast(f"✅ 检测到新文件：{_ws.get('last_file','')}（{_ws.get('last_rows',0)} 行）", icon="📂")
